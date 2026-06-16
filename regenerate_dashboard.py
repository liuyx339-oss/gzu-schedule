"""
Standalone dashboard regenerator — reads existing Schedule Excel, generates HTML.
Run: python regenerate_dashboard.py
"""
import json
import os
import re
from collections import defaultdict
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, "Schedule_2026-06.xlsx")
HTML_OUT = os.path.join(BASE_DIR, "Schedule_Dashboard_2026-06.html")

TARGET_HOURS_FULL = 176.0
TARGET_HOURS_80 = 140.8

DISPLAY_NAME_MAP = {
    "Dustin Huang": "Dustin Huang",
    "Dustin Huang (US)": "Dustin Huang (US)",
}

SHIFT_TIME_STR = {
    "D": "08:30-17:30", "D1": "08:30-17:00", "D2": "09:00-17:30", "D3": "09:30-18:00",
    "D4": "09:00-18:00", "D5": "08:30-18:00", "D6": "07:30-15:30",
    "C": "07:40-16:10", "C1": "08:00-16:30", "L": "08:00-20:00",
    "H1": "07:40-11:40", "H2": "08:30-12:30", "H3": "13:30-17:30", "T": "08:00-12:00",
    "N": "17:30-08:00", "N2": "17:30-07:30", "N3": "18:00-08:00",
    "L/N": "08:00-08:00", "OnCall": "OnCall",
}

SHIFT_COLORS = {
    "D": "#4CAF50", "D1": "#66BB6A", "D2": "#81C784", "D3": "#A5D6A7",
    "D4": "#43A047", "D5": "#388E3C", "D6": "#2E7D32",
    "C": "#4CAF50", "C1": "#66BB6A", "L": "#8BC34A",
    "H1": "#29B6F6", "H2": "#4FC3F7", "H3": "#81D4FA", "T": "#B3E5FC",
    "N": "#1a73e8", "N2": "#1565C0", "N3": "#0D47A1",
    "L/N": "#FF9800", "OnCall": "#9E9E9E", "off": "#F5F5F5",
}

# Role-key to sheet-name mapping
ROLE_SHEETS = {
    "放射医生": ("放射医生排班", "放射_放射医生", True),
    "放射技师": ("放射技师排班", "放射_放射技师", False),
    "B超医生": ("超声医生排班", "超声_B超医生", False),
}


def read_schedule_sheet(wb, sheet_name, is_rad_doc=False):
    """Read a schedule sheet from Excel, return staff list and dates"""
    ws = wb[sheet_name]
    headers = [str(c.value or "") for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]

    # Find date columns
    dates = []
    date_cols = []
    for i, h in enumerate(headers):
        if "月" in h and "日" in h:
            date_cols.append(i)
            dates.append(h)

    staff_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or "").strip()
        if not name or "人员" in name:
            continue
        display = DISPLAY_NAME_MAP.get(name, name)
        hours = float(row[1]) if row[1] is not None else 0.0
        target = float(row[2]) if row[2] is not None else 0.0
        is_backup = "备班" in name
        ln_count = int(row[3]) if is_rad_doc and row[3] is not None else 0

        schedule = {}
        for di, ci in enumerate(date_cols):
            val = str(row[ci] or "").strip()
            ds = dates[di]
            schedule[ds] = val

        staff_list.append({
            "name": display,
            "internal_name": name,
            "hours": round(hours, 1),
            "target": target,
            "is_backup": is_backup,
            "ln_count": ln_count,
            "schedule": schedule,
        })

    return {"staff": staff_list, "dates": dates}


def main():
    print("Reading Excel:", XLSX_PATH)
    wb = openpyxl.load_workbook(XLSX_PATH)
    print("Available sheets:", wb.sheetnames)

    roles_data = {}
    for role_name, (sheet_name, role_key, is_rad_doc) in ROLE_SHEETS.items():
        if sheet_name in wb.sheetnames:
            roles_data[role_name] = read_schedule_sheet(wb, sheet_name, is_rad_doc)
            roles_data[role_name]["shift_colors"] = SHIFT_COLORS
            roles_data[role_name]["shift_times"] = SHIFT_TIME_STR
            # Demand is not in the dashboard data, but we keep empty samples
            roles_data[role_name]["demand_samples"] = {}
            print(f"  {role_name}: {len(roles_data[role_name]['staff'])} staff, "
                  f"{len(roles_data[role_name]['dates'])} dates")

    # Compute statistics
    all_hours = []
    backup_hours = 0
    backup_count = 0
    ft_count = 0
    for role_name, role_data in roles_data.items():
        for p in role_data["staff"]:
            if p["is_backup"]:
                if p["hours"] > 0:
                    backup_hours += p["hours"]
                    backup_count += 1
            else:
                if p["hours"] > 0:
                    all_hours.append(p["hours"])
                    ft_count += 1

    avg_hours = round(sum(all_hours) / len(all_hours), 1) if all_hours else 0
    all_dates = list(roles_data.values())[0]["dates"] if roles_data else []
    month_str = f"{all_dates[0]} ~ {all_dates[-1]}" if all_dates else "N/A"

    data = {
        "month": month_str,
        "roles": roles_data,
        "statistics": {
            "total_fulltime": ft_count,
            "total_backup": backup_count,
            "total_backup_hours": round(backup_hours, 1),
            "avg_fulltime_hours": avg_hours,
        },
    }

    # Render HTML using safe replacement
    json_data = json.dumps(data, ensure_ascii=False, default=str)

    # Read template from this script's string
    html = _get_html_template()
    replacements = [
        ("__MONTH__", month_str),
        ("__JSON_DATA__", json_data),
        ("__TARGET_FULL__", str(TARGET_HOURS_FULL)),
        ("__TARGET_80__", str(TARGET_HOURS_80)),
        ("__STAT_FT__", str(ft_count)),
        ("__STAT_AVG__", str(avg_hours)),
        ("__STAT_BU_H__", str(round(backup_hours, 1))),
        ("__STAT_BU_N__", str(backup_count)),
    ]
    for placeholder, value in replacements:
        html = html.replace(placeholder, value)

    with open(HTML_OUT, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\n✅ Dashboard regenerated: {HTML_OUT}")
    print(f"   Size: {len(html)} bytes")


def _get_html_template():
    return r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>排班仪表盘</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:"Microsoft YaHei","SimHei",sans-serif; background:#f0f2f5; color:#333; min-height:100vh; }
.header { background:linear-gradient(135deg,#1a73e8 0%,#0d47a1 100%); color:#fff; padding:20px 32px; text-align:center; }
.header h1 { font-size:24px; font-weight:600; margin:0; }
.header .subtitle { font-size:13px; opacity:0.85; margin-top:4px; }
.stats-bar { display:flex; gap:16px; padding:16px 20px; max-width:1400px; margin:0 auto; flex-wrap:wrap; justify-content:center; }
.stat-card { background:#fff; border-radius:8px; padding:16px 24px; box-shadow:0 1px 4px rgba(0,0,0,0.08); text-align:center; min-width:140px; }
.stat-card .value { font-size:28px; font-weight:700; color:#1a73e8; }
.stat-card .label { font-size:12px; color:#666; margin-top:4px; }
.container { max-width:1400px; margin:0 auto; padding:0 20px 20px; }
.tabs { display:flex; gap:4px; margin-bottom:16px; background:#fff; border-radius:8px; padding:4px; box-shadow:0 1px 4px rgba(0,0,0,0.08); }
.tab { flex:1; text-align:center; padding:10px 20px; cursor:pointer; border-radius:6px; font-size:14px; font-weight:600; transition:all 0.2s; background:#fff; border:none; color:#333; }
.tab:hover { background:#e8f0fe; }
.tab.active { background:#1a73e8; color:#fff; }
.legend { display:flex; gap:12px; flex-wrap:wrap; padding:12px 16px; background:#fff; border-radius:8px; margin-bottom:12px; box-shadow:0 1px 4px rgba(0,0,0,0.08); font-size:12px; align-items:center; }
.legend-item { display:inline-flex; align-items:center; gap:4px; white-space:nowrap; }
.legend-dot { width:14px; height:14px; border-radius:3px; display:inline-block; flex-shrink:0; }
.roster-grid { background:#fff; border-radius:8px; padding:16px; box-shadow:0 1px 4px rgba(0,0,0,0.08); }
.roster-grid h2 { font-size:16px; margin:0 0 12px 0; color:#1a73e8; }
.roster-wrapper { overflow-x:auto; }
table.schedule { border-collapse:collapse; width:max-content; min-width:100%; font-size:12px; }
table.schedule th, table.schedule td { border:1px solid #e0e0e0; padding:6px 5px; text-align:center; white-space:nowrap; }
table.schedule thead th { background:#f5f5f5; font-weight:600; position:sticky; top:0; z-index:2; }
table.schedule .name-col { min-width:90px; position:sticky; left:0; background:#fff; z-index:1; font-weight:600; text-align:left; padding-left:8px; }
table.schedule tbody tr:hover .name-col { background:#f0f7ff; }
table.schedule .stats-col { min-width:55px; }
table.schedule .shift-cell { font-size:11px; min-width:56px; cursor:pointer; transition:all 0.15s; border-radius:3px; position:relative; }
table.schedule .shift-cell:hover { transform:scale(1.08); z-index:3; box-shadow:0 2px 8px rgba(0,0,0,0.2); }
table.schedule .cell-backup { border:2px dashed #F44336 !important; }
table.schedule .cell-ln { border:3px solid #FF9800 !important; font-weight:bold; }
table.schedule .cell-oncall::after { content:"📞"; position:absolute; top:0px; right:1px; font-size:8px; line-height:1; }
.popup { display:none; position:fixed; top:50%; left:50%; transform:translate(-50%,-50%); background:#fff; border-radius:12px; padding:24px; box-shadow:0 8px 32px rgba(0,0,0,0.2); z-index:1000; max-width:650px; width:90%; max-height:80vh; overflow-y:auto; }
.popup.active { display:block; }
.popup h3 { margin:0 0 12px 0; color:#1a73e8; }
.popup table { width:100%; font-size:13px; border-collapse:collapse; margin-bottom:12px; }
.popup table td, .popup table th { padding:6px 8px; border-bottom:1px solid #eee; text-align:left; }
.overlay { display:none; position:fixed; inset:0; background:rgba(0,0,0,0.4); z-index:999; }
.overlay.active { display:block; }
</style>
</head>
<body>
<div class="header">
    <h1>🏥 放射/超声 月度排班仪表盘</h1>
    <div class="subtitle">__MONTH__ | 全职目标: __TARGET_FULL__h | 80%阈值: __TARGET_80__h</div>
</div>

<div class="stats-bar">
    <div class="stat-card"><div class="value">__STAT_FT__</div><div class="label">全职人员</div></div>
    <div class="stat-card"><div class="value">__STAT_AVG__h</div><div class="label">全职人均工时</div></div>
    <div class="stat-card"><div class="value">__STAT_BU_H__h</div><div class="label">备班总工时</div></div>
    <div class="stat-card"><div class="value">__STAT_BU_N__</div><div class="label">备班人数</div></div>
</div>

<div class="container">
    <div class="tabs" id="tabs"></div>
    <div class="legend" id="legend"></div>
    <div class="roster-grid">
        <div id="roster"><div style="text-align:center;padding:40px;color:#999;">加载中...</div></div>
    </div>
</div>

<div class="overlay" id="overlay" onclick="closePopup()"></div>
<div class="popup" id="popup"></div>

<script>
var SCHEDULE_DATA = __JSON_DATA__;
var TARGET_HOURS = __TARGET_FULL__;
var TARGET_80 = __TARGET_80__;
var currentRole = '放射医生';

function init() {
    var tabs = document.getElementById('tabs');
    var roles = Object.keys(SCHEDULE_DATA.roles);
    for (var i = 0; i < roles.length; i++) {
        (function(role) {
            var btn = document.createElement('button');
            btn.className = 'tab' + (role === currentRole ? ' active' : '');
            btn.textContent = role;
            btn.onclick = function() { currentRole = role; renderAll(); };
            tabs.appendChild(btn);
        })(roles[i]);
    }
    renderAll();
}

function renderAll() {
    var tabEls = document.querySelectorAll('.tab');
    for (var i = 0; i < tabEls.length; i++) {
        var t = tabEls[i];
        if (t.textContent === currentRole) { t.classList.add('active'); }
        else { t.classList.remove('active'); }
    }
    renderLegend();
    renderRoster();
}

function renderLegend() {
    var roleData = SCHEDULE_DATA.roles[currentRole];
    var colors = roleData.shift_colors;
    var legend = document.getElementById('legend');
    var html = '<b>班次图例:</b> ';
    var shown = {};
    var keys = Object.keys(colors);
    for (var i = 0; i < keys.length; i++) {
        var shift = keys[i];
        if (shown[shift] || shift === 'off') continue;
        shown[shift] = true;
        var color = colors[shift];
        var time = roleData.shift_times[shift] || '';
        html += '<span class="legend-item"><span class="legend-dot" style="background:' + color + '"></span> ' + shift + (time ? ' (' + time + ')' : '') + '</span> ';
    }
    html += '<span class="legend-item"><span class="legend-dot" style="background:#F5F5F5;border:1px solid #ddd"></span> 休息</span> ';
    html += '<span class="legend-item" style="color:#1a73e8;font-weight:600">🔵 超80%</span> ';
    html += '<span class="legend-item" style="color:#F44336;font-weight:600">🔴 备班</span>';
    legend.innerHTML = html;
}

function renderRoster() {
    var roleData = SCHEDULE_DATA.roles[currentRole];
    var roster = document.getElementById('roster');
    var dates = roleData.dates;

    var html = '<h2>' + currentRole + ' 排班表</h2>';
    html += '<div class="roster-wrapper"><table class="schedule"><thead><tr>';
    html += '<th class="name-col">人员</th><th class="stats-col">工时</th><th class="stats-col">剩余</th>';
    if (currentRole === '放射医生') html += '<th class="stats-col">L/N</th>';
    for (var i = 0; i < dates.length; i++) {
        html += '<th>' + dates[i].slice(3) + '</th>';
    }
    html += '</tr></thead><tbody>';

    var staffList = roleData.staff;
    for (var si = 0; si < staffList.length; si++) {
        var person = staffList[si];
        html += '<tr>';
        html += '<td class="name-col">' + person.name + (person.is_backup ? ' 🔄' : '') + '</td>';

        var over80 = person.hours > TARGET_80 && !person.is_backup;
        html += '<td class="stats-col" style="' + (over80 ? 'color:#1a73e8;font-weight:600' : '') + '">' + person.hours + 'h</td>';

        var remain = person.target > 0 ? (person.target - person.hours).toFixed(1) : '-';
        var remainStyle = (person.target > 0 && parseFloat(remain) < 0) ? 'color:red' : '';
        html += '<td class="stats-col" style="' + remainStyle + '">' + remain + 'h</td>';

        if (currentRole === '放射医生') html += '<td class="stats-col">' + (person.ln_count || '') + '</td>';

        for (var di = 0; di < dates.length; di++) {
            var ds = dates[di];
            var shiftVal = person.schedule[ds] || '';
            var shifts = shiftVal ? shiftVal.split(' + ').filter(function(s) { return s; }) : [];
            var bg = '#FAFAFA';
            var text = '-';
            var extraClass = '';

            if (shifts.length > 0 && shiftVal !== 'None' && shiftVal !== 'null') {
                text = shiftVal;
                bg = '#e0e0e0';
                var mainShift = null;
                for (var k = 0; k < shifts.length; k++) {
                    if (shifts[k] !== 'OnCall') { mainShift = shifts[k]; break; }
                }
                if (!mainShift) mainShift = shifts[0];
                var cleanShift = mainShift.replace('OnCall', '').trim();
                if (roleData.shift_colors[cleanShift]) bg = roleData.shift_colors[cleanShift];

                if (person.is_backup && cleanShift) extraClass += ' cell-backup';
                for (var k2 = 0; k2 < shifts.length; k2++) {
                    if (shifts[k2].indexOf('L/N') >= 0) { extraClass += ' cell-ln'; break; }
                }
                for (var k3 = 0; k3 < shifts.length; k3++) {
                    if (shifts[k3].indexOf('OnCall') >= 0) { extraClass += ' cell-oncall'; break; }
                }
                if (over80) text = '<span style="color:#1a73e8;font-weight:600">' + text + '</span>';
                if (person.is_backup) text = '<span style="color:#F44336;font-weight:600">' + text + '</span>';
            }

            html += '<td class="shift-cell' + extraClass + '" style="background:' + bg + '" onclick="showDayDetail(\'' + ds + '\')" title="' + ds + ': ' + (shiftVal || '休息') + '">' + text + '</td>';
        }
        html += '</tr>';
    }
    html += '</tbody></table></div>';
    roster.innerHTML = html;
}

function showDayDetail(dateStr) {
    var popup = document.getElementById('popup');
    var overlay = document.getElementById('overlay');
    var roleData = SCHEDULE_DATA.roles[currentRole];

    var html = '<h3>📅 ' + dateStr + ' — ' + currentRole + '</h3>';
    html += '<table><thead><tr><th>人员</th><th>班次</th><th>时间</th></tr></thead><tbody>';
    var staffList = roleData.staff;
    var hasData = false;
    for (var si = 0; si < staffList.length; si++) {
        var person = staffList[si];
        var shiftVal = person.schedule[dateStr] || '';
        if (shiftVal && shiftVal !== 'None' && shiftVal !== 'null') {
            hasData = true;
            var parts = shiftVal.split(' + ');
            var timeParts = [];
            for (var k = 0; k < parts.length; k++) {
                timeParts.push(roleData.shift_times[parts[k].trim()] || parts[k].trim());
            }
            var times = timeParts.join(' + ');
            html += '<tr><td>' + person.name + (person.is_backup ? ' 🔄' : '') + '</td><td>' + shiftVal + '</td><td>' + times + '</td></tr>';
        }
    }
    if (!hasData) html += '<tr><td colspan="3" style="text-align:center;color:#999">当日无人排班</td></tr>';
    html += '</tbody></table>';

    popup.innerHTML = html;
    popup.classList.add('active');
    overlay.classList.add('active');
}

function closePopup() {
    var popup = document.getElementById('popup');
    var overlay = document.getElementById('overlay');
    if (popup) popup.classList.remove('active');
    if (overlay) overlay.classList.remove('active');
}

document.addEventListener('keydown', function(e) {
    if (e.key === 'Escape') closePopup();
});

// Start
if (document.readyState === 'loading') {
    document.addEventListener('DOMContentLoaded', init);
} else {
    init();
}
</script>
</body>
</html>"""


if __name__ == "__main__":
    main()
