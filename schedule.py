# ==========================================
# 排班优化系统 — V3
# Pipeline: 预测工作量 → HC需求转换 → L/N预分配 →
#           Stage1(80%工时CP-SAT) → Stage2(20%工时CP-SAT) →
#           Stage3(备班CP-SAT全覆盖) → OnCall分配 →
#           Dustin跨角色处理 → Excel输出(分类标注) → Web可视化
#
# 核心原则:
#   1. 统一分层覆盖: 80%池 → 20%池 → 备班池 (所有需求一致)
#   2. 硬性全覆盖: 80% + 20% + 备班 必须覆盖所有需求(Stage3无slack)
#   3. 放射技师备班仅按小时(无班型)
#   4. Dustin: 放射优先(一三五优先) → 剩余支持超声 → 抵扣0.5HC
#   5. 超声每日总人数 ≤ 5
# ==========================================
import os
import sys
import math
import json
import argparse
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime, timedelta

sys.stdout.reconfigure(line_buffering=True, encoding='utf-8')

# ==========================================
# 1. 配置 & 常量
# ==========================================

TOLERANCE_MINUTES = 25

# --- 负荷率 (每小时有效工作分钟 = 60 * load_rate) ---
LOAD_RATE = {
    "放射医生": {0:0.930988,1:0.937606,2:0.933171,3:1.0,4:1.0,5:0.955709,6:0.868295,7:0.868396,8:0.776812,9:0.728898,10:0.708848,11:0.711359,12:0.70014,13:0.741611,14:0.756309,15:0.766785,16:0.791551,17:0.808238,18:0.85161,19:0.870316,20:0.857747,21:0.891201,22:0.891172,23:0.926642},
    "放射技师": {0:0.894781,1:0.916714,2:0.93822,3:0.903828,4:0.856494,5:0.930688,6:0.849995,7:0.92332,8:0.86096,9:0.801839,10:0.608063,11:0.743238,12:0.721842,13:0.728564,14:0.727277,15:0.738215,16:0.738477,17:0.794821,18:0.87881,19:0.874276,20:0.858698,21:0.87809,22:0.883282,23:0.911768},
    "B超医生": {0:0.939,1:0.88374,2:0.8804,3:0.939,4:0.939,5:0.939,6:0.939,7:0.9924,8:0.9488,9:0.9692,10:0.8692,11:0.8536,12:0.74042,13:0.72396,14:0.61706,15:0.53536,16:0.54538,17:0.6105,18:0.75438,19:0.86406,20:0.84002,21:0.87574,22:0.87734,23:0.8828},
}

# --- 班次定义 ---
SHIFT_DICT = {
    "D":  (8.5, 17.5, 8.5, "白班"),   "D1": (8.5, 17.0, 8.0, "白班"),
    "D2": (9.0, 17.5, 8.0, "白班"),   "D3": (9.5, 18.0, 8.0, "白班"),
    "D4": (9.0, 18.0, 8.5, "白班"),   "D5": (8.5, 18.0, 9.0, "白班"),
    "D6": (7.5, 15.5, 8.0, "白班"),   "C":  (7.67, 16.17, 8.0, "白班"),
    "C1": (8.0, 16.5, 8.0, "白班"),   "L":  (8.0, 20.0, 12.0, "白班"),
    "H1": (7.67, 11.67, 4.0, "半天班"), "H2": (8.5, 12.5, 4.0, "半天班"),
    "H3": (13.5, 17.5, 4.0, "半天班"),  "T":  (8.0, 12.0, 4.0, "半天班"),
    "N":  (17.5, 32.0, 14.5, "夜班"),  "N2": (17.5, 31.5, 14.0, "夜班"),
    "N3": (18.0, 32.0, 14.0, "夜班"),
    "L/N": (8.0, 32.0, 24.0, "24H班"),
}

SHIFT_TIME_STR = {
    "D": "08:30-17:30", "D1": "08:30-17:00", "D2": "09:00-17:30", "D3": "09:30-18:00",
    "D4": "09:00-18:00", "D5": "08:30-18:00", "D6": "07:30-15:30",
    "C": "07:40-16:10", "C1": "08:00-16:30", "L": "08:00-20:00",
    "H1": "07:40-11:40", "H2": "08:30-12:30", "H3": "13:30-17:30", "T": "08:00-12:00",
    "N": "17:30-08:00", "N2": "17:30-07:30", "N3": "18:00-08:00",
    "L/N": "08:00-08:00",
}

SHIFT_COVERAGE = {
    "D":  [0,0,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,1,1, 0,0,0,0,0,0],
    "D1": [0,0,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,1,0, 0,0,0,0,0,0],
    "D2": [0,0,0,0,0,0,0,0, 0,1,1,1,1,1,1,1,1,1, 0,0,0,0,0,0],
    "D3": [0,0,0,0,0,0,0,0, 0,1,1,1,1,1,1,1,1,1, 0,0,0,0,0,0],
    "D4": [0,0,0,0,0,0,0,0, 0,1,1,1,1,1,1,1,1,1, 0,0,0,0,0,0],
    "D5": [0,0,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,1,1, 0,0,0,0,0,0],
    "D6": [0,0,0,0,0,0,0,1, 1,1,1,1,1,1,1,1,0,0, 0,0,0,0,0,0],
    "C":  [0,0,0,0,0,0,0,1, 1,1,1,1,1,1,1,1,1,0, 0,0,0,0,0,0],
    "C1": [0,0,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,1,0, 0,0,0,0,0,0],
    "L":  [0,0,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,1,1,1,1,0,0,0,0],
    "H1": [0,0,0,0,0,0,0,1, 1,1,1,1, 0,0,0,0,0,0,0,0,0,0,0,0],
    "H2": [0,0,0,0,0,0,0,0, 1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,0],
    "H3": [0,0,0,0,0,0,0,0, 0,0,0,0,0,1,1,1,1,1, 0,0,0,0,0,0],
    "T":  [0,0,0,0,0,0,0,0, 1,1,1,1, 0,0,0,0,0,0,0,0,0,0,0,0],
    "N":  [1,1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,1,1,1,1,1,1,1],
    "N2": [1,1,1,1,1,1,1,0, 0,0,0,0,0,0,0,0,0,1,1,1,1,1,1,1],
    "N3": [1,1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,1,1,1,1,1,1],
    "L/N":[1,1,1,1,1,1,1,1, 1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1],
}

ROLE_SHIFTS = {
    "放射医生": ["D","D1","D2","D3","D4","D5","D6","C","C1","L","L/N","N","N2","N3"],
    "放射技师": ["D","D1","D2","D3","D4","D5","D6","C","C1","L","H1","H2","H3","N","N2","N3","L/N"],
    "B超医生": ["D","D1","D2","D3","D4","D5","D6","C","C1","H1","H2","H3"],
}

NIGHT_SHIFTS = {"N", "N2", "N3"}
DAY_SHIFTS = {"D","D1","D2","D3","D4","D5","D6","C","C1","L","H1","H2","H3","T"}
# 全天白班（不含半天班 H1/H2/H3/T）— 基础保障硬约束使用
FULL_DAY_SHIFTS = {"D","D1","D2","D3","D4","D5","D6","C","C1","L"}

TARGET_HOURS_FULL = None   # 在 main() 中根据月份动态计算: 当月工作日数 × 8
TARGET_HOURS_80 = None     # TARGET_HOURS_FULL × 0.8

# --- 角色配置 ---
ROLE_CONFIG = {
    "放射医生": {
        "day_shifts": 1, "night_shifts": 1, "ln_per_month": 2,
        "coverage_24h": False, "has_oncall": False,
        "backup_shift_based": True, "night_prefer_backup": True,
    },
    "放射技师": {
        "day_shifts": 2, "night_shifts": 1, "ln_per_month": 2,
        "coverage_24h": True, "has_oncall": True,
        "backup_shift_based": False, "night_prefer_backup": False,
    },
    "B超医生": {
        "day_shifts": 2, "night_shifts": 0, "ln_per_month": 0,
        "coverage_24h": False, "has_oncall": True,
        "backup_shift_based": True, "night_prefer_backup": False,
    },
}

# CP-SAT 阶段权重
S1_COVERAGE_WEIGHT = 1_000_000
S1_DAILY_REQ_WEIGHT = 500_000
S1_NIGHT_REQ_WEIGHT = 800_000
S1_DUSTIN_MWF_WEIGHT = 150_000  # Dustin 周三周五强激励
S1_BALANCE_WEIGHT = 50_000

S2_COVERAGE_WEIGHT = 1_000_000

S3_BACKUP_MINIMIZE = 1_000

# --- Staff fallback ---
STAFF_FALLBACK = {
    "rad_docs_full": ["li zhenhuan", "Dustin Huang"],
    "rad_docs_pt": ["Zhou ChunXiang", "Liang ZhiYing", "Ling Jian", "Liang Ruiyun",
                     "Liu Zengwei", "Chen Yingqian", "zhujun", "wangshuai"],
    "rad_techs_full": ["Zheng Xiaochun", "Zhang Meng", "Ma Linlin", "Yang Yongjun", "Yi Hong", "Liu Shuting"],
    "rad_techs_pt": ["ZHONG Minzhi", "LUO Hui", "CHEN Jiajun"],
    "us_docs_full": ["Xu Jing", "Liu Xiaoyan", "Lu Liyu", "doctor hou"],
    "us_docs_pt": ["Tang Shengwen", "Wu yingheng", "Yan Can", "Zhou Huiling",
                    "Liu Yuanyuan", "Wu yanling", "Wang Huimin"],
}
STAFF_FALLBACK_BACKUP = {
    "rad_docs": "放射医生备班",
    "rad_techs": "放射技师备班",
    "us_docs": "超声医生备班",
}

DUSTIN_RAD = "Dustin Huang"
DUSTIN_US = "Dustin Huang (US)"

DISPLAY_NAME = {
    "Dustin Huang": "Dustin Huang",
    "Dustin Huang (US)": "Dustin Huang (US)",
}

# ==========================================
# 2. 工时目标计算
# ==========================================

def _count_workdays(year, month):
    """计算指定月份的工作日数(周一至周五)。用于动态确定月目标工时。"""
    import calendar
    total = 0
    for day in range(1, calendar.monthrange(year, month)[1] + 1):
        wd = datetime(year, month, day).weekday()
        if wd < 5:  # 0=Monday, 4=Friday
            total += 1
    return total


# ==========================================
# 3. 人员加载
# ==========================================

def load_staff_from_feishu():
    """从飞书多维表格拉取人员。失败返回 None。"""
    app_id = os.environ.get("FEISHU_APP_ID", "cli_aaa8d24639b8dcd8")
    app_secret = os.environ.get("FEISHU_APP_SECRET", "b0ayVQKIuUGmvzRu9YCm9gpZHUzniNz1")
    try:
        import requests
        resp = requests.post(
            "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
            json={"app_id": app_id, "app_secret": app_secret}, timeout=15)
        token = resp.json().get("tenant_access_token", "")
        if not token:
            print("⚠️ 飞书token获取失败，使用fallback人员列表")
            return None
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        app_token, table_id = "MiRrw2dILig6I2k7wU7ceV0on9e", "tbl8f0tku6yPwc2V"
        all_records, page_token = [], None
        while True:
            url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records"
            params = {"page_size": 500}
            if page_token: params["page_token"] = page_token
            r = requests.get(url, headers=headers, params=params, timeout=30)
            data = r.json()
            if data.get("code") != 0:
                print(f"⚠️ 飞书API错误: {data.get('msg')}，使用fallback人员列表")
                return None
            items = data.get("data", {}).get("items", [])
            all_records.extend(items)
            if not data.get("data", {}).get("has_more"): break
            page_token = data.get("data", {}).get("page_token")

        staff = {"rad_docs_full": [], "rad_docs_pt": [], "rad_docs_backup": [],
                 "rad_techs_full": [], "rad_techs_pt": [], "rad_techs_backup": [],
                 "us_docs_full": [], "us_docs_pt": [], "us_docs_backup": []}
        for rec in all_records:
            fields = rec.get("fields", {})
            name, dept, position, emp_type = "", "", "", "全职"
            for key in fields:
                val = fields[key]
                if "姓名" in key:
                    if not name:
                        name = str(val[0]["text"]) if isinstance(val, list) else str(val)
                elif "科室" in key:
                    dept = str(val).strip() if not isinstance(val, list) else str(val[0] if val else "").strip()
                elif key.strip() == "类型":
                    position = str(val).strip() if not isinstance(val, list) else str(val[0] if val else "").strip()
                elif "雇佣形式" in key:
                    v = str(val).strip() if not isinstance(val, list) else str(val[0] if val else "").strip()
                    emp_type = v
                elif "角色" in key:  # 旧表兼容
                    role_str = str(val).strip() if not isinstance(val, list) else str(val[0] if val else "").strip()
                    if "超声" in role_str or "B超" in role_str:
                        dept, position = "超声", "医生"
                    elif "放射医生" in role_str:
                        dept, position = "放射", "医生"
                    elif "放射技师" in role_str:
                        dept, position = "放射", "技师"
            if not name:
                continue
            # 科室+职位 → staff key prefix
            if dept == "超声":
                prefix = "us_docs"
            elif dept == "放射" and position == "医生":
                prefix = "rad_docs"
            elif dept == "放射" and position == "技师":
                prefix = "rad_techs"
            else:
                continue
            # 雇佣形式 → suffix
            if emp_type == "兼职":
                key_suffix = "pt"
            elif emp_type == "备班":
                key_suffix = "backup"
            else:
                key_suffix = "full"
            staff_key = f"{prefix}_{key_suffix}"
            if staff_key not in staff:
                staff_key = f"{prefix}_full"  # fallback
            if name not in staff.get(staff_key, []):
                staff.setdefault(staff_key, []).append(name)
        total = sum(len(v) for v in staff.values())
        if total == 0: print("⚠️ 飞书返回空数据，使用fallback人员列表"); return None
        print(f"✅ 从飞书加载 {total} 名人员"); return staff
    except Exception as e:
        print(f"⚠️ 飞书API异常: {e}，使用fallback人员列表"); return None


def build_staff(staff_raw=None):
    if staff_raw is None: staff_raw = STAFF_FALLBACK
    staff = {
        "放射医生": {"fulltime": list(staff_raw.get("rad_docs_full", [])),
                      "parttime": list(staff_raw.get("rad_docs_pt", [])),
                      "backup": list(staff_raw.get("rad_docs_backup", [])) or [STAFF_FALLBACK_BACKUP["rad_docs"]]},
        "放射技师": {"fulltime": list(staff_raw.get("rad_techs_full", [])),
                      "parttime": list(staff_raw.get("rad_techs_pt", [])),
                      "backup": list(staff_raw.get("rad_techs_backup", [])) or [STAFF_FALLBACK_BACKUP["rad_techs"]]},
        "B超医生": {"fulltime": list(staff_raw.get("us_docs_full", [])),
                      "parttime": list(staff_raw.get("us_docs_pt", [])),
                      "backup": list(staff_raw.get("us_docs_backup", [])) or [STAFF_FALLBACK_BACKUP["us_docs"]]},
    }
    staff["Dustin_US"] = DUSTIN_US
    return staff


# ==========================================
# 3A. 请假数据加载
# ==========================================

LEAVE_BASE_TOKEN = "K9FdbW1mpaQE9gsWpX8c40Kvnub"
LEAVE_TABLE_ID = "tbluJrGR8bbn3RHM"


def _parse_time_range(time_str):
    """
    解析时间段字符串如 "8:30-5:30" 或 "8:30-17:30" 为 (start_hour_decimal, end_hour_decimal)。
    返回 (start_h, end_h) 或 None。
    """
    if not time_str or '-' not in time_str:
        return None
    try:
        parts = time_str.replace('：', ':').split('-')
        start_str, end_str = parts[0].strip(), parts[1].strip()

        def _to_decimal(s):
            s = s.strip()
            if ':' in s:
                h, m = s.split(':')
                return int(h) + int(m) / 60
            return float(s)

        start_h = _to_decimal(start_str)
        end_h = _to_decimal(end_str)
        return (start_h, end_h)
    except (ValueError, IndexError):
        return None


def _get_blocked_hours(start_h, end_h):
    """
    将开始/结束时间(decimal hours)转换为被阻塞的小时索引列表。
    例如 8:30-17:30 → 阻塞小时 8,9,10,11,12,13,14,15,16,17
    注意 17:30 结束意味着 17:00-18:00 这个小时也算被占用
    """
    blocked = set()
    h = int(start_h)  # 从开始小时开始
    while h < end_h:
        if 0 <= h < 24:
            blocked.add(h)
        h += 1
    return blocked


# 人员姓名映射: 飞书请假表姓名 → schedule.py 内部姓名
LEAVE_NAME_MAP = {
    "Dustin Huang": "Dustin Huang",
    "li zhenhuan": "li zhenhuan",
    "Zheng Xiaochun": "Zheng Xiaochun",
    "Zhang Meng": "Zhang Meng",
    "Ma Linlin": "Ma Linlin",
    "Yang Yongjun": "Yang Yongjun",
    "Yi Hong": "Yi Hong",
    "Liu Shuting": "Liu Shuting",
    "Xu Jing": "Xu Jing",
    "Liu Xiaoyan": "Liu Xiaoyan",
    "Lu Liyu": "Lu Liyu",
    "doctor hou": "doctor hou",
}


def load_leaves_from_feishu():
    """
    从飞书请假库 (K9FdbW1mpaQE9gsWpX8c40Kvnub) 加载请假记录。
    返回: {person_name: {date_str: set of blocked_hours}} 或 None (失败时)
    """
    app_id = os.environ.get("FEISHU_APP_ID", "cli_aaa8d24639b8dcd8")
    app_secret = os.environ.get("FEISHU_APP_SECRET", "b0ayVQKIuUGmvzRu9YCm9gpZHUzniNz1")
    try:
        import requests
        resp = requests.post(
            "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
            json={"app_id": app_id, "app_secret": app_secret}, timeout=15)
        token = resp.json().get("tenant_access_token", "")
        if not token:
            print("⚠️ 请假库token获取失败")
            return None
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

        all_records, page_token = [], None
        while True:
            url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{LEAVE_BASE_TOKEN}/tables/{LEAVE_TABLE_ID}/records"
            params = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            r = requests.get(url, headers=headers, params=params, timeout=30)
            data = r.json()
            if data.get("code") != 0:
                print(f"⚠️ 请假库API错误: {data.get('msg')}")
                return None
            items = data.get("data", {}).get("items", [])
            all_records.extend(items)
            if not data.get("data", {}).get("has_more"):
                break
            page_token = data.get("data", {}).get("page_token")

        leaves = defaultdict(dict)  # {person: {date_str: set of blocked_hours}}
        parsed_count = 0
        for rec in all_records:
            fields = rec.get("fields", {})
            name_raw = ""
            date_ts = None
            time_range_str = ""

            for key in fields:
                val = fields[key]
                kl = key.lower()
                if "姓名" in key:
                    name_raw = str(val[0]["text"]) if isinstance(val, list) else str(val)
                elif "pto" in kl or "时间" in key or "日期" in key:
                    # 时间戳 (毫秒)
                    if isinstance(val, (int, float)):
                        date_ts = int(val)
                elif "时间段" in key or "时段" in key:
                    time_range_str = str(val) if not isinstance(val, list) else str(val[0] if val else "")

            if not name_raw or not date_ts or not time_range_str:
                continue

            # 映射姓名
            person = LEAVE_NAME_MAP.get(name_raw, name_raw)

            # 解析日期 (毫秒时间戳 → 日期字符串)
            from datetime import datetime, timezone, timedelta
            dt = datetime.fromtimestamp(date_ts / 1000, tz=timezone(timedelta(hours=8)))
            date_str = dt.strftime("%m月%d日")

            # 解析时间段
            time_parsed = _parse_time_range(time_range_str)
            if not time_parsed:
                continue
            blocked = _get_blocked_hours(time_parsed[0], time_parsed[1])

            if date_str not in leaves[person]:
                leaves[person][date_str] = set()
            leaves[person][date_str] |= blocked
            parsed_count += 1

        if parsed_count == 0:
            print("   (请假库无有效记录)")
        else:
            print(f"   ✅ 加载 {parsed_count} 条请假记录, 涉及 {len(leaves)} 人")
            for person, date_blocks in sorted(leaves.items()):
                for ds in sorted(date_blocks.keys()):
                    hrs = sorted(date_blocks[ds])
                    print(f"      {person}: {ds} 请假 {hrs[0]}:00-{hrs[-1]+1}:00")
        return dict(leaves) if parsed_count > 0 else {}

    except Exception as e:
        print(f"⚠️ 请假库异常: {e}")
        return None


def _apply_leave_to_staff(staff, leaves, date_strs):
    """
    将请假约束转换为每位员工每天的"禁排小时"集合。
    返回: {person: {day_index: set of forbidden_hours}}
    """
    if not leaves:
        return {}

    result = defaultdict(dict)
    date_index = {ds: d for d, ds in enumerate(date_strs)}

    for person, date_blocks in leaves.items():
        for ds, blocked_hours in date_blocks.items():
            if ds not in date_index:
                continue
            d = date_index[ds]
            result[person][d] = blocked_hours

    return dict(result)


def _apply_leave_constraints_cpsat(model, x, person_idx_map, shifts_list,
                                    leave_constraints, d):
    """
    在CP-SAT中对请假日应用约束：请假医生的所有覆盖请假时段的班次变量=0。
    """
    if not leave_constraints:
        return

    for person, p_idx in person_idx_map.items():
        if person not in leave_constraints:
            continue
        if d not in leave_constraints[person]:
            continue
        blocked_hours = leave_constraints[person][d]
        for s, shift in enumerate(shifts_list):
            if shift not in SHIFT_COVERAGE:
                continue
            cov = SHIFT_COVERAGE[shift]
            # 如果这个班次覆盖了任何一个请假时段 → 禁止
            for h in blocked_hours:
                if 0 <= h < 24 and cov[h]:
                    model.Add(x[p_idx, d, s] == 0)
                    break


# ==========================================
# 3. 数据预处理
# ==========================================

def load_and_preprocess_demand(csv_path, month_year):
    print("\n" + "="*60)
    print(f"📊 Phase 1: 数据预处理 (容忍度={TOLERANCE_MINUTES}min, 仅白天7-18 HC转换)")
    print("="*60)
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    print(f"   读取 {len(df)} 行")
    df['ds'] = pd.to_datetime(df['ds'], errors='coerce')
    df = df.dropna(subset=['ds'])
    df['pred_doc_minutes'] = df['pred_doc_minutes'].fillna(0)
    df['pred_tech_minutes'] = df['pred_tech_minutes'].fillna(0)
    df = df[(df['ds'].dt.year == month_year[0]) & (df['ds'].dt.month == month_year[1])]
    if df.empty: raise ValueError(f"❌ 数据中没有 {month_year[0]}-{month_year[1]:02d} 的数据！")

    # --- 体检中心 doc_minutes 推迟到下午 ---
    # 体检中心的医生报告不需要当场完成，均匀分布到当天13:00-17:00
    if 'eps_dept_desc' in df.columns:
        tj_mask = (df['大分类'] == '超声') & df['eps_dept_desc'].str.contains('Health Management', na=False)
        if tj_mask.any():
            tj_doc_moved = 0
            for date_val in df.loc[tj_mask, 'ds'].dt.date.unique():
                date_tj = tj_mask & (df['ds'].dt.date == date_val)
                total_doc = df.loc[date_tj, 'pred_doc_minutes'].sum()
                if total_doc <= 0:
                    continue
                # 清零原始小时的doc
                df.loc[date_tj, 'pred_doc_minutes'] = 0
                # 均匀分布到全天 8:00-17:00 (10小时)
                per_hour = total_doc / 10
                for h in [8, 9, 10, 11, 12, 13, 14, 15, 16, 17]:
                    pm_mask = (df['ds'].dt.date == date_val) & (df['大分类'] == '超声') & (df['ds'].dt.hour == h)
                    if pm_mask.any():
                        pm_indices = df[pm_mask].index
                        per_row = per_hour / len(pm_indices)
                        df.loc[pm_indices, 'pred_doc_minutes'] += per_row
                tj_doc_moved += total_doc
            print(f"   体检doc推迟到下午: {tj_doc_moved:.0f}min ({tj_doc_moved/60:.0f}h)")

    df_grouped = df.groupby(['ds', '大分类']).agg(
        pred_doc_minutes=('pred_doc_minutes', 'sum'),
        pred_tech_minutes=('pred_tech_minutes', 'sum')).reset_index()
    df_grouped['hour'] = df_grouped['ds'].dt.hour
    df_grouped['date'] = df_grouped['ds'].dt.date
    print(f"   聚合后 {len(df_grouped)} 条")

    df_grouped = _merge_small_demand(df_grouped)

    us_mask = df_grouped['大分类'] == '超声'
    df_grouped.loc[us_mask, 'pred_doc_minutes'] += df_grouped.loc[us_mask, 'pred_tech_minutes']
    df_grouped.loc[us_mask, 'pred_tech_minutes'] = 0

    all_dates = sorted(df_grouped['date'].unique())
    date_strs = [d.strftime("%m月%d日") for d in all_dates]
    hourly_hc = defaultdict(lambda: defaultdict(lambda: np.zeros(24)))

    for _, row in df_grouped.iterrows():
        date_str = row['date'].strftime("%m月%d日")
        h = int(row['hour'])
        dept = row['大分类']
        if dept == '放射':
            hourly_hc[date_str]['放射_放射医生'][h] = max(
                hourly_hc[date_str]['放射_放射医生'][h],
                _compute_hc_v2(row['pred_doc_minutes'], '放射医生', h))
            hourly_hc[date_str]['放射_放射技师'][h] = max(
                hourly_hc[date_str]['放射_放射技师'][h],
                _compute_hc_v2(row['pred_tech_minutes'], '放射技师', h))
        elif dept == '超声':
            hourly_hc[date_str]['超声_B超医生'][h] = max(
                hourly_hc[date_str]['超声_B超医生'][h],
                _compute_hc_v2(row['pred_doc_minutes'], 'B超医生', h))

    # 峰值平滑: 避免单小时尖刺拉高全天排班
    hourly_hc = _smooth_peak_hc(hourly_hc, date_strs)

    print(f"   生成 {len(date_strs)} 天 HC需求矩阵 (容忍度={TOLERANCE_MINUTES}分钟)")
    _print_demand_summary(hourly_hc, date_strs)
    return dict(hourly_hc), date_strs, all_dates


def _merge_small_demand(df):
    df = df.sort_values(['大分类', 'date', 'hour']).copy()
    for dept in df['大分类'].unique():
        mask = df['大分类'] == dept
        dept_idx = df[mask].index
        for i in range(len(dept_idx) - 1):
            idx, next_idx = dept_idx[i], dept_idx[i + 1]
            for col in ['pred_doc_minutes', 'pred_tech_minutes']:
                if df.at[idx, col] < 5:
                    df.at[next_idx, col] += df.at[idx, col]
                    df.at[idx, col] = 0
    return df


def _compute_hc_v2(workload_mins, role, hour):
    """HC转换：仅白天(7-18)参与。夜间由硬约束保证1个夜班覆盖。"""
    if workload_mins <= 0: return 0
    if hour < 7 or hour > 18: return 0  # 夜间不参与HC转换, 夜班由CP-SAT硬约束覆盖
    role_key = role if role in LOAD_RATE else ('放射医生' if '放射' in role else 'B超医生')
    load_rate = LOAD_RATE.get(role_key, {}).get(hour, 0.8)
    effective_mins = 60 * load_rate
    if effective_mins <= 0: return 1 if workload_mins > 0 else 0
    n = 1
    remaining = workload_mins - effective_mins
    while remaining > TOLERANCE_MINUTES:
        n += 1; remaining -= effective_mins
    return min(n, 10)


def _print_demand_summary(hourly_hc, date_strs):
    print(f"\n{'日期':10} | {'放射技师 白/夜':15} | {'放射医生 白/夜':15} | {'B超医生 白/夜':15}")
    print("-"*65)
    for ds in date_strs[:5]:
        tech = hourly_hc[ds].get('放射_放射技师', np.zeros(24))
        doc = hourly_hc[ds].get('放射_放射医生', np.zeros(24))
        us = hourly_hc[ds].get('超声_B超医生', np.zeros(24))
        print(f"{ds:10} | {_max_day_night(tech)} | {_max_day_night(doc)} | {_max_day_night(us)}")
    if len(date_strs) > 5: print(f"... 共 {len(date_strs)} 天")


def _max_day_night(arr):
    day = int(max(arr[8:18])) if any(arr[8:18]) else 0
    night = int(max(max(arr[18:24]), max(arr[0:8]))) if any(arr[18:24]) or any(arr[0:8]) else 0
    return f"白:{day} 夜:{night}"


def _smooth_peak_hc(hourly_hc, date_strs):
    """峰值平滑: 仅白天(7-18)处理。上午(7-12)、下午(13-18)各6小时。
    规则: max_HC在块内出现≥3小时→保留; 出现<3小时→cap到次高值(或出现≥3次的值)。
    夜间不参与——由硬约束保证1个夜班覆盖。"""
    from collections import Counter
    morning = list(range(7, 13))    # 07:00-12:00 (6h)
    afternoon = list(range(13, 19))  # 13:00-18:00 (6h)

    sm_count = 0
    for ds in date_strs:
        for role_key in list(hourly_hc[ds].keys()):
            arr = hourly_hc[ds][role_key]
            for block_name, hours in [("morning", morning), ("afternoon", afternoon)]:
                vals = [int(arr[h]) for h in hours]
                cnt = Counter(v for v in vals if v > 0)
                if len(cnt) <= 1:
                    continue
                max_val = max(cnt.keys())
                # 出现≥3小时 → 保留
                if cnt[max_val] >= 3:
                    continue
                # 出现<3小时的尖刺 → 找下一个出现≥3次的值作为cap
                cap = None
                for v in sorted(cnt.keys(), reverse=True):
                    if cnt[v] >= 3:
                        cap = v
                        break
                if cap is None:
                    # 没有≥3次的值，取次高值
                    keys = sorted(cnt.keys(), reverse=True)
                    cap = keys[1] if len(keys) >= 2 else keys[0]
                if cap is not None:
                    for h in hours:
                        if int(arr[h]) > cap:
                            sm_count += 1
                            arr[h] = cap
    if sm_count > 0:
        print(f"   峰值平滑: {sm_count} 个尖刺被cap (仅白天7-18)")
    return hourly_hc


# ==========================================
# 4. L/N 预分配 (Phase 2)
# ==========================================

def pre_allocate_ln(hourly_hc, date_strs, staff, all_dates):
    """
    Phase 2: L/N (24H班) 预分配。
    每位全职放射医生/技师每月2个L/N，均匀间隔分布。
    L/N 属于独立类别，不计入 80%/20%/备班 分层。
    返回: (ln_schedule, ln_hours, ln_skip_dates)
    """
    print("\n" + "="*60)
    print("🔧 Phase 2: L/N (24H班) 预分配")
    print("="*60)

    ln_schedule = defaultdict(dict)
    ln_hours = defaultdict(float)
    ln_skip_dates = defaultdict(set)  # L/N当天不能排其他班(L/N覆盖24h)

    for role in ['放射医生', '放射技师']:
        fulltime = staff[role]['fulltime']
        cfg = ROLE_CONFIG[role]
        ln_per_person = cfg['ln_per_month']
        if ln_per_person <= 0 or not fulltime:
            continue

        assignments = _distribute_ln_shifts(fulltime, date_strs, ln_per_person)
        for person, dates in assignments.items():
            for ds in dates:
                ln_schedule[person][ds] = 'L/N'
                ln_hours[person] += SHIFT_DICT['L/N'][2]
                ln_skip_dates[person].add(ds)
                print(f"    L/N: {DISPLAY_NAME.get(person, person)} → {ds}")

    # 统计
    for role in ['放射医生', '放射技师']:
        ft = staff[role]['fulltime']
        print(f"  [{role}] L/N分配:")
        for p in ft:
            ln_dates = list(ln_schedule.get(p, {}).keys())
            if ln_dates:
                print(f"    {DISPLAY_NAME.get(p, p):25} L/N×{len(ln_dates)} ({', '.join(sorted(ln_dates))})")

    return dict(ln_schedule), dict(ln_hours), dict(ln_skip_dates)


def _distribute_ln_shifts(fulltime_staff, date_strs, count_per_person=2):
    """均匀间隔分配L/N班次"""
    n_people = len(fulltime_staff)
    if n_people == 0:
        return {}
    total_ln = n_people * count_per_person
    n_days = len(date_strs)
    assignments = defaultdict(list)
    spacing = n_days / total_ln if total_ln > 0 else n_days
    for i in range(total_ln):
        person = fulltime_staff[i % n_people]
        day_idx = int(i * spacing)
        if day_idx < n_days:
            ds = date_strs[day_idx]
            assignments[person].append(ds)
    return dict(assignments)


# ==========================================
# 5. 三阶段 CP-SAT 求解器 (Phase 3-5)
# ==========================================

# --- 5A. 通用工具 ---

def _get_date_weekday(ds, all_dates, date_strs):
    """根据日期字符串判断星期几 (0=周一, 6=周日)"""
    idx = date_strs.index(ds)
    dt = all_dates[idx]
    return dt.weekday()


def _get_role_key(role_name):
    return {
        '放射医生': '放射_放射医生',
        '放射技师': '放射_放射技师',
        'B超医生': '超声_B超医生',
    }[role_name]


def _build_shift_list(role_name, demand_by_date, date_strs):
    """筛选与需求相关的可用班次"""
    available = ROLE_SHIFTS[role_name]
    role_key = _get_role_key(role_name)
    used = set()
    for ds in date_strs:
        d = demand_by_date.get(ds, {}).get(role_key, np.zeros(24))
        for s in available:
            if s in SHIFT_COVERAGE:
                cov = SHIFT_COVERAGE[s]
                if any(cov[h] == 1 and d[h] > 0 for h in range(24)):
                    used.add(s)
    # 确保常用班次
    defaults = {
        '放射医生': ['D', 'L', 'L/N', 'N'],
        '放射技师': ['D', 'D2', 'D6', 'N', 'N2', 'L/N', 'H1', 'H3'],
        'B超医生': ['D', 'D5', 'D2', 'H1', 'H2', 'H3', 'T'],
    }
    used.update(defaults.get(role_name, ['D']))
    return [s for s in available if s in used]


def _compute_stage_coverage(person_shifts_for_date):
    """计算某天的基础覆盖(来自之前阶段的排班)。返回 24-element np.array。"""
    cov = np.zeros(24)
    for shift_name in person_shifts_for_date:
        if shift_name and shift_name in SHIFT_COVERAGE:
            sc = SHIFT_COVERAGE[shift_name]
            for h in range(24):
                if sc[h]:
                    cov[h] += 1
    return cov


# --- 5B. Stage 1: 80%工时池 ---

def solve_stage1_80pct(hourly_hc, date_strs, staff, ln_schedule, ln_skip_dates,
                        ln_hours, all_dates, roles=None, leave_constraints=None):
    """
    Stage 1: 80%工时池 CP-SAT
    - 变量: 仅全职人员
    - 硬约束: 每人总工时 ≤ 140.8h (含L/N), 每天≤1班(含L/N), 工作规则
    - 目标: 最大化需求覆盖 + 每日最低班次要求 + 一三五Dustin优先
    - roles: 可选角色列表，默认全部三个角色
    返回: (schedule, hours) — 所有本阶段排班标注 "80%"
    """
    from ortools.sat.python import cp_model
    if roles is None:
        roles = ['放射医生', '放射技师', 'B超医生']
    print("\n" + "="*60)
    print(f"🧮 Phase 3: Stage 1 — 80%工时池 CP-SAT ({', '.join(roles)})")
    print("="*60)

    result_schedule = defaultdict(dict)
    result_hours = defaultdict(float)

    all_results = {}

    for role_name in roles:
        print(f"\n--- {role_name} Stage 1 (80%池) ---")
        fulltime = staff[role_name]['fulltime']
        backup = staff[role_name]['backup']
        cfg = ROLE_CONFIG[role_name]
        role_key = _get_role_key(role_name)

        if not fulltime:
            print(f"  无全职人员，跳过")
            continue

        all_staff = fulltime  # Stage 1 只用全职
        n_staff = len(all_staff)
        n_days = len(date_strs)

        shifts_list = _build_shift_list(role_name, hourly_hc, date_strs)
        n_shifts = len(shifts_list)
        shift_hours = [SHIFT_DICT[s][2] for s in shifts_list]
        shift_is_night = [1 if s in NIGHT_SHIFTS else 0 for s in shifts_list]
        shift_is_ln = [1 if s == 'L/N' else 0 for s in shifts_list]
        shift_is_day = [1 if s in DAY_SHIFTS else 0 for s in shifts_list]

        # 计算每人已有工时（来自L/N预分配）和每天已有排班
        existing_hours = {p: ln_hours.get(p, 0) for p in all_staff}
        existing_shift_d = {}  # {person: {d: shift_name}}
        for p_idx, p in enumerate(all_staff):
            existing_shift_d[p] = {}
            for d, ds in enumerate(date_strs):
                ln_shift = ln_schedule.get(p, {}).get(ds, '')
                if ln_shift:
                    existing_shift_d[p][d] = ln_shift

        model = cp_model.CpModel()

        # Variables: x[p, d, s]
        x = {}
        for p in range(n_staff):
            for d in range(n_days):
                for s in range(n_shifts):
                    x[p, d, s] = model.NewBoolVar(f's1_{role_name[:2]}_{p}_{d}_{s}')

        # --- Constraints ---

        # C1: 每人每天最多1班 (与L/N互斥)
        for p in range(n_staff):
            person = all_staff[p]
            for d, ds in enumerate(date_strs):
                has_ln = 1 if d in existing_shift_d.get(person, {}) else 0
                # L/N skip: 次日不能排班
                in_skip = 1 if ds in ln_skip_dates.get(person, set()) else 0
                if has_ln or in_skip:
                    model.Add(sum(x[p, d, s] for s in range(n_shifts)) == 0)
                else:
                    model.Add(sum(x[p, d, s] for s in range(n_shifts)) <= 1)

        # C2: 月工时 ≤ TARGET_HOURS_FULL (共享上限)
        TOL_DECIHOURS = 0
        for p in range(n_staff):
            person = all_staff[p]
            base_hrs = existing_hours.get(person, 0)
            total_dec = sum(x[p, d, s] * int(shift_hours[s] * 10)
                          for d in range(n_days) for s in range(n_shifts))
            total_with_base = total_dec + int(base_hrs * 10)
            model.Add(total_with_base <= int(TARGET_HOURS_FULL * 10))

        # C3: L/N限制 (每人每月 ≤ 2, 含已分配)
        for p in range(n_staff):
            person = all_staff[p]
            base_ln_count = len(ln_schedule.get(person, {}))
            supp_ln = sum(x[p, d, s] for d in range(n_days) for s in range(n_shifts) if shift_is_ln[s])
            model.Add(supp_ln <= max(0, cfg['ln_per_month'] - base_ln_count))

        # C4: 不连续夜班 (含L/N)
        for p in range(n_staff):
            person = all_staff[p]
            for d in range(n_days - 1):
                # 当天夜班
                base_night_curr = 1 if existing_shift_d.get(person, {}).get(d, '') in (NIGHT_SHIFTS | {'L/N'}) else 0
                base_night_next = 1 if existing_shift_d.get(person, {}).get(d + 1, '') in (NIGHT_SHIFTS | {'L/N'}) else 0
                supp_night_curr = sum(x[p, d, s] for s in range(n_shifts) if shift_is_night[s] or shift_is_ln[s])
                supp_night_next = sum(x[p, d + 1, s] for s in range(n_shifts) if shift_is_night[s] or shift_is_ln[s])
                model.Add(base_night_curr + base_night_next + supp_night_curr + supp_night_next <= 1)

        # C5: 夜班后不排白班
        for p in range(n_staff):
            person = all_staff[p]
            for d in range(n_days - 1):
                base_night = 1 if existing_shift_d.get(person, {}).get(d, '') in (NIGHT_SHIFTS | {'L/N'}) else 0
                supp_night = sum(x[p, d, s] for s in range(n_shifts) if shift_is_night[s] or shift_is_ln[s])
                supp_day_next = sum(x[p, d + 1, s] for s in range(n_shifts) if shift_is_day[s])
                if base_night:
                    model.Add(supp_day_next == 0)
                else:
                    model.Add(supp_night + supp_day_next <= 1)

        # ================================================================
        # 放射医生: 简化轮替模型
        # - 全职: 每天≥1个全天白班, Dustin Wed+Fri, L/N, 0夜班, 176h
        # - 兼职: 每天1人夜班(轮转, 若数学不可行由slack兜底→备班)
        # - 不考虑需求覆盖 (需求由备班满足)
        # ================================================================
        if role_name == '放射医生':
            full_day_is = [1 if s in FULL_DAY_SHIFTS else 0 for s in shifts_list]

            # 兼职人员 (独立的 night-only pool)
            parttime = staff[role_name].get('parttime', [])
            n_pt = len(parttime)
            pt_hours_target = 80.0  # 兼职月夜班上限 (~5.5个N班, 28天÷8人≈3.5)
            pt_slack_vars = {}

            # 为兼职创建独立的夜班变量 (y[pt_idx, d, s])
            y = {}
            # 兼职只用 N (17:30-08:00, 14.5h)
            night_shifts_list = ['N']
            night_s_indices = [i for i, s in enumerate(shifts_list) if s == 'N']
            if n_pt > 0:
                for pt in range(n_pt):
                    for d in range(n_days):
                        for si in night_s_indices:
                            y[pt, d, si] = model.NewBoolVar(f's1_pt_{pt}_{d}_{si}')

                # 兼职约束: 每人每天最多1个夜班
                for pt in range(n_pt):
                    for d in range(n_days):
                        pt_day_vars = [y[pt, d, si] for si in night_s_indices]
                        if pt_day_vars:
                            model.Add(sum(pt_day_vars) <= 1)

                # 兼职月工时上限
                for pt in range(n_pt):
                    pt_total = sum(y[pt, d, si] * int(SHIFT_DICT[night_shifts_list[i]][2] * 10)
                                  for d in range(n_days) for i, si in enumerate(night_s_indices))
                    model.Add(pt_total <= int(pt_hours_target * 10))
                    # 每人至少2个夜班（公平轮转，28天÷8人≈3.5）
                    model.Add(pt_total >= 2 * int(SHIFT_DICT[night_shifts_list[0]][2] * 10) - 5)

                # 不连续夜班 (兼职)
                for pt in range(n_pt):
                    for d in range(n_days - 1):
                        pt_n1 = sum(y[pt, d, si] for si in night_s_indices)
                        pt_n2 = sum(y[pt, d + 1, si] for si in night_s_indices)
                        model.Add(pt_n1 + pt_n2 <= 1)

            # --- 全职约束 ---
            for d, ds in enumerate(date_strs):
                ln_covers = 0
                for p in range(n_staff):
                    person = all_staff[p]
                    if existing_shift_d.get(person, {}).get(d, '') == 'L/N':
                        ln_covers = 1
                        break

                if ln_covers:
                    day_vars = [x[p, d, s] for p in range(n_staff) for s in range(n_shifts)
                               if not shift_is_night[s] and not shift_is_ln[s]]
                    if day_vars:
                        model.Add(sum(day_vars) == 0)
                else:
                    # 软交替: ≥1白班, 两人同天→轻微惩罚
                    all_day_vars = [x[p, d, s] for p in range(n_staff) for s in range(n_shifts)
                                   if not shift_is_night[s] and not shift_is_ln[s]]
                    if all_day_vars:
                        model.Add(sum(all_day_vars) >= 1)

                # 全职0夜班
                ft_night_vars = [x[p, d, s] for p in range(n_staff) for s in range(n_shifts) if shift_is_night[s]]
                if ft_night_vars:
                    model.Add(sum(ft_night_vars) == 0)

                # 兼职夜班: 每天=1 (硬约束)
                if n_pt > 0:
                    pt_night_vars_d = [y[pt, d, si] for pt in range(n_pt) for si in night_s_indices]
                    if ln_covers:
                        if pt_night_vars_d:
                            model.Add(sum(pt_night_vars_d) == 0)
                    else:
                        if pt_night_vars_d:
                            ptsl = model.NewIntVar(0, 1, f's1_ptsl_{d}')
                            pt_slack_vars[d] = ptsl
                            model.Add(sum(pt_night_vars_d) + ptsl == 1)

            # --- Objective ---
            objective_terms = []

            # Dustin Wed+Fri 硬约束
            if DUSTIN_RAD in fulltime:
                dustin_p = all_staff.index(DUSTIN_RAD)
                for d, ds in enumerate(date_strs):
                    wd = _get_date_weekday(ds, all_dates, date_strs)
                    if wd in (2, 4):
                        has_ln = existing_shift_d.get(DUSTIN_RAD, {}).get(d, '') == 'L/N'
                        after_ln = d > 0 and existing_shift_d.get(DUSTIN_RAD, {}).get(d - 1, '') == 'L/N'
                        if not has_ln and not after_ln:
                                dv = [x[dustin_p, d, s] for s in range(n_shifts)
                                      if not shift_is_night[s] and not shift_is_ln[s]]
                                if dv:
                                    dustin_day = sum(dv)
                                    objective_terms.append(dustin_day * (S1_COVERAGE_WEIGHT * 5))

            # 强奖励多排班 + 严重惩罚超目标
            for p in range(n_staff):
                person = all_staff[p]
                total_dec = sum(x[p, d, s] * int(shift_hours[s] * 10)
                              for d in range(n_days) for s in range(n_shifts))
                base_dec = int(existing_hours.get(person, 0) * 10)
                target_dec = int(TARGET_HOURS_FULL * 10)
                objective_terms.append(total_dec * 1000)
                # 超过目标 → 极高惩罚
                over = model.NewIntVar(0, target_dec * 2, f's1_over_{p}')
                model.Add(over >= total_dec + base_dec - target_dec)
                objective_terms.append(over * (-S1_COVERAGE_WEIGHT))

            # 同天惩罚: 两人同时白班→惩罚
            for d, ds in enumerate(date_strs):
                li_var = sum(x[0, d, s] for s in range(n_shifts) if not shift_is_night[s] and not shift_is_ln[s])
                du_var = sum(x[1, d, s] for s in range(n_shifts) if not shift_is_night[s] and not shift_is_ln[s])
                both = model.NewIntVar(0, 1, f's1_both_{d}')
                model.Add(both >= li_var + du_var - 1)
                model.Add(both <= li_var)
                model.Add(both <= du_var)
                objective_terms.append(both * (-5000))

            # 全职工时均衡
            if len(fulltime) >= 2:
                for p in range(n_staff):
                    person = all_staff[p]
                    total_dec = sum(x[p, d, s] * int(shift_hours[s] * 10)
                                  for d in range(n_days) for s in range(n_shifts))
                    base_h = int(existing_hours.get(person, 0) * 10)
                    avg_target = int(TARGET_HOURS_FULL * 10)
                    dev = model.NewIntVar(0, int(TARGET_HOURS_FULL * 10), f's1_dev_{p}')
                    model.Add(dev >= total_dec + base_h - avg_target)
                    model.Add(dev >= avg_target - (total_dec + base_h))
                    objective_terms.append(dev * (-S1_BALANCE_WEIGHT * 10))

            # 兼职夜班缺口惩罚
            for sl in pt_slack_vars.values():
                objective_terms.append(sl * (-S1_COVERAGE_WEIGHT * 5))

            # 兼职夜班均衡
            if n_pt >= 2:
                for pt in range(n_pt):
                    pt_total = sum(y[pt, d, si] * int(SHIFT_DICT[night_shifts_list[i]][2] * 10)
                                  for d in range(n_days) for i, si in enumerate(night_s_indices))
                    avg_pt = int(pt_hours_target * 10)
                    pt_dev = model.NewIntVar(0, int(pt_hours_target * 10), f's1_ptdev_{pt}')
                    model.Add(pt_dev >= pt_total - avg_pt)
                    model.Add(pt_dev >= avg_pt - pt_total)
                    objective_terms.append(pt_dev * (-S1_BALANCE_WEIGHT * 10))

            model.Maximize(sum(objective_terms))

            # Solve for radiologists
            solver = cp_model.CpSolver()
            solver.parameters.max_time_in_seconds = 120
            solver.parameters.num_search_workers = 8
            status = solver.Solve(model)
            print(f"   {role_name} Stage1 求解: {solver.StatusName(status)}")

            role_result = {}
            if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
                for p, person in enumerate(all_staff):
                    for d, ds in enumerate(date_strs):
                        for s, shift in enumerate(shifts_list):
                            if solver.Value(x[p, d, s]):
                                existing = role_result.get(person, {}).get(ds, '')
                                if existing:
                                    role_result.setdefault(person, {})[ds] = existing + ' + ' + shift
                                else:
                                    role_result.setdefault(person, {})[ds] = shift
                                result_hours[person] += shift_hours[s]
                for p in fulltime:
                    result_hours[p] = result_hours.get(p, 0) + ln_hours.get(p, 0)

                # 提取兼职夜班结果
                if n_pt > 0:
                    for pt in range(n_pt):
                        pt_name = parttime[pt]
                        for d, ds in enumerate(date_strs):
                            for i, si in enumerate(night_s_indices):
                                if solver.Value(y[pt, d, si]):
                                    role_result.setdefault(pt_name, {})[ds] = night_shifts_list[i]
                                    result_hours[pt_name] += SHIFT_DICT[night_shifts_list[i]][2]
                    # 打印兼职统计
                    for pt in range(n_pt):
                        pt_name = parttime[pt]
                        pt_hrs = result_hours.get(pt_name, 0)
                        pt_days = sum(1 for ds in date_strs if ds in role_result.get(pt_name, {}))
                        print(f"   兼职 {pt_name}: {pt_hrs:.1f}h (夜班 {pt_days} 天)")

            all_results[role_name] = role_result
            continue  # 放射医生已处理，跳过后续逻辑

        # ================================================================
        # ================================================================
        # 放射技师 / B超医生: 先硬需求→再80%池→再20%池→备班
        # ================================================================
        day_slack_vars = {}
        night_slack_vars = {}
        h24_slack_vars = {}
        dustin_wf_slack = {}
        full_day_is = [1 if s in FULL_DAY_SHIFTS else 0 for s in shifts_list]

        for d, ds in enumerate(date_strs):
            base_coverage = np.zeros(24)
            base_night_count = 0
            base_full_day_count = 0
            for p in range(n_staff):
                person = all_staff[p]
                shift = existing_shift_d.get(person, {}).get(d, '')
                if shift and shift in SHIFT_COVERAGE:
                    cov_arr = SHIFT_COVERAGE[shift]
                    for hh in range(24):
                        if cov_arr[hh]:
                            base_coverage[hh] += 1
                    if shift in NIGHT_SHIFTS or shift == 'L/N':
                        base_night_count += 1
                    if shift in FULL_DAY_SHIFTS or shift == 'L/N':
                        base_full_day_count += 1

            # C0: 请假硬约束 — 请假时段禁止排任何覆盖该时段的班次
            if leave_constraints:
                person_idx_map = {}
                for p, person in enumerate(all_staff):
                    if person in leave_constraints and d in leave_constraints[person]:
                        person_idx_map[person] = p
                _apply_leave_constraints_cpsat(model, x, person_idx_map, shifts_list,
                                                leave_constraints, d)

            supp_ln_vars_d = [x[p, d, s] for p in range(n_staff) for s in range(n_shifts) if shift_is_ln[s]]

            # 全天白班数量 (高惩罚slack)
            day_target = cfg['day_shifts']
            if role_name == 'B超医生':
                wd = _get_date_weekday(ds, all_dates, date_strs)
                if wd == 6:
                    day_target = 2  # 周日2个白班
                elif wd in (1, 3):
                    day_target = 3  # 周二周四≥3人(保证PM=3)

            if day_target > 0:
                supp_full_day = [x[p, d, s] for p in range(n_staff) for s in range(n_shifts)
                                if full_day_is[s]]
                # 放射医生: 硬约束≥1全天白班 (L/N也算)，极小 slack 兜底仅2人情况
                if cfg.get('night_prefer_backup'):
                    if base_full_day_count < 1:
                        dslack2 = model.NewIntVar(0, 1, f's1_rddslack_{d}')
                        day_slack_vars[d] = dslack2
                        model.Add(sum(supp_full_day) + sum(supp_ln_vars_d) + base_full_day_count + dslack2 >= 1)
                else:
                    # 放射技师: 硬约束=2个全天白班, ≤3人(仪器限制), L/N日除外
                    if role_name == '放射技师':
                        # 白天总数 ≤3 (含L/N覆盖)
                        all_day = [x[p, d, s] for p in range(n_staff) for s in range(n_shifts)
                                  if not shift_is_night[s] and not shift_is_ln[s]]
                        model.Add(sum(all_day) + base_full_day_count <= 3)
                        # 恰好2个全天白班
                        if base_full_day_count >= day_target:
                            model.Add(sum(supp_full_day) + sum(supp_ln_vars_d) == 0)
                        else:
                            model.Add(sum(supp_full_day) + sum(supp_ln_vars_d) + base_full_day_count == day_target)
                    elif role_name == 'B超医生':
                        # 全天白班 ≥ day_target (硬)
                        model.Add(sum(supp_full_day) + sum(supp_ln_vars_d) + base_full_day_count >= day_target)
                        # 每天总人数 ≤4 (任何班型)
                        all_us = [x[p, d, s] for p in range(n_staff) for s in range(n_shifts)
                                 if not shift_is_night[s] and not shift_is_ln[s]]
                        model.Add(sum(all_us) + base_full_day_count <= 4)
                    else:
                        dslack = model.NewIntVar(0, day_target, f's1_dslack_{d}')
                        day_slack_vars[d] = dslack
                        model.Add(sum(supp_full_day) + sum(supp_ln_vars_d) + base_full_day_count + dslack >= day_target)

            # 夜班数量 (放射医生硬约束≥1, slack fallback)
            if cfg['night_shifts'] > 0:
                supp_night_vars = [x[p, d, s] for p in range(n_staff) for s in range(n_shifts) if shift_is_night[s]]
                if cfg.get('night_prefer_backup'):
                    # 放射医生: 夜班优先由备班覆盖，保持软约束
                    if base_night_count < 1:
                        nslack = model.NewIntVar(0, 1, f's1_rdnslack_{d}')
                        night_slack_vars[d] = nslack
                        model.Add(sum(supp_night_vars) + sum(supp_ln_vars_d) + base_night_count + nslack >= cfg['night_shifts'])
                else:
                    # 放射技师: 硬约束每天恰好1个夜班 (无slack)
                    if role_name == '放射技师':
                        model.Add(sum(supp_night_vars) + sum(supp_ln_vars_d) + base_night_count == cfg['night_shifts'])
                    else:
                        nslack = model.NewIntVar(0, cfg['night_shifts'], f's1_nslack_{d}')
                        night_slack_vars[d] = nslack
                        model.Add(sum(supp_night_vars) + sum(supp_ln_vars_d) + base_night_count + nslack >= cfg['night_shifts'])

            # 24h覆盖 (仅放射技师，极高惩罚slack)
            if cfg['coverage_24h']:
                for h in range(24):
                    coverage_vars = [x[p, d, s] for p in range(n_staff) for s, shift in enumerate(shifts_list)
                                   if h < len(SHIFT_COVERAGE.get(shift, [])) and SHIFT_COVERAGE[shift][h] == 1]
                    if coverage_vars:
                        total_base = int(base_coverage[h])
                        h24slack = model.NewIntVar(0, 2, f's1_h24slack_{d}_{h}')
                        h24_slack_vars[d, h] = h24slack
                        model.Add(sum(coverage_vars) + total_base + h24slack >= 1)

        # C8: B超医生 下午约束 (Wed+Fri=2人, Tue+Thu=3人) — 硬约束
        if role_name == 'B超医生':
            pm_is = [1 if (s in FULL_DAY_SHIFTS or s == 'H3') else 0 for s in shifts_list]
            pm_dates_wf = pm_dates_tt = 0
            for d, ds in enumerate(date_strs):
                wd = _get_date_weekday(ds, all_dates, date_strs)
                pm_vars = [x[p, d, s] for p in range(n_staff) for s in range(n_shifts) if pm_is[s]]
                if not pm_vars:
                    continue
                if wd in (2, 4):
                    model.Add(sum(pm_vars) == 2)
                    pm_dates_wf += 1
                elif wd in (1, 3):
                    model.Add(sum(pm_vars) == 3)
                    pm_dates_tt += 1
            if pm_dates_wf > 0 or pm_dates_tt > 0:
                print(f"   [B超] PM硬约束: Wed+Fri=2人({pm_dates_wf}天), Tue+Thu=3人({pm_dates_tt}天)")

        # C9: B超医生固定配对约束 (仅Stage 1 80%池)
        # 合法配对: Liu+Xu, Xu+Hou, Xu+Lu, Lu+Hou (禁止 Liu+Lu, Liu+Hou)
        if role_name == 'B超医生':
            liu_idx = next((i for i, p in enumerate(fulltime) if 'Liu' in p), None)
            lu_idx = next((i for i, p in enumerate(fulltime) if 'Lu' in p and 'Liu' not in p), None)
            hou_idx = next((i for i, p in enumerate(fulltime) if 'hou' in p.lower()), None)
            for d in range(n_days):
                full_day_vars = {}
                for p_idx in [liu_idx, lu_idx, hou_idx]:
                    if p_idx is not None:
                        full_day_vars[p_idx] = sum(x[p_idx, d, s] for s in range(n_shifts) if full_day_is[s])
                # 禁止 Liu + Lu
                if liu_idx is not None and lu_idx is not None:
                    model.Add(full_day_vars[liu_idx] + full_day_vars[lu_idx] <= 1)
                # 禁止 Liu + Hou
                if liu_idx is not None and hou_idx is not None:
                    model.Add(full_day_vars[liu_idx] + full_day_vars[hou_idx] <= 1)

        # C7: 覆盖需求 (极高惩罚软约束 — 优先用全职，不够的留给备班)
        coverage_slack_vars = {}
        for d, ds in enumerate(date_strs):
            demand = hourly_hc.get(ds, {}).get(role_key, np.zeros(24))
            base_coverage = np.zeros(24)
            for p in range(n_staff):
                person = all_staff[p]
                shift = existing_shift_d.get(person, {}).get(d, '')
                if shift and shift in SHIFT_COVERAGE:
                    for hh in range(24):
                        if SHIFT_COVERAGE[shift][hh]:
                            base_coverage[hh] += 1

            for h in range(24):
                if demand[h] <= 0:
                    continue
                coverage_vars = [x[p, d, s] for p in range(n_staff) for s, shift in enumerate(shifts_list)
                               if h < len(SHIFT_COVERAGE.get(shift, [])) and SHIFT_COVERAGE[shift][h] == 1]
                if coverage_vars:
                    needed = int(demand[h])
                    sl = model.NewIntVar(0, max(1, needed), f's1_cov_{d}_{h}')
                    coverage_slack_vars[d, h] = sl
                    model.Add(sum(coverage_vars) + int(base_coverage[h]) + sl >= needed)

        # Dustin Wed+Fri: 极高惩罚——强制周三周五上放射全天白班
        if role_name == '放射医生' and DUSTIN_RAD in fulltime:
            dustin_p = all_staff.index(DUSTIN_RAD)
            for d, ds in enumerate(date_strs):
                wd = _get_date_weekday(ds, all_dates, date_strs)
                if wd in (2, 4):  # 周三(2)、周五(4)
                    dustin_day_vars = [x[dustin_p, d, s] for s in range(n_shifts)
                                      if full_day_is[s] or shift_is_ln[s]]
                    dustin_slack = model.NewIntVar(0, 1, f's1_dustinWFs_{d}')
                    dustin_wf_slack[d] = dustin_slack
                    model.Add(sum(dustin_day_vars) + dustin_slack >= 1)

        # --- Objective ---
        objective_terms = []

        # 需求覆盖缺口 → 极高惩罚 (P0: 优先全职覆盖)
        for sl in coverage_slack_vars.values():
            objective_terms.append(sl * (-S1_COVERAGE_WEIGHT * 10))

        # 基础保障缺口 → 高惩罚 (P1: 日间/夜班人数)
        for sl in day_slack_vars.values():
            objective_terms.append(sl * (-S1_COVERAGE_WEIGHT * 5))
        for sl in night_slack_vars.values():
            objective_terms.append(sl * (-S1_COVERAGE_WEIGHT * 5))
        for sl in h24_slack_vars.values():
            objective_terms.append(sl * (-S1_COVERAGE_WEIGHT * 5))

        # Dustin Wed+Fri 缺口 → 极高惩罚 (P0级别)
        for sl in dustin_wf_slack.values():
            objective_terms.append(sl * (-S1_COVERAGE_WEIGHT * 10))

        # 放射医生: 全职上夜班 → 惩罚 (鼓励用备班覆盖夜班)
        if cfg.get('night_prefer_backup'):
            for p, person in enumerate(all_staff):
                if person in fulltime:
                    for d in range(n_days):
                        for s in range(n_shifts):
                            if shift_is_night[s]:
                                objective_terms.append(x[p, d, s] * (-S1_COVERAGE_WEIGHT * 3))

        # 工时均衡
        if len(fulltime) >= 2:
            ft_indices = list(range(n_staff))
            for p in ft_indices:
                person = all_staff[p]
                total_dec = sum(x[p, d, s] * int(shift_hours[s])
                              for d in range(n_days) for s in range(n_shifts))
                base_h = int(existing_hours.get(person, 0) * 10)
                avg_target = int(TARGET_HOURS_FULL * 10)
                dev = model.NewIntVar(0, int(TARGET_HOURS_FULL * 10), f's1_dev_{p}')
                model.Add(dev >= total_dec + base_h - avg_target)
                model.Add(dev >= avg_target - (total_dec + base_h))
                objective_terms.append(dev * (-S1_BALANCE_WEIGHT))

        model.Maximize(sum(objective_terms))

        # Solve
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 120
        solver.parameters.num_search_workers = 8
        status = solver.Solve(model)
        print(f"   {role_name} Stage1 求解: {solver.StatusName(status)}")

        role_result = {}
        if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            for p, person in enumerate(all_staff):
                for d, ds in enumerate(date_strs):
                    for s, shift in enumerate(shifts_list):
                        if solver.Value(x[p, d, s]):
                            existing = role_result.get(person, {}).get(ds, '')
                            if existing:
                                role_result.setdefault(person, {})[ds] = existing + ' + ' + shift
                            else:
                                role_result.setdefault(person, {})[ds] = shift
                            result_hours[person] += shift_hours[s]
            # 加上L/N工时（始终计入，确保Stage2的176h上限包含L/N）
            for p in fulltime:
                result_hours[p] = result_hours.get(p, 0) + ln_hours.get(p, 0)

        all_results[role_name] = role_result

    # 合并结果: 标记来源为 "80%"
    for role_name, role_result in all_results.items():
        for person, date_shifts in role_result.items():
            for ds, shift in date_shifts.items():
                result_schedule[person][ds] = (shift, "80%")

    # L/N也加入schedule（来源: "L/N"）
    for person, date_shifts in ln_schedule.items():
        for ds, shift in date_shifts.items():
            existing = result_schedule[person].get(ds, None)
            if existing:
                # L/N已经在这个日期, 来源保持L/N
                result_schedule[person][ds] = (shift, "L/N")
            else:
                result_schedule[person][ds] = (shift, "L/N")

    return dict(result_schedule), dict(result_hours)


# --- 5C. Stage 2: 20%工时池 ---

def solve_stage2_20pct(hourly_hc, date_strs, staff, stage1_schedule, stage1_hours,
                        ln_schedule, ln_skip_dates, all_dates, leave_constraints=None):
    """
    Stage 2: 20%工时池 CP-SAT
    - 固定 Stage 1 排班
    - 变量: 全职人员额外班次
    - 硬约束: 每人总工时 ≤ 176h, 每天总共≤1班(含Stage1)
    - 目标: 最大化覆盖 Stage 1 未覆盖的剩余需求
    返回: (schedule, hours) — 新增排班标注 "20%"
    """
    from ortools.sat.python import cp_model
    print("\n" + "="*60)
    print("🧮 Phase 4: Stage 2 — 20%工时池 CP-SAT")
    print("="*60)

    result_schedule = defaultdict(dict)
    result_hours = defaultdict(float)

    all_results = {}

    for role_name in ['放射医生', '放射技师', 'B超医生']:
        print(f"\n--- {role_name} Stage 2 (20%池) ---")
        fulltime = staff[role_name]['fulltime']
        cfg = ROLE_CONFIG[role_name]
        role_key = _get_role_key(role_name)

        if not fulltime:
            continue

        n_staff = len(fulltime)
        n_days = len(date_strs)

        shifts_list = _build_shift_list(role_name, hourly_hc, date_strs)
        n_shifts = len(shifts_list)
        shift_hours = [SHIFT_DICT[s][2] for s in shifts_list]
        shift_is_night = [1 if s in NIGHT_SHIFTS else 0 for s in shifts_list]
        shift_is_ln = [1 if s == 'L/N' else 0 for s in shifts_list]
        shift_is_day = [1 if s in DAY_SHIFTS else 0 for s in shifts_list]
        full_day_is = [1 if s in FULL_DAY_SHIFTS else 0 for s in shifts_list]
        TOL_DECIHOURS = 0  # 8h tolerance (lower bound = TARGET - 8h)

        # 计算Stage1每天的覆盖和每人已用工时
        s1_hours = {p: stage1_hours.get(p, 0) for p in fulltime}
        s1_shift_d = {}
        for p_idx, p in enumerate(fulltime):
            s1_shift_d[p] = {}
            for d, ds in enumerate(date_strs):
                entry = stage1_schedule.get(p, {}).get(ds, None)
                if entry:
                    s1_shift_d[p][d] = entry[0] if isinstance(entry, tuple) else entry

        model = cp_model.CpModel()

        # Variables: x[p, d, s]
        x = {}
        for p in range(n_staff):
            for d in range(n_days):
                for s in range(n_shifts):
                    x[p, d, s] = model.NewBoolVar(f's2_{role_name[:2]}_{p}_{d}_{s}')

        # --- Constraints ---

        # C1: 每人每天最多1班 (含Stage1 + L/N)
        for p in range(n_staff):
            person = fulltime[p]
            for d, ds in enumerate(date_strs):
                has_s1 = 1 if d in s1_shift_d.get(person, {}) else 0
                in_skip = 1 if ds in ln_skip_dates.get(person, set()) else 0
                if has_s1 or in_skip:
                    model.Add(sum(x[p, d, s] for s in range(n_shifts)) == 0)
                else:
                    model.Add(sum(x[p, d, s] for s in range(n_shifts)) <= 1)

        # C0: 请假硬约束
        if leave_constraints:
            for d, ds in enumerate(date_strs):
                person_idx_map = {}
                for p, person in enumerate(fulltime):
                    if person in leave_constraints and d in leave_constraints[person]:
                        person_idx_map[person] = p
                _apply_leave_constraints_cpsat(model, x, person_idx_map, shifts_list,
                                                leave_constraints, d)

        for p in range(n_staff):
            person = fulltime[p]
            base_hrs = s1_hours.get(person, 0)
            total_dec = sum(x[p, d, s] * int(shift_hours[s] * 10)
                          for d in range(n_days) for s in range(n_shifts))
            cap_dec = int((TARGET_HOURS_FULL - base_hrs) * 10)
            if cap_dec < 0:
                cap_dec = 0
            total_with_base = total_dec + int(base_hrs * 10)
            target_dec = int(TARGET_HOURS_FULL * 10)
            model.Add(total_with_base <= target_dec)
            model.Add(total_with_base >= target_dec - TOL_DECIHOURS)

        # C3: L/N限制
        for p in range(n_staff):
            person = fulltime[p]
            s1_ln = sum(1 for v in s1_shift_d.get(person, {}).values() if v == 'L/N')
            supp_ln = sum(x[p, d, s] for d in range(n_days) for s in range(n_shifts) if shift_is_ln[s])
            model.Add(supp_ln <= max(0, cfg['ln_per_month'] - s1_ln))

        # C4: 不连续夜班
        for p in range(n_staff):
            person = fulltime[p]
            for d in range(n_days - 1):
                s1_night_curr = 1 if s1_shift_d.get(person, {}).get(d, '') in (NIGHT_SHIFTS | {'L/N'}) else 0
                s1_night_next = 1 if s1_shift_d.get(person, {}).get(d + 1, '') in (NIGHT_SHIFTS | {'L/N'}) else 0
                supp_night_curr = sum(x[p, d, s] for s in range(n_shifts) if shift_is_night[s] or shift_is_ln[s])
                supp_night_next = sum(x[p, d + 1, s] for s in range(n_shifts) if shift_is_night[s] or shift_is_ln[s])
                model.Add(s1_night_curr + s1_night_next + supp_night_curr + supp_night_next <= 1)

        # C5: 夜班后不排白班
        for p in range(n_staff):
            person = fulltime[p]
            for d in range(n_days - 1):
                s1_night = 1 if s1_shift_d.get(person, {}).get(d, '') in (NIGHT_SHIFTS | {'L/N'}) else 0
                supp_night = sum(x[p, d, s] for s in range(n_shifts) if shift_is_night[s] or shift_is_ln[s])
                supp_day_next = sum(x[p, d + 1, s] for s in range(n_shifts) if shift_is_day[s])
                if s1_night:
                    model.Add(supp_day_next == 0)
                else:
                    model.Add(supp_night + supp_day_next <= 1)

        # C6: 覆盖剩余需求 (硬约束，无slack! — 优先消耗全职20%池)
        for d, ds in enumerate(date_strs):
            demand = hourly_hc.get(ds, {}).get(role_key, np.zeros(24))
            # Stage1覆盖
            s1_coverage = np.zeros(24)
            for p in range(n_staff):
                person = fulltime[p]
                shift = s1_shift_d.get(person, {}).get(d, '')
                if shift and shift in SHIFT_COVERAGE:
                    for hh in range(24):
                        if SHIFT_COVERAGE[shift][hh]:
                            s1_coverage[hh] += 1
            remaining = np.maximum(demand - s1_coverage, 0)

            for h in range(24):
                if remaining[h] <= 0:
                    continue
                coverage_vars = [x[p, d, s] for p in range(n_staff) for s, shift in enumerate(shifts_list)
                               if h < len(SHIFT_COVERAGE.get(shift, [])) and SHIFT_COVERAGE[shift][h] == 1]
                if coverage_vars:
                    needed = int(remaining[h])
                    # 硬约束: 尽量满足，但上限为全职人数
                    model.Add(sum(coverage_vars) <= min(needed, n_staff))

        # C7: 放射技师每天≤1夜班 (硬上限, 防止Stage2在Stage1夜班日叠加第二个夜班)
        if role_name == '放射技师':
            for d, ds in enumerate(date_strs):
                s1_night_d = sum(1 for p in range(n_staff)
                               if s1_shift_d.get(fulltime[p], {}).get(d, '') in (NIGHT_SHIFTS | {'L/N'}))
                supp_night_d = sum(x[p, d, s] for p in range(n_staff) for s in range(n_shifts)
                                  if shift_is_night[s] or shift_is_ln[s])
                model.Add(supp_night_d + s1_night_d <= 1)

        # B超 PM约束: Stage1已保证, Stage2不限制(merge时硬切)

        # --- Objective: 最大化全职工时使用 ---
        objective_terms = []

        # 放射医生 Stage2: 夜班惩罚 (20%池也不排夜班，但约束优先)
        if cfg.get('night_prefer_backup'):
            for p in range(n_staff):
                person = fulltime[p]
                for d in range(n_days):
                    for s in range(n_shifts):
                        if shift_is_night[s]:
                            objective_terms.append(x[p, d, s] * (-S2_COVERAGE_WEIGHT // 10))

        for p in range(n_staff):
            person = fulltime[p]
            cap = max(0, int((TARGET_HOURS_FULL - s1_hours.get(person, 0)) * 10))
            if cap > 0:
                total10 = sum(x[p, d, s] * int(shift_hours[s] * 10)
                            for d in range(n_days) for s in range(n_shifts))
                objective_terms.append(total10)  # 奖励多排班

        model.Maximize(sum(objective_terms))

        # Solve
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 120
        solver.parameters.num_search_workers = 8
        status = solver.Solve(model)
        print(f"   {role_name} Stage2 求解: {solver.StatusName(status)}")

        role_result = {}
        if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            for p, person in enumerate(fulltime):
                for d, ds in enumerate(date_strs):
                    for s, shift in enumerate(shifts_list):
                        if solver.Value(x[p, d, s]):
                            role_result.setdefault(person, {})[ds] = shift
                            result_hours[person] += shift_hours[s]

        all_results[role_name] = role_result

    # 合并: 标记来源 "20%"
    for role_name, role_result in all_results.items():
        for person, date_shifts in role_result.items():
            for ds, shift in date_shifts.items():
                result_schedule[person][ds] = (shift, "20%")

    return dict(result_schedule), dict(result_hours)


# --- 5D. Stage 3: 备班池全覆盖 ---

def solve_stage3_backup(hourly_hc, date_strs, staff,
                         stage1_schedule, stage1_hours,
                         stage2_schedule, stage2_hours,
                         ln_schedule, ln_skip_dates, all_dates,
                         leave_constraints=None):
    """
    Stage 3: 备班池 CP-SAT — 硬性全覆盖
    - 固定 Stage 1+2 排班
    - 变量: 备班人员
      - 放射医生/B超医生: 整班变量 (backup_shift_based=True)
      - 放射技师: 仅按小时变量 (backup_shift_based=False)
    - 硬约束: 所有剩余需求必须覆盖 (无slack!)
    - 目标: 最小化备班使用量
    返回: (schedule, hours) — 标注 "备班"
    """
    from ortools.sat.python import cp_model
    print("\n" + "="*60)
    print("🧮 Phase 5: Stage 3 — 备班池全覆盖 CP-SAT")
    print("="*60)

    result_schedule = defaultdict(dict)
    result_hours = defaultdict(float)

    all_remaining = {}  # 记录每个角色最终未覆盖的需求

    for role_name in ['放射医生', '放射技师', 'B超医生']:
        print(f"\n--- {role_name} Stage 3 (备班池) ---")
        fulltime = staff[role_name]['fulltime']
        backup = staff[role_name]['backup']
        cfg = ROLE_CONFIG[role_name]
        role_key = _get_role_key(role_name)

        if not backup:
            print(f"  无备班人员，跳过")
            continue

        all_staff = backup
        n_staff_bk = len(backup)
        n_days = len(date_strs)
        backup_shift_based = cfg['backup_shift_based']

        # 收集Stage1+2所有人员的排班
        all_prev_shifts = {}  # {person: {d: shift_name}}
        for p in fulltime:
            all_prev_shifts[p] = {}
            for d, ds in enumerate(date_strs):
                s1 = stage1_schedule.get(p, {}).get(ds, None)
                s2 = stage2_schedule.get(p, {}).get(ds, None)
                shifts = []
                if s1: shifts.append(s1[0] if isinstance(s1, tuple) else s1)
                if s2: shifts.append(s2[0] if isinstance(s2, tuple) else s2)
                if shifts: all_prev_shifts[p][d] = ' + '.join(shifts)

        for p in backup:
            all_prev_shifts[p] = {}
            # 备班人员之前的排班（可能来自前两阶段如果未来扩展）

        if backup_shift_based:
            # 放射医生/B超医生: 备班可排整班
            shifts_list = _build_shift_list(role_name, hourly_hc, date_strs)
            n_shifts = len(shifts_list)
            shift_hours = [SHIFT_DICT[s][2] for s in shifts_list]

            model = cp_model.CpModel()

            # Variables: x[s, d, s_idx] for backup pool (备班是池子，很多人!)
            x = {}
            # 备班池无人数上限 — 用单索引表示所有可用备班
            # 实际: 每天对每种班次最多需要的数量 = demand所需的最大HC
            max_backup_per_day = 8  # 备班池最多一天出8个人(足够大)
            for bk_idx in range(max_backup_per_day):
                for d in range(n_days):
                    for s in range(n_shifts):
                        x[bk_idx, d, s] = model.NewBoolVar(f's3s_{role_name[:2]}_{bk_idx}_{d}_{s}')

            # C1: 备班池每人每天最多1班 (每个备班个体)
            for bk_idx in range(max_backup_per_day):
                for d in range(n_days):
                    model.Add(sum(x[bk_idx, d, s] for s in range(n_shifts)) <= 1)

            # C2: L/N限制 - 备班不上L/N
            for bk_idx in range(max_backup_per_day):
                for d in range(n_days):
                    for s in range(n_shifts):
                        if shifts_list[s] == 'L/N':
                            model.Add(x[bk_idx, d, s] == 0)

            # C3: 同一个备班不连续夜班
            shift_is_night_bk = [1 if s in NIGHT_SHIFTS else 0 for s in shifts_list]
            for bk_idx in range(max_backup_per_day):
                for d in range(n_days - 1):
                    n_curr = sum(x[bk_idx, d, s] for s in range(n_shifts) if shift_is_night_bk[s])
                    n_next = sum(x[bk_idx, d + 1, s] for s in range(n_shifts) if shift_is_night_bk[s])
                    model.Add(n_curr + n_next <= 1)

            # C3b: 放射医生 — 每天必须恰好1个夜班 (硬约束!)
            if cfg.get('night_prefer_backup'):
                for d, ds in enumerate(date_strs):
                    # 全职已排夜班数
                    ft_night = 0
                    for person in fulltime:
                        shift_str = all_prev_shifts.get(person, {}).get(d, '')
                        if shift_str:
                            for sname in shift_str.split(' + '):
                                if sname in NIGHT_SHIFTS or sname == 'L/N':
                                    ft_night += 1
                    # 还需要备班补的夜班数
                    needed = max(0, 1 - ft_night)
                    if needed > 0:
                        backup_night = [x[bk_idx, d, s] for bk_idx in range(max_backup_per_day)
                                      for s in range(n_shifts) if shift_is_night_bk[s]]
                        # 备班必须补上夜班缺口
                        model.Add(sum(backup_night) >= needed)

            # C4: 硬性全覆盖 (无slack! 备班池无限)
            for d, ds in enumerate(date_strs):
                demand = hourly_hc.get(ds, {}).get(role_key, np.zeros(24))
                prev_cov = np.zeros(24)
                for person, shifts_d in all_prev_shifts.items():
                    shift_str = shifts_d.get(d, '')
                    if shift_str:
                        for sname in shift_str.split(' + '):
                            if sname and sname in SHIFT_COVERAGE:
                                for hh in range(24):
                                    if SHIFT_COVERAGE[sname][hh]:
                                        prev_cov[hh] += 1

                remaining = np.maximum(demand - prev_cov, 0)

                for h in range(24):
                    if remaining[h] <= 0:
                        continue
                    coverage_vars = [x[bk_idx, d, s] for bk_idx in range(max_backup_per_day)
                                   for s, shift in enumerate(shifts_list)
                                   if h < len(SHIFT_COVERAGE.get(shift, [])) and SHIFT_COVERAGE[shift][h] == 1]
                    if coverage_vars:
                        needed = max(1, int(math.ceil(remaining[h])))
                        model.Add(sum(coverage_vars) >= needed)

            # Objective: 最小化备班使用量
            objective_terms = []
            for bk_idx in range(max_backup_per_day):
                for d in range(n_days):
                    for s in range(n_shifts):
                        objective_terms.append(x[bk_idx, d, s] * (-S3_BACKUP_MINIMIZE * int(shift_hours[s])))
            model.Maximize(sum(objective_terms))

            # Solve
            solver = cp_model.CpSolver()
            solver.parameters.max_time_in_seconds = 120
            solver.parameters.num_search_workers = 8
            status = solver.Solve(model)
            print(f"   {role_name} Stage3(整班) 求解: {solver.StatusName(status)}")

            if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
                person = backup[0]  # 所有备班池的班次合并到备班名下
                total_bk_hours = 0.0
                for bk_idx in range(max_backup_per_day):
                    for d, ds in enumerate(date_strs):
                        for s, shift in enumerate(shifts_list):
                            if solver.Value(x[bk_idx, d, s]):
                                existing = result_schedule.get(person, {}).get(ds, None)
                                if existing:
                                    old = existing[0] if isinstance(existing, tuple) else existing
                                    result_schedule[person][ds] = (old + ' + ' + shift, "备班")
                                else:
                                    result_schedule[person][ds] = (shift, "备班")
                                total_bk_hours += shift_hours[s]
                result_hours[person] = total_bk_hours
                print(f"   备班池使用: {total_bk_hours:.0f}h (整班模式)")
            else:
                print(f"   ❌ 备班整班模式不可行！")


        else:
            # 放射技师: 备班仅按小时 (无班型!)
            model = cp_model.CpModel()

            # Variables: bk[p, d, h] — 备班人员p在第d天的第h小时上班
            bk = {}
            for p in range(n_staff_bk):
                for d in range(n_days):
                    for h in range(24):
                        bk[p, d, h] = model.NewBoolVar(f's3h_{role_name[:2]}_{p}_{d}_{h}')

            # C1: 备班每人每天不超过12小时
            for p in range(n_staff_bk):
                for d in range(n_days):
                    model.Add(sum(bk[p, d, h] for h in range(24)) <= 12)

            # C2: 硬性全覆盖 (无slack!)
            for d, ds in enumerate(date_strs):
                demand = hourly_hc.get(ds, {}).get(role_key, np.zeros(24))
                # Stage1+2覆盖
                prev_cov = np.zeros(24)
                for person, shifts_d in all_prev_shifts.items():
                    shift_str = shifts_d.get(d, '')
                    if shift_str:
                        for sname in shift_str.split(' + '):
                            if sname and sname in SHIFT_COVERAGE:
                                for hh in range(24):
                                    if SHIFT_COVERAGE[sname][hh]:
                                        prev_cov[hh] += 1

                remaining = np.maximum(demand - prev_cov, 0)

                for h in range(24):
                    if remaining[h] <= 0:
                        continue
                    needed = max(1, int(math.ceil(remaining[h])))
                    coverage_vars = [bk[p, d, h] for p in range(n_staff_bk)]
                    model.Add(sum(coverage_vars) >= needed)

            # Objective: 最小化备班总工时
            objective_terms = []
            for p in range(n_staff_bk):
                for d in range(n_days):
                    for h in range(24):
                        objective_terms.append(bk[p, d, h] * (-S3_BACKUP_MINIMIZE))
            model.Maximize(sum(objective_terms))

            # Solve
            solver = cp_model.CpSolver()
            solver.parameters.max_time_in_seconds = 120
            solver.parameters.num_search_workers = 8
            status = solver.Solve(model)
            print(f"   {role_name} Stage3(小时制) 求解: {solver.StatusName(status)}")

            if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
                for p, person in enumerate(backup):
                    for d, ds in enumerate(date_strs):
                        bk_hours_today = []
                        for h in range(24):
                            if solver.Value(bk[p, d, h]):
                                bk_hours_today.append(h)
                        if bk_hours_today:
                            ranges = _merge_hour_ranges(bk_hours_today)
                            time_str = ', '.join(ranges)
                            label = f"备班({time_str})"
                            result_schedule[person][ds] = (label, "备班")
                            result_hours[person] += len(bk_hours_today)

    return dict(result_schedule), dict(result_hours)


def _solve_backup_hourly_fallback(role_name, hourly_hc, date_strs, fulltime, backup,
                                   role_key, all_prev_shifts, result_schedule, result_hours):
    """整班备班不可行时的按小时回退求解器。
    第一尝试: 硬性全覆盖(≤12h/天)。若仍不可行，转软约束最小化未覆盖。"""
    from ortools.sat.python import cp_model
    n_staff_bk = len(backup)
    n_days = len(date_strs)

    def _build_hourly_model(use_slack=False):
        model = cp_model.CpModel()
        bk = {}
        for p in range(n_staff_bk):
            for d in range(n_days):
                for h in range(24):
                    bk[p, d, h] = model.NewBoolVar(f's3hfb_{p}_{d}_{h}')

        slack_vars = {}

        # 每天≤12小时 (若只有1人且需求>12h，适当放宽到14h)
        daily_cap = 14 if (n_staff_bk == 1) else 12
        for p in range(n_staff_bk):
            for d in range(n_days):
                model.Add(sum(bk[p, d, h] for h in range(24)) <= daily_cap)

        # 覆盖约束
        for d, ds in enumerate(date_strs):
            demand = hourly_hc.get(ds, {}).get(role_key, np.zeros(24))
            prev_cov = np.zeros(24)
            for person, shifts_d in all_prev_shifts.items():
                shift_str = shifts_d.get(d, '')
                if shift_str:
                    for sname in shift_str.split(' + '):
                        if sname and sname in SHIFT_COVERAGE:
                            for hh in range(24):
                                if SHIFT_COVERAGE[sname][hh]:
                                    prev_cov[hh] += 1
            remaining = np.maximum(demand - prev_cov, 0)
            for h in range(24):
                if remaining[h] <= 0:
                    continue
                needed = min(max(1, int(math.ceil(remaining[h]))), n_staff_bk)
                coverage_vars = [bk[p, d, h] for p in range(n_staff_bk)]
                if use_slack:
                    sl = model.NewIntVar(0, needed, f'slack_{d}_{h}')
                    slack_vars[d, h] = sl
                    model.Add(sum(coverage_vars) + sl >= needed)
                else:
                    model.Add(sum(coverage_vars) >= needed)

        # Objective
        obj = []
        if use_slack:
            # 高惩罚未覆盖缺口 + 最小化备班工时
            for sl in slack_vars.values():
                obj.append(sl * (-1_000_000))
        for p in range(n_staff_bk):
            for d in range(n_days):
                for h in range(24):
                    obj.append(bk[p, d, h] * (-S3_BACKUP_MINIMIZE))
        model.Maximize(sum(obj))
        return model, bk, slack_vars

    # Try 1: hard coverage
    model, bk, _ = _build_hourly_model(use_slack=False)
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30
    solver.parameters.num_search_workers = 4
    status = solver.Solve(model)

    if status == cp_model.INFEASIBLE:
        print(f"   ⚠️ 小时硬覆盖仍不可行，使用软约束(最小化缺口)...")
        model, bk, slack_vars = _build_hourly_model(use_slack=True)
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 60
        solver.parameters.num_search_workers = 4
        status = solver.Solve(model)

    print(f"   {role_name} Stage3(回退小时) 求解: {solver.StatusName(status)}")

    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        for p, person in enumerate(backup):
            for d, ds in enumerate(date_strs):
                bk_hours_today = []
                for h in range(24):
                    if solver.Value(bk[p, d, h]):
                        bk_hours_today.append(h)
                if bk_hours_today:
                    ranges = _merge_hour_ranges(bk_hours_today)
                    time_str = ', '.join(ranges)
                    label = f"备班({time_str})"
                    result_schedule[person][ds] = (label, "备班")
                    result_hours[person] += len(bk_hours_today)
    else:
        print(f"   ❌ {role_name} 备班仍不可行！部分需求可能无法覆盖")


def _merge_hour_ranges(hours):
    """合并连续小时: [9,10,11,14,15] → ['09:00-12:00', '14:00-16:00']"""
    if not hours: return []
    sorted_hours = sorted(set(hours))
    ranges, start, end = [], sorted_hours[0], sorted_hours[0]
    for h in sorted_hours[1:]:
        if h == end + 1: end = h
        else:
            ranges.append(f"{start:02d}:00-{end+1:02d}:00")
            start = h; end = h
    ranges.append(f"{start:02d}:00-{end+1:02d}:00")
    return ranges


# ==========================================
# 6. 合并最终排班 + OnCall分配 (Phase 6)
# ==========================================

def merge_and_oncall(stage1_schedule, stage1_hours,
                     stage2_schedule, stage2_hours,
                     stage3_schedule, stage3_hours,
                     date_strs, staff, all_dates):
    """
    合并三阶段排班 + 分配 OnCall（不计工时）。
    返回:
      - final_schedule: {person: {ds: (shift_str, category)}}
      - final_hours: {person: total_hours}
      - category_hours: {person: {"80%": h, "20%": h, "备班": h, "L/N": h}}
      - oncall_schedule: {person: {ds: True}}
    """
    print("\n" + "="*60)
    print("🔧 Phase 6: 合并排班 + OnCall分配")
    print("="*60)

    final_schedule = defaultdict(dict)
    final_hours = defaultdict(float)
    category_hours = defaultdict(lambda: {"80%": 0.0, "20%": 0.0, "备班": 0.0, "L/N": 0.0})
    oncall_schedule = defaultdict(dict)

    # 合并三阶段
    all_stages = [
        (stage1_schedule, "80%"),
        (stage2_schedule, "20%"),
        (stage3_schedule, "备班"),
    ]

    for stage_sched, category in all_stages:
        for person, date_shifts in stage_sched.items():
            for ds, value in date_shifts.items():
                shift_name = value[0] if isinstance(value, tuple) else value
                actual_cat = category
                # L/N 固定类别
                if shift_name == 'L/N' or 'L/N' in shift_name:
                    actual_cat = "L/N"

                existing = final_schedule[person].get(ds, None)
                if existing:
                    old_shift = existing[0] if isinstance(existing, tuple) else existing
                    final_schedule[person][ds] = (old_shift + ' + ' + shift_name, actual_cat)
                else:
                    final_schedule[person][ds] = (shift_name, actual_cat)

                shift_hrs = _get_shift_hours(shift_name)
                final_hours[person] += shift_hrs
                category_hours[person][actual_cat] += shift_hrs

    # 兼职放射医生: 合并前清理白天班 + 备班清理重复夜班
    for role_name in ['放射医生']:
        parttime = staff[role_name].get('parttime', [])
        backup_names = staff[role_name].get('backup', [])
        for pt_name in parttime:
            if pt_name in final_schedule:
                for ds in list(final_schedule[pt_name].keys()):
                    val = final_schedule[pt_name][ds]
                    shift_str = val[0] if isinstance(val, tuple) else str(val)
                    cat = val[1] if isinstance(val, tuple) else "80%"
                    night_only = [s for s in shift_str.split(' + ') if s.strip() in NIGHT_SHIFTS]
                    if night_only:
                        final_schedule[pt_name][ds] = (' + '.join(night_only), cat)
                        # 当天已有兼职夜班 → 备班不排夜班
                        for bk_name in backup_names:
                            if bk_name in final_schedule and ds in final_schedule[bk_name]:
                                bk_val = final_schedule[bk_name][ds]
                                bk_shift = bk_val[0] if isinstance(bk_val, tuple) else str(bk_val)
                                bk_cat = bk_val[1] if isinstance(bk_val, tuple) else "备班"
                                bk_day = [s for s in bk_shift.split(' + ') if s.strip() not in NIGHT_SHIFTS]
                                if bk_day:
                                    final_schedule[bk_name][ds] = (' + '.join(bk_day), bk_cat)
                                else:
                                    del final_schedule[bk_name][ds]
                    else:
                        del final_schedule[pt_name][ds]

    


    # B超 Wed+Fri 下午=2: 硬切多余为H2
    us_ft = [p for p in staff['B超医生']['fulltime'] if 'US' not in p]
    for d, ds in enumerate(date_strs):
        if all_dates[d].weekday() in (2, 4):
            pm_list = [(p, final_schedule[p][ds]) for p in us_ft
                      if p in final_schedule and ds in final_schedule[p]
                      and any(s.strip() in FULL_DAY_SHIFTS or s.strip() == 'H3'
                           for s in str(final_schedule[p][ds][0] if isinstance(final_schedule[p][ds], tuple) else final_schedule[p][ds]).split(' + '))]
            while len(pm_list) > 2:
                p_to_move = pm_list[-1][0]
                old_val = final_schedule[p_to_move][ds]
                old_cat = old_val[1] if isinstance(old_val, tuple) else '80%'
                final_schedule[p_to_move][ds] = ('H2', old_cat)
                pm_list.pop()

    # 重新算全部工时（已清理兼职）
    final_hours.clear()
    for person in list(category_hours.keys()):
        for cat in category_hours[person]:
            category_hours[person][cat] = 0.0

    for person, date_shifts in final_schedule.items():
        total = 0.0
        for ds, value in date_shifts.items():
            if isinstance(value, tuple):
                shift_str, cat = value
            else:
                shift_str, cat = value, "80%"
            hrs = _get_shift_hours(shift_str)
            total += hrs
            category_hours[person][cat] += hrs
        final_hours[person] = total

    

        # 80/20 + OT 分离: ≤140.8→80%, >140.8→20%, >176→OT(不计入80/20)
    ot_hours = defaultdict(float)
    for person in list(final_schedule.keys()):
        if person not in final_schedule: continue
        accumulated = 0.0
        for ds, value in sorted(final_schedule[person].items()):
            shift_str = value[0] if isinstance(value, tuple) else str(value)
            hrs = _get_shift_hours(shift_str)
            accumulated += hrs
            old_cat = value[1] if isinstance(value, tuple) and value[1] else "80%"
            if accumulated > TARGET_HOURS_FULL:
                final_schedule[person][ds] = (shift_str, "OT")
                ot_hours.setdefault(person, 0)
                ot_hours[person] += hrs
            elif accumulated > TARGET_HOURS_80 and old_cat not in ("20%", "备班", "L/N"):
                final_schedule[person][ds] = (shift_str, "20%")

    # 重新算工时(反映20%分离)
    for person in list(category_hours.keys()):
        for cat in category_hours[person]:
            category_hours[person][cat] = 0.0
    for person, date_shifts in final_schedule.items():
        for ds, value in date_shifts.items():
            shift_str = value[0] if isinstance(value, tuple) else str(value)
            cat = value[1] if isinstance(value, tuple) else "80%"
            hrs = _get_shift_hours(shift_str)
            if cat == 'OT':
                ot_hours.setdefault(person, 0)
                ot_hours[person] += hrs
            else:
                category_hours[person][cat] += hrs



    # --- OnCall分配 ---
    for role_name in ['放射技师', 'B超医生']:
        cfg = ROLE_CONFIG[role_name]
        fulltime = staff[role_name]['fulltime']
        if not cfg['has_oncall']:
            continue
        if not fulltime:
            continue

        print(f"\n  [{role_name}] OnCall分配:")

        if role_name == 'B超医生':
            # OnCall: 4 US doctors (excl DUSTIN_US) + Dustin (rad days only, max 6)
            us_real = [p for p in fulltime if 'US' not in p and p != DUSTIN_US]
            if DUSTIN_US in oncall_schedule: del oncall_schedule[DUSTIN_US]
            oc_count = {p: 0 for p in us_real + [DUSTIN_RAD]}
            dustin_max = 6
            for ds in date_strs:
                dustin_ok = ds in final_schedule.get(DUSTIN_RAD, {}) and oc_count[DUSTIN_RAD] < dustin_max
                off_today = [p for p in us_real if not final_schedule.get(p, {}).get(ds)]
                if off_today:
                    best = min(off_today, key=lambda p: oc_count[p])
                elif dustin_ok:
                    best = DUSTIN_RAD
                else:
                    best = min(us_real, key=lambda p: oc_count[p])
                save_name = DUSTIN_US if best == DUSTIN_RAD else best
                oncall_schedule[save_name][ds] = True
                oc_count[best] += 1
            for p in us_real + [DUSTIN_RAD]:
                cnt = oc_count.get(p, 0)
                if cnt > 0:
                    print(f"    {DISPLAY_NAME.get(p, p):25} OnCall×{cnt}")
        else:
            # 放射技师 OnCall
            n_oc = len(fulltime)
            oncall_idx = 0
            for ds in date_strs:
                assigned = False
                for _ in range(n_oc * 2):
                    person = fulltime[oncall_idx % n_oc]
                    oncall_idx += 1
                    if not final_schedule.get(person, {}).get(ds):
                        oncall_schedule[person][ds] = True
                        assigned = True
                        break
                if not assigned:
                    person = fulltime[oncall_idx % n_oc]
                    oncall_idx += 1
                    oncall_schedule[person][ds] = True
            for person in fulltime:
                cnt = sum(1 for v in oncall_schedule.get(person, {}).values() if v is True)
                if cnt > 0:
                    print(f"    {DISPLAY_NAME.get(person, person):25} OnCall×{cnt}")

    # --- 半天班合并 + OT计算 ---
    # _merge_half_shifts 会把新增OT合并到传入的ot_hours中
    final_schedule, final_hours, category_hours, merge_ot = _merge_half_shifts(
        final_schedule, final_hours, category_hours, date_strs, staff)
    for k, v in merge_ot.items():
        ot_hours[k] += v

    # 工时已由 Stage 1+2 CP-SAT 约束保证 ≤176h，不做额外裁切

    
    # Debug dump
    if DUSTIN_US in oncall_schedule:
        dus_count = sum(1 for v in oncall_schedule.get(DUSTIN_US, {}).values() if v is True)
        dru_count = sum(1 for v in oncall_schedule.get(DUSTIN_RAD, {}).values() if v is True)
        print(f"[ONCALL DUMP] DUSTIN_US={dus_count} DUSTIN_RAD={dru_count}")
    # 工时由CP-SAT控制
    return dict(final_schedule), dict(final_hours), dict(category_hours), dict(oncall_schedule), dict(ot_hours)


def _merge_half_shifts(final_schedule, final_hours, category_hours, date_strs, staff):
    """
    检测同一天出现 H1+H3 或 H2+H3 的情况 → 合并为全职医生的全天班。
    全职可以OT(加班)，合并后工时超出8.5h部分计入OT。
    返回: (updated_schedule, updated_hours, updated_category_hours, ot_hours)
    """
    ot_hours = defaultdict(float)
    half_pairs = {('H1', 'H3'): ('D6', 10.0), ('H2', 'H3'): ('D', 8.5)}  # (半天1,半天2) -> (合并班型, 工时)
    merged_count = 0

    for ds in date_strs:
        for (h1, h3), (merged_shift, merged_hours) in half_pairs.items():
            # 找人
            h1_person = h3_person = None
            h1_cat = h3_cat = None
            for person, shifts in final_schedule.items():
                if ds not in shifts:
                    continue
                val = shifts[ds]
                shift_str = val[0] if isinstance(val, tuple) else str(val)
                cat = val[1] if isinstance(val, tuple) else "80%"
                if h1 in shift_str.split(' + ') and h1_person is None:
                    h1_person = person; h1_cat = cat
                if h3 in shift_str.split(' + ') and h3_person is None:
                    h3_person = person; h3_cat = cat
            if not h1_person or not h3_person:
                continue

            # 合并给全职医生（如果有）或h1的人
            target = h1_person if not h1_person.startswith('备班') else h3_person
            if target.startswith('备班') and not h3_person.startswith('备班'):
                target = h3_person
            cat = h1_cat if target == h1_person else h3_cat

            # 删除原来两个半天班
            half_to_remove = {h1_person: h1, h3_person: h3}
            for p, half in half_to_remove.items():
                if ds not in final_schedule.get(p, {}):
                    continue  # 同一人的两个半天，第一个迭代已删掉
                old_val = final_schedule[p][ds]
                old_shift = old_val[0] if isinstance(old_val, tuple) else str(old_val)
                old_cat = old_val[1] if isinstance(old_val, tuple) else "80%"
                # 删掉对应的半天
                new_parts = [s.strip() for s in old_shift.split(' + ') if s.strip() not in (h1, h3)]
                if new_parts:
                    final_schedule[p][ds] = (' + '.join(new_parts), old_cat)
                    # 减去被删除的半天工时
                    cat_hours_deduct = SHIFT_DICT.get(h1 if p == h1_person else h3, (0,0,0,0))[2]
                    category_hours[p][old_cat] = max(0, category_hours[p].get(old_cat, 0) - cat_hours_deduct)
                    final_hours[p] -= cat_hours_deduct
                else:
                    del final_schedule[p][ds]
                    cat_hours_deduct = SHIFT_DICT.get(h1 if p == h1_person else h3, (0,0,0,0))[2]
                    category_hours[p][old_cat] = max(0, category_hours[p].get(old_cat, 0) - cat_hours_deduct)
                    final_hours[p] -= cat_hours_deduct

            # 给target加合并后的全天班
            existing = final_schedule[target].get(ds, None)
            if existing:
                old_shift = existing[0] if isinstance(existing, tuple) else str(existing)
                final_schedule[target][ds] = (old_shift + ' + ' + merged_shift, cat)
            else:
                final_schedule[target][ds] = (merged_shift, cat)

            # 更新工时: 前8.5h入池子, 超出部分入OT
            standard_hrs = 8.5
            category_hours[target][cat] += min(merged_hours, standard_hrs)
            if merged_hours > standard_hrs:
                ot_hours[target] += (merged_hours - standard_hrs)
            final_hours[target] += merged_hours
            merged_count += 1

    if merged_count > 0:
        print(f"   半天班合并: {merged_count} 个半天组合 → 全天班 (OT: {sum(ot_hours.values()):.0f}h)")

    return final_schedule, final_hours, category_hours, ot_hours


def _get_shift_hours(shift_str):
    """计算班次字符串的总工时，处理组合班次和备班(小时)"""
    if not shift_str:
        return 0.0
    total = 0.0
    parts = shift_str.split(' + ')
    for part in parts:
        part = part.strip()
        if part.startswith('备班('):
            # 备班(09:00-12:00, 14:00-16:00) → 计算小时数
            time_part = part[3:-1]  # 去掉 '备班(' 和 ')'
            for rng in time_part.split(','):
                rng = rng.strip()
                if '-' in rng:
                    start_str, end_str = rng.split('-')
                    start_h = int(start_str.split(':')[0])
                    end_h = int(end_str.split(':')[0])
                    total += max(0, end_h - start_h)
        elif part in SHIFT_DICT:
            total += SHIFT_DICT[part][2]
    return total


# ==========================================
# 6B. 超声医生楼层备注 (4/9/B1) 公平轮转
# ==========================================

def assign_ultrasound_notes(final_schedule, date_strs, staff):
    """
    为超声医生每日排班添加楼层备注: 4(四楼) / 9(九楼) / B1(二楼+床边)。
    4位全职医生公平轮转，每人每月各楼层天数尽量均衡。
    返回: {person: {ds: note_str}}
    """
    NOTES = ['4', '9', '2(B1)']
    ultrasound_fulltime = staff['B超医生']['fulltime']
    # 过滤掉Dustin_US
    ultrasound_fulltime = [p for p in ultrasound_fulltime if 'US' not in p]

    # 记录每人每类note的累计天数
    note_counts = {p: {n: 0 for n in NOTES} for p in ultrasound_fulltime}

    result = defaultdict(dict)

    for d, ds in enumerate(date_strs):
        # 当天有全天白班的超声医生
        working = []
        for p in ultrasound_fulltime:
            if ds in final_schedule.get(p, {}):
                val = final_schedule[p][ds]
                shift = val[0] if isinstance(val, tuple) else str(val)
                # 只有有全天白班的人才需要分配楼层
                has_full_day = any(s in FULL_DAY_SHIFTS for s in shift.split(' + '))
                if has_full_day:
                    working.append(p)

        if not working:
            continue

        # 当天需要分配的note数量 = min(上班人数, 3)
        n_notes = min(len(working), len(NOTES))

        # 先分配B1: 每天至少1人
        b1_person = min(working, key=lambda p: note_counts[p]['2(B1)'])
        result[b1_person][ds] = '2(B1)'
        note_counts[b1_person]['2(B1)'] += 1
        working.remove(b1_person)
        # 再分配4和9 (剩余人数≤2人)
        remaining_notes = ['4', '9']
        for ni in range(min(len(working), 2)):
            note = remaining_notes[ni]
            best_p = min(working, key=lambda p: note_counts[p][note])
            result[best_p][ds] = note
            note_counts[best_p][note] += 1
            working.remove(best_p)

    # 打印统计
    print(f"\n  [超声楼层备注] 4(四楼) / 9(九楼) / B1(二楼+床边):")
    for p in ultrasound_fulltime:
        counts = {n: sum(1 for ds2 in result.get(p, {}) if result[p][ds2] == n) for n in NOTES}
        display = DISPLAY_NAME.get(p, p)
        print(f"    {display:25} 4楼x{counts['4']}  9楼x{counts['9']}  2(B1)x{counts['2(B1)']}")

    return dict(result)


# ==========================================
# 7. Dustin 跨角色处理 (Phase 7)
# ==========================================

def apply_dustin_cross_role(final_schedule, final_hours, category_hours,
                             hourly_hc, date_strs, staff):
    """
    Dustin 跨角色处理:
    1. 记录 Dustin 放射排班 (已在 final_schedule 中)
    2. 计算 Dustin 超声可用工时: max(0, TARGET_HOURS_FULL - dustin_rad_hours)
    3. Dustin 上放射班的日期: 超声需求 -= 0.5
    4. Dustin 超声排班上限 = 超声可用工时
    5. 超声每天总人数 ≤ 5

    注意: 这个函数在超声排班完成后调用，用于后处理/验证。
    实际的超声需求抵扣和上限约束在 solve_stage1_80pct 的超声部分实现。
    """
    print("\n" + "="*60)
    print("🔧 Phase 7: Dustin 跨角色处理")
    print("="*60)

    dustin_rad_hours = final_hours.get(DUSTIN_RAD, 0)
    dustin_us_available = max(0, TARGET_HOURS_FULL - dustin_rad_hours)

    print(f"   Dustin 放射已用: {dustin_rad_hours:.1f}h")
    print(f"   Dustin 超声可用: {dustin_us_available:.1f}h")

    # Dustin上放射班的日期列表
    dustin_rad_dates = set()
    for ds, value in final_schedule.get(DUSTIN_RAD, {}).items():
        if isinstance(value, tuple):
            shift_str = value[0]
        else:
            shift_str = value
        if shift_str:
            dustin_rad_dates.add(ds)

    if dustin_rad_dates:
        print(f"   Dustin上放射班日期: {len(dustin_rad_dates)}天")
        print(f"   超声需求抵扣: 这些日期超声HC -= 0.5")

    # 检查超声每日人数
    us_fulltime = staff['B超医生']['fulltime']
    us_all = us_fulltime + staff['B超医生']['backup']
    over5_dates = []
    for ds in date_strs:
        count = 0
        for person in us_all:
            if ds in final_schedule.get(person, {}):
                count += 1
        if DUSTIN_US in final_schedule and ds in final_schedule.get(DUSTIN_US, {}):
            count += 1
        if count > 5:
            over5_dates.append(ds)

    if over5_dates:
        print(f"   ⚠️ 超声超5人日期: {over5_dates}")
    else:
        print(f"   ✅ 超声每日≤5人，合规")

    return dustin_us_available, dustin_rad_dates


def apply_dustin_us_deduction(hourly_hc, dustin_rad_dates, date_strs):
    """
    在超声排班前调用：抵扣Dustin放射班日期的超声需求。
    返回新的超声需求矩阵（深拷贝，仅在超声求解时使用）。
    """
    role_key = '超声_B超医生'
    adjusted_hc = defaultdict(lambda: defaultdict(lambda: np.zeros(24)))

    for ds in date_strs:
        original = hourly_hc.get(ds, {}).get(role_key, np.zeros(24)).copy()
        if ds in dustin_rad_dates:
            original = np.maximum(original - 0.5, 0)
        adjusted_hc[ds][role_key] = original

    # 保留其他role_key
    for ds in date_strs:
        for rk in ['放射_放射医生', '放射_放射技师']:
            adjusted_hc[ds][rk] = hourly_hc.get(ds, {}).get(rk, np.zeros(24)).copy()

    return dict(adjusted_hc)


# ==========================================
# 8. Excel 输出 (Phase 8) — 分类标注
# ==========================================

def generate_excel(final_schedule, final_hours, category_hours, hourly_hc,
                    date_strs, staff, oncall_schedule, ot_hours, us_notes,
                    pto_dates, output_path):
    """生成排班Excel (V3: 分类标注 80%/20%/备班/L/N)"""
    print("\n" + "="*60)
    print("📝 Phase 8: 生成Excel输出 (V3 分类标注)")
    print("="*60)

    rad_docs = staff['放射医生']['fulltime'] + staff['放射医生']['backup'] + staff['放射医生'].get('parttime', [])
    rad_techs = staff['放射技师']['fulltime'] + staff['放射技师'].get('parttime', []) + staff['放射技师']['backup']
    us_docs = staff['B超医生']['fulltime'] + staff['B超医生'].get('parttime', []) + staff['B超医生']['backup']
    if DUSTIN_US in final_schedule and DUSTIN_US not in us_docs:
        us_docs.append(DUSTIN_US)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # --- 需求矩阵 sheets ---
        for role_key, sheet_label in [('放射_放射医生', '放射医生'), ('放射_放射技师', '放射技师'), ('超声_B超医生', 'B超医生')]:
            sheet_name = f"需求_{sheet_label}"
            rows = []
            for ds in date_strs:
                row = {"日期": ds}
                hourly = hourly_hc.get(ds, {}).get(role_key, np.zeros(24))
                for h in range(24):
                    row[f"{h:02d}:00"] = int(hourly[h])
                rows.append(row)
            df_demand = pd.DataFrame(rows)
            if not df_demand.empty:
                df_demand.set_index("日期", inplace=True)
                df_demand.to_excel(writer, sheet_name=sheet_name)

        # --- 排班 sheets (含分类标注) ---
        def write_group(sheet_name, members, role_name):
            data = []
            for name in members:
                row = {"人员": DISPLAY_NAME.get(name, name)}
                cat = category_hours.get(name, {"80%": 0, "20%": 0, "备班": 0, "L/N": 0})
                row["80%工时"] = round(cat.get("80%", 0), 1)
                row["20%工时"] = round(cat.get("20%", 0), 1)
                row["备班工时"] = round(cat.get("备班", 0), 1)
                row["L/N工时"] = round(cat.get("L/N", 0), 1)
                ot = ot_hours.get(name, 0)
                if ot > 0:
                    row["OT工时"] = round(ot, 1)
                row["总工时"] = round(final_hours.get(name, 0), 1)
                row["目标"] = TARGET_HOURS_FULL if not name.startswith("备班") else ""
                row["剩余"] = round(TARGET_HOURS_FULL - final_hours.get(name, 0), 1) if not name.startswith("备班") else ""
                oncall_count = len(oncall_schedule.get(name, {}))
                if oncall_count > 0:
                    row["OnCall"] = oncall_count

                for ds in date_strs:
                    entry = final_schedule.get(name, {}).get(ds, None)
                    oncall_today = ds in oncall_schedule.get(name, {})
                    is_pto = ds in pto_dates.get(name, set())
                    if is_pto:
                        display = "PTO"
                    elif entry:
                        if isinstance(entry, tuple):
                            shift_val, cat_label = entry
                            time_parts = []
                            for part in shift_val.split(' + '):
                                part_clean = part.strip()
                                if part_clean.startswith('备班('):
                                    time_parts.append(part_clean)
                                else:
                                    time_parts.append(SHIFT_TIME_STR.get(part_clean, part_clean))
                            time_str = " + ".join(time_parts)
                            display = f"{shift_val}\n{time_str}\n[{cat_label}]"
                        else:
                            shift_val = str(entry)
                            time_str = SHIFT_TIME_STR.get(shift_val, shift_val)
                            display = f"{shift_val}\n{time_str}"
                    else:
                        display = ""
                    if oncall_today:
                        display += "\n📞"
                    # 超声医生楼层备注
                    if role_name == 'B超医生' and ds in us_notes.get(name, {}):
                        display += f"\n[{us_notes[name][ds]}]"
                    row[ds] = display
                data.append(row)

            df = pd.DataFrame(data)
            df.set_index("人员", inplace=True)
            df.to_excel(writer, sheet_name=sheet_name)

            # 应用样式
            workbook = writer.book
            worksheet = workbook[sheet_name]
            # PTO红色背景
            red_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
            for row_idx in range(2, len(data) + 2):
                for col_idx in range(len(date_strs)):
                    cell = worksheet.cell(row=row_idx, column=col_idx + stats_col_count + 1)
                    if cell.value and 'PTO' in str(cell.value):
                        cell.fill = red_fill
                        cell.font = Font(bold=True, color="FFFFFF")
            _apply_v3_excel_styling(worksheet, df, sheet_name, date_strs, category_hours)

        stats_col_count = 8  # 人员列之后空8列再放排班日期

        write_group("放射医生排班", rad_docs, "放射医生")
        write_group("放射技师排班", rad_techs, "放射技师")
        write_group("超声医生排班", us_docs, "B超医生")

        # --- 统计汇总 ---
        summary_data = []
        all_persons = rad_docs + rad_techs + us_docs
        seen = set()
        for name in all_persons:
            if name in seen or name == DUSTIN_US:
                continue
            seen.add(name)
            cat = category_hours.get(name, {"80%": 0, "20%": 0, "备班": 0, "L/N": 0})
            total = final_hours.get(name, 0)

            role = ""
            if name in rad_docs: role = "放射医生"
            elif name in rad_techs: role = "放射技师"
            elif name in us_docs: role = "B超医生"

            summary_data.append({
                "人员": DISPLAY_NAME.get(name, name),
                "角色": role,
                "身份": "全职" if not name.startswith("备班") else "备班",
                "80%工时": round(cat.get("80%", 0), 1),
                "20%工时": round(cat.get("20%", 0), 1),
                "备班工时": round(cat.get("备班", 0), 1),
                "L/N工时": round(cat.get("L/N", 0), 1),
                "总工时": round(total, 1),
                "目标工时": TARGET_HOURS_FULL if not name.startswith("备班") else "",
                "OnCall": len(oncall_schedule.get(name, {})),
            })

        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="统计汇总", index=False)
        workbook = writer.book
        workbook._sheets = [workbook["统计汇总"]] + [s for s in workbook._sheets if s.title != "统计汇总"]

    print(f"✅ Excel已保存: {output_path}")
    return output_path


def _apply_v3_excel_styling(worksheet, df, sheet_name, date_strs, category_hours):
    """V3 Excel样式: 按分类着色"""
    stats_col_count = 8  # 人员 + 80% + 20% + 备班 + L/N + OT + 总工时 + 目标 + 剩余 + OnCall -> 实际从人员列后开始算
    # 统计列数: 80%工时, 20%工时, 备班工时, L/N工时, 总工时, 目标, 剩余, (OnCall)
    header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    stat_cols = 0
    for cell_val in header_row:
        if cell_val in ('80%工时', '20%工时', '备班工时', 'L/N工时', '总工时', '目标', '剩余', 'OnCall'):
            stat_cols += 1
        else:
            break

    for row_idx in range(len(df)):
        person = df.index[row_idx]
        cat = category_hours.get(person, {})
        excel_row = row_idx + 2

        for col_idx, ds in enumerate(date_strs):
            excel_col = col_idx + stat_cols + 2
            cell = worksheet.cell(row=excel_row, column=excel_col)
            shift_val = str(cell.value or "")

            if not shift_val or shift_val == "None":
                continue

            # 边框编码 (无背景色)
            if "[备班]" in shift_val and "L/N" not in shift_val:
                cell.font = Font(color="C00000")  # 深红 - 备班
            elif "[L/N]" in shift_val:
                cell.font = Font(bold=True, color="CC6600")  # 橙色粗体 - L/N
            elif "[20%]" in shift_val:
                cell.font = Font(color="1a73e8")  # 蓝色 - 20%
            # 80% → 默认黑色，不加样式


# ==========================================
# 9. Dashboard Web 仪表盘 (Phase 9)
# ==========================================
# 9. Dashboard Web 仪表盘 (Phase 9)
# ==========================================


def generate_dashboard_html(final_schedule, final_hours, category_hours, hourly_hc,
                             date_strs, staff, oncall_schedule, us_notes,
                             pto_dates, ot_hours, output_path):
    """生成独立HTML仪表盘 (V3: 含分类工时面板)"""
    print("\n" + "="*60)
    print("🌐 Phase 9: 生成Web仪表盘 (V3)")
    print("="*60)

    def _build_role_data(role_name, members, role_key):
        staff_list = []
        for name in members:
            display = DISPLAY_NAME.get(name, name)
            person_schedule = {}
            person_cat = {}
            for ds in date_strs:
                entry = final_schedule.get(name, {}).get(ds, None)
                if entry and isinstance(entry, tuple):
                    person_schedule[ds] = entry[0]
                    person_cat[ds] = entry[1]
                elif entry:
                    person_schedule[ds] = str(entry)
                    person_cat[ds] = ""
                else:
                    person_schedule[ds] = ""
                    person_cat[ds] = ""

            is_oncall_today = {}
            for ds in date_strs:
                is_oncall_today[ds] = ds in oncall_schedule.get(name, {})
            person_notes = {}
            for ds in date_strs:
                if role_name == 'B超医生' and ds in us_notes.get(name, {}):
                    person_notes[ds] = us_notes[name][ds]
            person_pto = {}
            for ds in date_strs:
                if ds in pto_dates.get(name, set()):
                    person_pto[ds] = True

            cat = category_hours.get(name, {"80%": 0, "20%": 0, "备班": 0, "L/N": 0})
            staff_list.append({
                "name": display,
                "internal_name": name,
                "hours": round(final_hours.get(name, 0), 1),
                "hours_ot": round(ot_hours.get(name, 0), 1),
                "target": TARGET_HOURS_FULL if not name.startswith("备班") else 0,
                "hours_80": round(cat.get("80%", 0), 1),
                "hours_20": round(cat.get("20%", 0), 1),
                "hours_backup": round(cat.get("备班", 0), 1),
                "hours_ln": round(cat.get("L/N", 0), 1),
                "is_backup": name.startswith("备班"),
                "oncall_count": 0 if role_name == '放射医生' else len(oncall_schedule.get(name, {})),
                "schedule": person_schedule,
                "category": person_cat,
                "notes": person_notes if role_name == 'B超医生' else {},
                "pto": person_pto,
                "is_oncall": is_oncall_today,
            })

        demand_samples = {}
        for ds in date_strs[:7]:
            demand_samples[ds] = [float(v) for v in hourly_hc.get(ds, {}).get(role_key, np.zeros(24))]

        shift_colors = {
            "D": "#4CAF50", "D1": "#66BB6A", "D2": "#81C784", "D3": "#A5D6A7",
            "D4": "#43A047", "D5": "#388E3C", "D6": "#2E7D32",
            "C": "#4CAF50", "C1": "#66BB6A", "L": "#8BC34A",
            "H1": "#29B6F6", "H2": "#4FC3F7", "H3": "#81D4FA", "T": "#B3E5FC",
            "N": "#1a73e8", "N2": "#1565C0", "N3": "#0D47A1",
            "L/N": "#FF9800",
            "OnCall": "#9E9E9E",
            "off": "#F5F5F5",
        }

        return {
            "staff": staff_list,
            "dates": date_strs,
            "shift_colors": shift_colors,
            "shift_times": SHIFT_TIME_STR,
            "demand_samples": demand_samples,
            "total_days": len(date_strs),
        }

    rad_docs = staff['放射医生']['fulltime'] + staff['放射医生']['backup'] + staff['放射医生'].get('parttime', [])
    rad_techs = staff['放射技师']['fulltime'] + staff['放射技师'].get('parttime', []) + staff['放射技师']['backup']
    us_docs = staff['B超医生']['fulltime'] + staff['B超医生'].get('parttime', []) + staff['B超医生']['backup']
    if DUSTIN_US in final_schedule and DUSTIN_US not in us_docs:
        us_docs.append(DUSTIN_US)

    # B超 最终硬切（在生成HTML前最后执行）
    us_ft_names = [p for p in staff['B超医生']['fulltime'] if 'US' not in p]
    for d, ds in enumerate(date_strs):
        wd = d % 7
        day_target = 1 if wd == 6 else 2  # Sun=1, others=2 full_day minimum
        pm_target = 2 if wd in (2, 4) else (3 if wd in (1, 3) else None)

        fd_people = []   # 全天白班 (D/C/L)
        hx_people = []   # 半天班 (H1/H2/H3)
        off_people = []  # 休息
        for p_name in us_ft_names:
            entry = final_schedule.get(p_name, {}).get(ds, None)
            if not entry:
                off_people.append(p_name)
                continue
            sv = entry[0] if isinstance(entry, tuple) else str(entry)
            cat = entry[1] if isinstance(entry, tuple) else '80%'
            parts = set(s.strip() for s in sv.split(' + '))
            if parts & FULL_DAY_SHIFTS:
                fd_people.append((p_name, sv, cat))
            else:
                hx_people.append((p_name, sv, cat))

        # 1. 全天白班不足 → 从半天/休息中拉人
        need_fd = day_target - len(fd_people)
        if need_fd > 0:
            # promote H1/H2 → D
            for p_name, sv, cat in hx_people[:need_fd]:
                final_schedule.setdefault(p_name, {})[ds] = ('D', cat)
                fd_people.append((p_name, 'D', cat))
            # still need → pull off doctors
            need_fd = day_target - len(fd_people)
            for p_name in off_people[:need_fd]:
                final_schedule.setdefault(p_name, {})[ds] = ('D', '80%')
                fd_people.append((p_name, 'D', '80%'))

        # 2. PM 约束 (Tue-Fri)
        if pm_target is not None:
            # All full-day people are PM (they cover afternoon)
            pm_people = list(fd_people)
            # Also include H3 from half-day people
            for p_name, sv, cat in hx_people:
                if 'H3' in sv.split(' + '):
                    pm_people.append((p_name, sv, cat))
            # Cut excess PM to H2
            for p_name, sv, cat in pm_people[pm_target:]:
                final_schedule.setdefault(p_name, {})[ds] = ('H2', cat)
                # Remove them from pm_people too
                pm_people = pm_people[:pm_target]
            # 不足: promote H2→D
            need_pm = pm_target - len(pm_people[:pm_target])
            remaining_h2 = [(n, s, c) for n, s, c in hx_people if n not in [x[0] for x in pm_people[:pm_target]]]
            for p_name, sv, cat in remaining_h2[:need_pm]:
                final_schedule.setdefault(p_name, {})[ds] = ('D', cat)
            # still need → pull off
            still_off = [p for p in off_people if p not in [x[0] for x in pm_people]]
            need_pm = pm_target - len(pm_people[:pm_target])
            for p_name in still_off[:need_pm]:
                final_schedule.setdefault(p_name, {})[ds] = ('D', '80%')

    # 星期后缀 (JS渲染时附加)
    wd_names = ['一','二','三','四','五','六','日']

    data = {
        "month": f"{date_strs[0]} ~ {date_strs[-1]}",
        "roles": {
            "放射医生": _build_role_data("放射医生", rad_docs, "放射_放射医生"),
            "放射技师": _build_role_data("放射技师", rad_techs, "放射_放射技师"),
            "B超医生": _build_role_data("B超医生", us_docs, "超声_B超医生"),
        },
        "statistics": {
            "total_fulltime": sum(1 for n in final_hours if not n.startswith("备班") and final_hours.get(n, 0) > 0),
            "total_backup": sum(1 for n in final_hours if n.startswith("备班")),
            "total_backup_hours": round(sum(final_hours.get(n, 0) for n in final_hours if n.startswith("备班")), 1),
            "avg_fulltime_hours": round(np.mean([final_hours.get(n, 0) for n in final_hours if not n.startswith("备班") and final_hours.get(n, 0) > 0]), 1) if final_hours else 0,
            "total_80_hours": round(sum(category_hours.get(n, {}).get("80%", 0) for n in final_hours), 1),
            "total_20_hours": round(sum(category_hours.get(n, {}).get("20%", 0) for n in final_hours), 1),
            "total_ln_hours": round(sum(category_hours.get(n, {}).get("L/N", 0) for n in final_hours), 1),
        },
    }

    html = _render_dashboard_html_v3(data)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"✅ 仪表盘已保存: {output_path}")
    return output_path


def _render_dashboard_html_v3(data):
    """渲染统一排班仪表盘 — 含密码保护 + 在线编辑 + 备注/需求 + V3分类面板"""
    json_data = json.dumps(data, ensure_ascii=False, default=str)

    html = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>月度排班表 - GZU</title>
<style>
/* === Password Gate === */
#pwdGate{position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.5);display:flex;align-items:center;justify-content:center;z-index:9999}
#pwdGate>div{background:#fff;border-radius:12px;padding:32px;box-shadow:0 8px 32px rgba(0,0,0,0.2);text-align:center;min-width:300px}
#pwdGate input{padding:10px 14px;border:1px solid #ddd;border-radius:6px;font-size:16px;width:100%;text-align:center;margin:12px 0}
#pwdGate button{padding:10px 28px;background:#1a73e8;color:#fff;border:none;border-radius:6px;font-size:14px;cursor:pointer}
#pwdGate .err{color:#c62828;font-size:12px;margin-top:6px}

/* === Base === */
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"Microsoft YaHei","SimHei",sans-serif;background:#f0f2f5;color:#333;min-height:100vh}
.header{background:linear-gradient(135deg,#1a73e8 0%,#0d47a1 100%);color:#fff;padding:18px 24px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px}
.header h1{font-size:20px}
.header .nav a{color:#fff;opacity:0.8;text-decoration:none;font-size:13px;margin-left:14px}

/* === Toolbar === */
.toolbar{display:flex;gap:10px;padding:12px 20px;align-items:center;flex-wrap:wrap;background:#fff;border-bottom:1px solid #eee}
.toolbar select{padding:6px 12px;border:1px solid #ddd;border-radius:6px;font-size:14px}
.toolbar button{padding:6px 14px;border:none;border-radius:6px;font-size:12px;font-weight:600;cursor:pointer}
.btn-edit{background:#fff3e0;color:#e65100;border:1px solid #ffcc80!important}
.btn-edit.active{background:#e65100;color:#fff}
.btn-save{background:#1a73e8;color:#fff;display:none}
.btn-save.show{display:inline-block}
.btn-export{background:#388E3C;color:#fff;display:none}
.btn-export.show{display:inline-block}
.btn-token{background:none;color:#999;font-size:11px!important;text-decoration:underline}
.token-area{display:none;padding:8px 20px;background:#fff}
.token-area.show{display:flex;gap:8px;align-items:center}
.token-area input{padding:6px 10px;border:1px solid #ddd;border-radius:4px;width:300px;font-size:13px}
.msg{font-size:12px;padding:2px 20px;background:#fff}
.msg.ok{color:#2e7d32}
.msg.err{color:#c62828}

/* === Stats Bar === */
.stats-bar{display:flex;gap:10px;padding:12px 20px;flex-wrap:wrap;background:#fff;justify-content:center}
.stat-card{background:#fff;border-radius:8px;padding:12px 20px;box-shadow:0 1px 4px rgba(0,0,0,0.08);text-align:center;min-width:110px}
.stat-card .value{font-size:22px;font-weight:700;color:#1a73e8}
.stat-card .label{font-size:11px;color:#666;margin-top:2px}
.stat-card.orange .value{color:#FF9800}
.stat-card.red .value{color:#F44336}
.stat-card.green .value{color:#4CAF50}
.stat-card.blue .value{color:#1565C0}

/* === Tabs === */
.tabs{display:flex;gap:4px;margin:0 20px 12px;background:#fff;border-radius:8px;padding:4px;box-shadow:0 1px 4px rgba(0,0,0,0.08)}
.tab{flex:1;text-align:center;padding:10px 20px;cursor:pointer;border-radius:6px;font-size:14px;font-weight:600;transition:all 0.2s;background:#fff;border:none;color:#333}
.tab:hover{background:#e8f0fe}
.tab.active{background:#1a73e8;color:#fff}

/* === Legend === */
.legend{display:flex;gap:8px;flex-wrap:wrap;padding:8px 20px;background:#fff;margin-bottom:8px;font-size:11px;align-items:center;box-shadow:0 1px 2px rgba(0,0,0,0.04)}
.legend-item{display:inline-flex;align-items:center;gap:3px;white-space:nowrap}
.legend-dot{width:12px;height:12px;border-radius:2px;display:inline-block;flex-shrink:0}
.cat-tag{padding:1px 5px;border-radius:3px;font-weight:600;font-size:10px}
.cat-80{background:#e0e0e0;color:#000}
.cat-20{background:#1a73e8;color:#fff}
.cat-bk{background:#F44336;color:#fff}
.cat-ln{background:#FF9800;color:#fff}

/* === Table === */
.roster-grid{background:#fff;border-radius:8px;padding:16px;margin:0 20px 12px;box-shadow:0 1px 4px rgba(0,0,0,0.08);overflow-x:auto}
.roster-grid h2{font-size:15px;margin:0 0 10px 0;color:#1a73e8}
table.schedule{border-collapse:collapse;width:max-content;min-width:100%;font-size:11px}
table.schedule th,table.schedule td{border:1px solid #e0e0e0;padding:4px 3px;text-align:center;white-space:nowrap}
table.schedule th{background:#f5f5f5;font-weight:600;position:sticky;top:0;z-index:2}
table.schedule .name-col{min-width:80px;position:sticky;left:0;background:#fff;z-index:1;font-weight:600;text-align:left;padding-left:6px}
table.schedule .stats-col{min-width:42px;font-size:10px}
table.schedule .shift-cell{font-size:10px;min-width:50px;transition:all 0.15s;border-radius:2px;position:relative;color:#000}
table.schedule .shift-cell.editable{cursor:pointer}
table.schedule .shift-cell.editable:hover{transform:scale(1.1);z-index:3;box-shadow:0 2px 8px rgba(0,0,0,0.25)}
table.schedule .shift-cell.cat-backup{border:2px dashed #F44336!important}
table.schedule .shift-cell.cat-ln{border:3px solid #FF9800!important;font-weight:bold}
table.schedule .shift-cell.cat-20{background:#E3F2FD!important;border:2px solid #1565C0!important;font-weight:bold}
table.schedule .shift-cell.cell-oncall::after{content:"📞";position:absolute;top:0;right:1px;font-size:7px;line-height:1}

/* === PTO === */
.pto-cell{background:#FF4444!important;color:#fff!important;font-weight:bold}

/* === Popup === */
.overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.4);z-index:999}
.overlay.active{display:block}
.popup{display:none;position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);background:#fff;border-radius:12px;padding:20px;box-shadow:0 8px 32px rgba(0,0,0,0.2);z-index:1000;max-width:600px;width:90%;max-height:85vh;overflow-y:auto}
.popup.active{display:block}
.popup h3{margin:0 0 10px 0;color:#1a73e8;font-size:15px}
.popup table{width:100%;font-size:12px;border-collapse:collapse;margin-bottom:10px}
.popup table td,.popup table th{padding:4px 6px;border-bottom:1px solid #eee;text-align:left}
.popup .btn-row{display:flex;gap:8px;flex-wrap:wrap;margin:8px 0}
.popup .btn-row button{padding:6px 12px;border:1px solid #ddd;border-radius:6px;font-size:11px;font-weight:600;cursor:pointer;background:#fff}
.popup .btn-row button:hover{background:#f0f0f0}
.popup .btn-row button.sel{background:#1a73e8;color:#fff;border-color:#1a73e8}
.popup .btn-save-changes{background:#1a73e8!important;color:#fff!important;padding:8px 18px!important;font-size:13px!important}

/* === Chart === */
.chart-container{width:100%;height:260px;margin-top:10px}

/* === Shift Reference === */
.shift-ref{margin:0 20px 16px}
.shift-ref .section-title{font-size:14px;font-weight:700;color:#1a73e8;padding:8px 12px;background:#e8f0fe;border-radius:8px;margin-bottom:8px;border-left:4px solid #1a73e8}
.shift-grid{display:flex;flex-wrap:wrap;gap:6px}
.shift-chip{display:flex;align-items:center;gap:6px;padding:5px 10px;border-radius:8px;font-size:11px;font-weight:600;min-width:100px;box-shadow:0 1px 2px rgba(0,0,0,0.06)}
.shift-chip .chip-name{color:#fff;padding:2px 7px;border-radius:4px;font-size:12px;min-width:36px;text-align:center}
.shift-chip .chip-time{color:#555;font-size:10px}
.shift-chip.chip-off{background:#f5f5f5}

/* === Notes / Requirements === */
.section-title{font-size:14px;font-weight:700;color:#1a73e8;padding:8px 12px;background:#e8f0fe;border-radius:8px;margin-bottom:8px;border-left:4px solid #1a73e8}
.notes-section{margin:0 20px 16px}
.notes-toolbar{display:flex;gap:8px;margin-bottom:8px}
.notes-toolbar button{padding:5px 14px;border:none;border-radius:6px;font-size:12px;font-weight:600;cursor:pointer}
.btn-add-note{background:#1a73e8;color:#fff}
.notes-list{display:flex;flex-direction:column;gap:6px}
.note-card{background:#fff;border-radius:8px;padding:10px 14px;box-shadow:0 1px 3px rgba(0,0,0,0.06);position:relative}
.note-card .note-meta{font-size:10px;color:#999;margin-bottom:4px}
.note-card .note-text{font-size:12px;color:#333;line-height:1.5;white-space:pre-wrap}
.note-card .note-actions{position:absolute;top:8px;right:10px;display:flex;gap:4px}
.note-card .note-actions button{font-size:10px;padding:1px 6px;border:1px solid #ddd;border-radius:4px;background:#fff;cursor:pointer}
.note-editor{display:none;background:#fff;border-radius:8px;padding:12px;box-shadow:0 2px 8px rgba(0,0,0,0.1);margin-bottom:8px}
.note-editor.show{display:block}
.note-editor textarea{width:100%;min-height:60px;border:1px solid #ddd;border-radius:6px;padding:8px;font-size:12px;font-family:inherit;resize:vertical}
.note-editor .editor-actions{display:flex;gap:6px;margin-top:6px;justify-content:flex-end}
.note-editor .editor-actions button{padding:4px 12px;border:none;border-radius:6px;font-size:11px;cursor:pointer}

.footer{text-align:center;padding:16px;font-size:11px;color:#999}

@media(max-width:768px){table.schedule{font-size:9px}table.schedule td,table.schedule th{padding:2px 1px}}
</style>
</head>
<body>

<!-- === PASSWORD GATE === -->
<div id="pwdGate"><div>
    <h2>月度排班表</h2>
    <p style="color:#666;font-size:13px;margin:8px 0">请输入访问密码</p>
    <input type="password" id="pwdInput" placeholder="输入密码" onkeydown="if(event.key=='Enter')checkPwd()">
    <button onclick="checkPwd()">确认</button>
    <div class="err" id="pwdErr"></div>
</div></div>

<div id="mainContent" style="display:none">

<div class="header">
    <h1>月度排班表 <span id="hdMonth"></span></h1>
    <div class="nav">
        <select id="monthSelect" onchange="switchMonth(this.value)" style="padding:4px 8px;border-radius:4px;border:1px solid rgba(255,255,255,0.3);background:rgba(255,255,255,0.15);color:#fff;font-size:12px;margin-right:10px">
            <option value="">切换月份...</option>
        </select>
        <a href="index.html">首页</a>
    </div>
</div>

<!-- === TOOLBAR === -->
<div class="toolbar">
    <select id="roleSelect" onchange="switchRole()">
        <option>放射医生</option>
        <option>放射技师</option>
        <option>B超医生</option>
    </select>
    <button class="btn-edit" id="btnEdit" onclick="toggleEdit()">编辑模式</button>
    <button class="btn-save" id="btnSave" onclick="saveChanges()">保存修改</button>
    <button class="btn-export" id="btnExport" onclick="exportExcel()">📥 导出Excel</button>
    <button class="btn-token" onclick="toggleToken()">设置Token</button>
</div>
<div class="token-area" id="tokenArea">
    <span style="font-size:11px;color:#666">GitHub Token:</span>
    <input type="password" id="tokenInput" placeholder="ghp_xxxxxxxxxxxx">
    <button onclick="saveToken()" style="padding:5px 10px;font-size:11px">保存</button>
</div>

<div class="msg" id="msg"></div>

<!-- === STATS BAR === -->
<div class="stats-bar">
    <div class="stat-card"><div class="value">__STAT_FT__</div><div class="label">全职人员</div></div>
    <div class="stat-card"><div class="value">__STAT_AVG__h</div><div class="label">全职人均工时</div></div>
    <div class="stat-card green"><div class="value">__STAT_80__h</div><div class="label">80%池</div></div>
    <div class="stat-card blue"><div class="value">__STAT_20__h</div><div class="label">20%池</div></div>
    <div class="stat-card orange"><div class="value">__STAT_LN__h</div><div class="label">L/N</div></div>
    <div class="stat-card red"><div class="value">__STAT_BU_H__h</div><div class="label">备班</div></div>
</div>

<!-- === TABS === -->
<div class="tabs" id="tabs"></div>

<!-- === LEGEND === -->
<div class="legend" id="legendBar">
    <b>图例：</b><span id="legendContent"></span>
</div>
<div class="legend">
    <b>分类：</b>
    <span class="cat-tag cat-80">80%池</span> 黑色 |
    <span class="cat-tag cat-20">20%池</span> 蓝色框 |
    <span class="cat-tag cat-ln">L/N</span> 橙色粗框 |
    <span class="cat-tag cat-bk">备班</span> 红色虚线框
</div>

<!-- === ROSTER === -->
<div class="roster-grid" id="roster"><div style="text-align:center;padding:30px;color:#999">加载中...</div></div>

<!-- === SHIFT REFERENCE === -->
<div class="shift-ref">
    <div class="section-title">班型参考</div>
    <div id="shiftRef"></div>
</div>

<!-- === NOTES === -->
<div class="notes-section">
    <div class="section-title">部署使用说明</div>
    <div class="notes-toolbar">
        <button class="btn-add-note" onclick="addNote()">+ 新增说明</button>
    </div>
    <div class="note-editor" id="noteEditor">
        <textarea id="noteText" placeholder="输入说明内容..."></textarea>
        <div class="editor-actions">
            <button onclick="cancelNote()" style="background:#eee;color:#666">取消</button>
            <button onclick="saveNote()" style="background:#1a73e8;color:#fff">保存</button>
        </div>
    </div>
    <div class="notes-list" id="notesList"></div>
</div>

<!-- === REQUIREMENTS === -->
<div class="notes-section">
    <div class="section-title">需求收集区</div>
    <div class="notes-toolbar">
        <button class="btn-add-note" onclick="addReq()">+ 新增需求</button>
    </div>
    <div class="note-editor" id="reqEditor">
        <textarea id="reqText" placeholder="输入需求内容..."></textarea>
        <div class="editor-actions">
            <button onclick="cancelReq()" style="background:#eee;color:#666">取消</button>
            <button onclick="saveReq()" style="background:#1a73e8;color:#fff">保存</button>
        </div>
    </div>
    <div class="notes-list" id="reqsList"></div>
</div>

</div><!-- end mainContent -->

<div class="footer">数据来源：schedule.py CP-SAT排班引擎 | 本地手动运行，月底推送更新</div>

<div class="overlay" id="overlay" onclick="closePopup()"></div>
<div class="popup" id="popup"></div>

<script src="https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js"></script>
<script>var AVAILABLE_MONTHS = ['2026-06', '2026-07'];</script>
<script>
// ============ DATA ============
var SCHEDULE_DATA = __JSON_DATA__;
var TARGET_HOURS = __TARGET_FULL__;
var TARGET_80 = __TARGET_80__;
var currentRole = '放射医生';
var editMode = false;
var edits = {};
var selectedCell = null;

// ============ PASSWORD ============
var PWD_HASH = 'Z3p1MjAyNg=='; // gzu2026
function checkPwd(){
    var inp = document.getElementById('pwdInput').value;
    if(btoa(inp) === PWD_HASH){
        document.getElementById('pwdGate').style.display = 'none';
        document.getElementById('mainContent').style.display = 'block';
        sessionStorage.setItem('_gzu_sched', '1');
        init();
    } else {
        document.getElementById('pwdErr').textContent = '密码错误';
    }
}
if(sessionStorage.getItem('_gzu_sched') === '1'){
    document.getElementById('pwdGate').style.display = 'none';
    document.getElementById('mainContent').style.display = 'block';
    window.addEventListener('DOMContentLoaded', init);
}

// ============ INIT ============
function init(){
    document.getElementById('hdMonth').textContent = SCHEDULE_DATA.month || '';
    loadMonthSelector();
    var tabs = document.getElementById('tabs');
    var roles = Object.keys(SCHEDULE_DATA.roles);
    for(var i=0; i<roles.length; i++){
        (function(role){
            var btn = document.createElement('button');
            btn.className = 'tab' + (role === currentRole ? ' active' : '');
            btn.textContent = role;
            btn.onclick = function(){ currentRole = role; document.getElementById('roleSelect').value = role; switchRole(); };
            tabs.appendChild(btn);
        })(roles[i]);
    }
    document.getElementById('roleSelect').value = currentRole;
    renderAll();
    loadNotes();
    loadReqs();
}

function loadMonthSelector(){
    var sel = document.getElementById('monthSelect');
    if(!sel) return;
    // Scan for schedule_YYYY-MM.html archives (generated by schedule.py)
    var months = window.AVAILABLE_MONTHS || []; /* populated by Python at generation time */
    if(months.length === 0){
        // Fallback: probe via fetch
        var now = new Date(); var y=now.getFullYear(); var m=now.getMonth()+1;
        for(var i=0;i<12;i++){
            var ym=y+'-'+String(m).padStart(2,'0');
            (function(ymL,yr,mn){var xhr=new XMLHttpRequest();
                xhr.open('HEAD','schedule_'+ymL+'.html?_='+Date.now(),true);
                xhr.onload=function(){if(xhr.status<400)months.push(ymL);months.sort().reverse();renderOptions(sel,months);};
                xhr.send();
            })(ym);
            m--; if(m===0){m=12;y--;}
        }
    } else {
        months.sort().reverse();
        renderOptions(sel, months);
    }
}
function renderOptions(sel, months){
    sel.innerHTML = '<option value=\"\">切换月份...</option>';
    for(var i=0;i<months.length;i++){
        var o=document.createElement('option');
        o.value='schedule_'+months[i]+'.html';
        o.textContent=months[i];
        sel.appendChild(o);
    }
}
function switchMonth(url){ if(url) window.location.href = url; }

function switchRole(){
    currentRole = document.getElementById('roleSelect').value;
    var tabEls = document.querySelectorAll('.tab');
    for(var i=0; i<tabEls.length; i++){
        tabEls[i].classList.toggle('active', tabEls[i].textContent === currentRole);
    }
    renderAll();
}

// ============ RENDER ============
function renderAll(){
    renderLegend();
    renderRoster();
    renderShiftRef();
}

function renderLegend(){
    var rd = SCHEDULE_DATA.roles[currentRole];
    var colors = rd.shift_colors;
    var h = '';
    var shown = {};
    for(var k in colors){
        if(shown[k] || k === 'off') continue;
        shown[k] = true;
        var time = rd.shift_times[k] || '';
        h += '<span class="legend-item"><span class="legend-dot" style="background:' + colors[k] + '"></span> ' + k + (time?' ('+time+')':'') + '</span> ';
    }
    h += '<span class="legend-item"><span class="legend-dot" style="background:#F5F5F5;border:1px solid #ddd"></span> 休息</span> ';
    h += '<span class="legend-item" style="color:#9E9E9E">📞 OnCall</span>';
    document.getElementById('legendContent').innerHTML = h;
}

function renderRoster(){
    var rd = SCHEDULE_DATA.roles[currentRole];
    var dates = rd.dates;

    // 拆分: 全职+兼职 (主表) vs 备班 (独立section)
    var regular = [];
    var backup = [];
    for(var si=0; si<rd.staff.length; si++){
        var p = rd.staff[si];
        if(p.is_backup || (''+p.name).indexOf('备班') >= 0){
            backup.push(p);
        } else {
            regular.push(p);
        }
    }

    function staffRow(p){
        var html = '<tr>';
        html += '<td class="name-col">' + p.name + (p.is_backup?' 🔄':'') + '</td>';
        html += '<td class="stats-col" style="' + (p.hours > TARGET_80 && !p.is_backup?'color:#1a73e8;font-weight:600':'') + '">' + p.hours + 'h</td>';
        html += '<td class="stats-col">' + p.hours_80 + 'h</td>';
        html += '<td class="stats-col" style="color:#1a73e8">' + (p.hours_20>0?p.hours_20+'h':'-') + '</td>';
        html += '<td class="stats-col" style="color:#F44336">' + (p.hours_backup>0?p.hours_backup+'h':'-') + '</td>';
        html += '<td class="stats-col" style="color:#FF9800;font-weight:600">' + (p.hours_ln>0?p.hours_ln+'h':'-') + '</td>';
        html += '<td class="stats-col">' + (p.target>0?p.target+'h':'-') + '</td>';
        html += '<td class="stats-col">' + (p.oncall_count||'') + '</td>';
        for(var di=0; di<dates.length; di++){
            var ds = dates[di];
            var shiftVal = p.schedule[ds] || '';
            var catVal = p.category[ds] || '';
            var isOncall = p.is_oncall && p.is_oncall[ds];
            var hasEdit = edits[p.internal_name] && edits[p.internal_name][ds] !== undefined;
            var displayShift = hasEdit ? edits[p.internal_name][ds] : shiftVal;
            var text = displayShift || '-';
            var extraClass = editMode ? ' editable' : '';
            var bg = displayShift ? '#F8F8F8' : '#FFFFFF';
            if(displayShift){
                if(catVal === '备班') extraClass += ' cat-backup';
                else if(catVal === 'L/N') extraClass += ' cat-ln';
                else if(catVal === '20%') extraClass += ' cat-20';
                var badge = '';
                if(catVal === '20%') badge = ' <sup style="background:#1565C0;color:#fff;padding:2px 4px;border-radius:3px;font-size:10px;font-weight:bold">20%</sup>';
                else if(catVal === '备班') badge = ' <sup style="background:#F44336;color:#fff;padding:1px 3px;border-radius:2px;font-size:8px">B</sup>';
                else if(catVal === 'L/N') badge = ' <sup style="background:#FF9800;color:#fff;padding:1px 3px;border-radius:2px;font-size:8px">LN</sup>';
                text = '<b>' + displayShift + '</b>' + badge;
            }
            if(p.pto && p.pto[ds]){
                html += '<td class="shift-cell cell-pto" style="background:#FF4444;color:#fff;font-weight:bold"><span>PTO</span></td>';
                continue;
            }
            if(isOncall){ text += ' 📞'; extraClass += ' cell-oncall'; }
            if(p.notes && p.notes[ds]){
                text += ' [' + p.notes[ds] + ']';
            }
            var onclick = editMode ? ('onclick="openEditPopup(\'' + p.internal_name + '\',\'' + ds + '\')"') : ('onclick="showDayDetail(\'' + ds + '\')"');
            html += '<td class="shift-cell' + extraClass + '" style="background:' + bg + '" ' + onclick + ' title="' + ds + ': ' + (displayShift||'休息') + ' [' + (catVal||'-') + ']">' + text + '</td>';
        }
        html += '</tr>';
        return html;
    }

    var header = '<th class="name-col">人员</th><th class="stats-col">工时</th><th class="stats-col">80%</th><th class="stats-col">20%</th><th class="stats-col">备班</th><th class="stats-col">L/N</th><th class="stats-col">目标</th><th class="stats-col">OnCall</th>';
    var wk=['一','二','三','四','五','六','日'];
    for(var di=0; di<dates.length; di++){ header += '<th>' + dates[di].slice(3) + '<br><small>' + wk[di%7] + '</small></th>'; }

    var h = '<h2>' + currentRole + ' 排班表</h2>';
    h += '<table class="schedule"><thead><tr>' + header + '</tr></thead><tbody>';
    for(var si=0; si<regular.length; si++){ h += staffRow(regular[si]); }
    h += '</tbody></table>';

    // === 备班独立 Section ===
    if(backup.length > 0){
        h += '<div class="roster-grid" style="margin-top:20px;border-top:3px solid #F44336">';
        h += '<h2 style="color:#F44336">' + currentRole + ' — 备班</h2>';
        h += '<table class="schedule"><thead><tr>' + header + '</tr></thead><tbody>';
        for(var si=0; si<backup.length; si++){ h += staffRow(backup[si]); }
        h += '</tbody></table></div>';
    }

    document.getElementById('roster').innerHTML = h;
}

function renderShiftRef(){
    var rd = SCHEDULE_DATA.roles[currentRole];
    var h = '<div class="shift-grid">';
    var shiftNames = ['D','D1','D2','D3','D4','D5','D6','C','C1','L','N','N1','N2','N3','L/N','T','T1','H','H1','H2','H3'];
    for(var i=0; i<shiftNames.length; i++){
        var s = shiftNames[i];
        var time = rd.shift_times[s];
        var color = rd.shift_colors[s];
        if(!time && !color) continue;
        var bg = color || '#9e9e9e';
        h += '<div class="shift-chip"><span class="chip-name" style="background:' + bg + '">' + s + '</span><span class="chip-time">' + (time||'-') + '</span></div>';
    }
    h += '<div class="shift-chip chip-off"><span class="chip-name" style="background:#eee;color:#999">OFF</span><span class="chip-time">休息</span></div>';
    h += '</div>';
    document.getElementById('shiftRef').innerHTML = h;
}
function getShiftDesc(s){
    if(s=='L/N') return '24小时长班';
    if(s.indexOf('D')===0) return s==='D'?'白班8h':'白班变体';
    if(s.indexOf('C')===0) return '弹性班';
    if(s.indexOf('N')===0) return s==='N'?'夜班':'夜班变体';
    if(s.indexOf('H')===0) return '半天班';
    if(s=='T') return '教学班';
    if(s==='L') return '长班';
    return '';
}

// ============ DAY DETAIL POPUP (read-only) ============
function showDayDetail(dateStr){
    if(editMode) return;
    var rd = SCHEDULE_DATA.roles[currentRole];
    var h = '<h3>📅 ' + dateStr + ' — ' + currentRole + '</h3>';
    h += '<table><thead><tr><th>人员</th><th>班次</th><th>时间</th><th>分类</th></tr></thead><tbody>';
    var hasData = false;
    for(var i=0; i<rd.staff.length; i++){
        var p = rd.staff[i];
        var sv = p.schedule[dateStr] || '';
        if(sv){
            hasData = true;
            var parts = sv.split(' + ');
            var times = [];
            for(var k=0; k<parts.length; k++){
                var pp = parts[k].trim();
                if(pp.indexOf('备班(')>=0) times.push(pp);
                else times.push(rd.shift_times[pp] || pp);
            }
            var cv = p.category[dateStr] || '';
            var cd = cv;
            if(cv==='20%') cd='<span style="color:#1a73e8;font-weight:600">20%池</span>';
            else if(cv==='备班') cd='<span style="color:#F44336;font-weight:600">备班</span>';
            else if(cv==='L/N') cd='<span style="color:#FF9800;font-weight:600">L/N</span>';
            h += '<tr><td>' + p.name + '</td><td>' + sv + '</td><td>' + times.join(' + ') + '</td><td>' + cd + '</td></tr>';
        }
    }
    if(!hasData) h += '<tr><td colspan="4" style="text-align:center;color:#999">当日无人排班</td></tr>';
    h += '</tbody></table>';
    h += '<div class="chart-container" id="dayChart"></div>';
    document.getElementById('popup').innerHTML = h;
    openOverlay();
    // Demand chart
    setTimeout(function(){
        var demand = rd.demand_samples[dateStr] || [];
        if(demand.length>0){
            var cd = document.getElementById('dayChart');
            if(cd && typeof echarts !== 'undefined'){
                var chart = echarts.init(cd);
                var hours = []; for(var hh=0; hh<24; hh++) hours.push(hh+':00');
                chart.setOption({
                    title:{text:dateStr+' 需求HC',textStyle:{fontSize:12}},
                    tooltip:{trigger:'axis'},
                    xAxis:{data:hours,axisLabel:{rotate:45,fontSize:9}},
                    yAxis:{name:'人数',minInterval:1},
                    series:[{name:'需求HC',type:'bar',data:demand,itemStyle:{color:'#1a73e8'}}],
                    grid:{left:40,right:15,top:35,bottom:45}
                });
            }
        }
    },200);
}

// ============ EDIT MODE ============
function toggleEdit(){
    var btn = document.getElementById('btnEdit');
    var saveBtn = document.getElementById('btnSave');
    var exportBtn = document.getElementById('btnExport');
    editMode = !editMode;
    if(editMode){
        btn.textContent = '退出编辑';
        btn.classList.add('active');
        saveBtn.classList.add('show');
        if(exportBtn) exportBtn.classList.add('show');
    } else {
        btn.textContent = '编辑模式';
        btn.classList.remove('active');
        saveBtn.classList.remove('show');
        if(exportBtn) exportBtn.classList.remove('show');
    }
    renderRoster();
}

var ALLOWED_SHIFTS = ['C','C1','D1','D2','D3','D4','D5','D6','N','N1','N2','N3','L','L/N','T','T1','H','H1','H2','H3','PTO','CTO','OFF',''];
var SHIFT_LABELS = {OFF:'休息'};

function openEditPopup(personName, dateStr){
    selectedCell = {name: personName, ds: dateStr};
    var rd = SCHEDULE_DATA.roles[currentRole];
    var person = null;
    for(var i=0; i<rd.staff.length; i++){ if(rd.staff[i].internal_name === personName){ person = rd.staff[i]; break; } }
    if(!person) return;
    var currentShift = edits[personName]&&edits[personName][dateStr]!==undefined ? edits[personName][dateStr] : (person.schedule[dateStr] || '');

    var h = '<h3>✏️ ' + person.name + ' — ' + dateStr + '</h3>';
    h += '<p style="font-size:11px;color:#666;margin-bottom:8px">当前班次：<b>' + (currentShift||'休息') + '</b></p>';
    h += '<div class="btn-row">';
    for(var i=0; i<ALLOWED_SHIFTS.length; i++){
        var s = ALLOWED_SHIFTS[i];
        var sel = s === currentShift ? ' sel' : '';
        var label = SHIFT_LABELS[s] || s;
        h += '<button class="btn-shift' + sel + '" onclick="selectShift(\'' + s + '\')">' + label + '</button>';
    }
    h += '</div>';
    h += '<button class="btn-save-changes" onclick="applyEdit()">确认修改</button>';
    document.getElementById('popup').innerHTML = h;
    openOverlay();
}

function selectShift(shift){
    var btns = document.querySelectorAll('.btn-shift');
    for(var i=0; i<btns.length; i++){ btns[i].classList.remove('sel'); }
    var btns2 = document.querySelectorAll('.btn-shift');
    for(var i=0; i<btns2.length; i++){
        if(btns2[i].textContent === shift || (shift==='OFF' && btns2[i].textContent==='休息')){
            btns2[i].classList.add('sel');
        }
    }
    if(!edits[selectedCell.name]) edits[selectedCell.name] = {};
    edits[selectedCell.name][selectedCell.ds] = shift;
    renderRoster();
    closePopup();
}

function applyEdit(){
    closePopup();
    renderRoster();
}

function openOverlay(){
    document.getElementById('overlay').classList.add('active');
    document.getElementById('popup').classList.add('active');
}
function closePopup(){
    document.getElementById('overlay').classList.remove('active');
    document.getElementById('popup').classList.remove('active');
}
document.addEventListener('keydown', function(e){ if(e.key==='Escape') closePopup(); });

// ============ SAVE TO GITHUB ============
function toggleToken(){
    document.getElementById('tokenArea').classList.toggle('show');
}
function saveToken(){
    var t = document.getElementById('tokenInput').value.trim();
    if(t) localStorage.setItem('gh_token', t);
    msg('Token已保存', 'ok');
}

async function saveChanges(){
    var token = localStorage.getItem('gh_token');
    if(!token){ msg('请先设置GitHub Token', 'err'); toggleToken(); return; }
    var count = 0;
    for(var p in edits){ for(var d in edits[p]){ count++; } }
    if(count === 0){ msg('没有修改', 'err'); return; }

    // Update SCHEDULE_DATA with edits
    var rd = SCHEDULE_DATA.roles[currentRole];
    for(var si=0; si<rd.staff.length; si++){
        var person = rd.staff[si];
        if(edits[person.internal_name]){
            for(var ds in edits[person.internal_name]){
                person.schedule[ds] = edits[person.internal_name][ds];
                person.category[ds] = '';
            }
        }
    }

    var jsonStr = JSON.stringify(SCHEDULE_DATA, null, 2);
    // UTF-8 → base64: encodeURIComponent handles all Unicode, then convert %XX→bytes
    var b64 = btoa(encodeURIComponent(jsonStr).replace(/%([0-9A-F]{2})/g, function(m, p) {
        return String.fromCharCode(parseInt(p, 16));
    }));
    var path = 'publish/schedule_data.json';
    try{
        var getUrl = 'https://api.github.com/repos/liuyx339-oss/gzu-schedule/contents/' + path;
        var resp = await fetch(getUrl, {headers: {'Authorization': 'token ' + token}});
        var d = await resp.json();
        var body = {message: 'Update schedule', content: b64, branch: 'master'};
        if(d.sha) body.sha = d.sha;
        var putResp = await fetch(getUrl, {method: 'PUT', headers: {'Authorization': 'token ' + token, 'Content-Type': 'application/json'}, body: JSON.stringify(body)});
        var r = await putResp.json();
        if(r.content){
            msg('已保存 ' + count + ' 处修改', 'ok');
            edits = {};
            renderRoster();
        } else {
            msg('保存失败 HTTP' + putResp.status + ': ' + (r.message||'?'), 'err');
        }
    } catch(e){ msg('网络错误 ' + e.message, 'err'); }
}

function exportExcel(){
    // Apply edits to get full schedule
    for(var p in edits){
        for(var d in edits[p]){
            for(var si=0; si<SCHEDULE_DATA.roles[currentRole].staff.length; si++){
                if(SCHEDULE_DATA.roles[currentRole].staff[si].internal_name === p){
                    SCHEDULE_DATA.roles[currentRole].staff[si].schedule[d] = edits[p][d];
                }
            }
        }
    }
    var rd = SCHEDULE_DATA.roles[currentRole];
    var dates = rd.dates;
    // Build CSV: name, stats, then each date
    var rows = [];
    var header = ['人员','总工时','80%','20%','备班','L/N','OT','目标','OnCall'];
    for(var i=0; i<dates.length; i++) header.push(dates[i]);
    rows.push(header);
    for(var si=0; si<rd.staff.length; si++){
        var p = rd.staff[si];
        var row = [p.name, p.hours, p.hours_80, p.hours_20, p.hours_backup, p.hours_ln, p.hours_ot||0, p.target, p.oncall_count||''];
        for(var di=0; di<dates.length; di++){
            var ds = dates[di];
            var sv = (edits[p.internal_name]&&edits[p.internal_name][ds]!==undefined) ? edits[p.internal_name][ds] : (p.schedule[ds]||'');
            row.push(sv||'-');
        }
        rows.push(row);
    }
    // Convert to CSV string
    var csv = '\\uFEFF' + rows.map(function(r){ return r.map(function(c){ return '"' + String(c).replace(/"/g,'\"\"') + '"'; }).join(','); }).join('\\n');
    var blob = new Blob([csv], {type: 'text/csv;charset=utf-8'});
    var url = URL.createObjectURL(blob);
    var a = document.createElement('a');
    a.href = url;
    a.download = currentRole + '_' + (new Date().toISOString().slice(0,10)) + '.csv';
    a.click();
    URL.revokeObjectURL(url);
    msg('已导出 Excel (CSV): ' + a.download, 'ok');
}

function msg(text, cls){
    var el = document.getElementById('msg');
    el.textContent = text;
    el.className = 'msg ' + (cls||'');
    setTimeout(function(){ el.textContent = ''; el.className = 'msg'; }, 4000);
}

// ============ NOTES CRUD ============
var NOTES_PATH = 'publish/notes.json';
var notes = [];
var editingNoteId = null;

async function loadNotes(){
    try{ var r = await fetch('notes.json?' + Date.now()); if(r.ok) notes = await r.json(); else notes = []; }
    catch(e){ notes = []; }
    renderNotes();
}
function renderNotes(){
    var h = '';
    if(!notes.length){ h = '<div style="color:#999;font-size:12px;text-align:center;padding:12px">暂无说明，点击"+ 新增说明"添加</div>'; }
    for(var i=0; i<notes.length; i++){
        var n = notes[i];
        h += '<div class="note-card"><div class="note-meta">' + n.time + ' | #' + (notes.length-i) + '</div><div class="note-text">' + n.text + '</div><div class="note-actions"><button onclick="editNote(' + i + ')">编辑</button><button onclick="delNote(' + i + ')">删除</button></div></div>';
    }
    document.getElementById('notesList').innerHTML = h;
}
function addNote(){ editingNoteId = null; document.getElementById('noteText').value = ''; document.getElementById('noteEditor').classList.add('show'); }
function editNote(idx){ editingNoteId = idx; document.getElementById('noteText').value = notes[idx].text; document.getElementById('noteEditor').classList.add('show'); }
function cancelNote(){ document.getElementById('noteEditor').classList.remove('show'); }
async function saveNote(){
    var text = document.getElementById('noteText').value.trim(); if(!text) return;
    var token = localStorage.getItem('gh_token'); if(!token){ msg('请先设置GitHub Token', 'err'); return; }
    var now = new Date().toLocaleString('zh-CN');
    if(editingNoteId !== null){ notes[editingNoteId].text = text; notes[editingNoteId].time = now; }
    else{ notes.push({text:text, time:now}); }
    notes.sort(function(a,b){ return b.time.localeCompare(a.time); });
    var b64 = function(s){var b=new TextEncoder().encode(s);return function(ab){var c=0x8000,r=[];for(var i=0;i<ab.length;i+=c)r.push(String.fromCharCode.apply(null,ab.subarray(i,i+c)));return btoa(r.join(""))}(b)}(JSON.stringify(notes,null,2));
    try{
        var getUrl = 'https://api.github.com/repos/liuyx339-oss/gzu-schedule/contents/' + NOTES_PATH;
        var d = await (await fetch(getUrl, {headers:{'Authorization':'token '+token}})).json();
        var body = {message:'Update notes', content:b64, branch:'master'}; if(d.sha) body.sha = d.sha;
        var r = await (await fetch(getUrl, {method:'PUT', headers:{'Authorization':'token '+token,'Content-Type':'application/json'}, body:JSON.stringify(body)})).json();
        if(r.content){ renderNotes(); cancelNote(); msg('✅ 已保存', 'ok'); } else { msg('保存失败: '+(r.message||'?'), 'err'); }
    } catch(e){ msg('网络错误: '+e.message, 'err'); }
}
async function delNote(idx){
    if(!confirm('确定删除？')) return; notes.splice(idx,1);
    var token = localStorage.getItem('gh_token'); if(!token) return;
    var b64 = function(s){var b=new TextEncoder().encode(s);return function(ab){var c=0x8000,r=[];for(var i=0;i<ab.length;i+=c)r.push(String.fromCharCode.apply(null,ab.subarray(i,i+c)));return btoa(r.join(""))}(b)}(JSON.stringify(notes,null,2));
    try{
        var getUrl = 'https://api.github.com/repos/liuyx339-oss/gzu-schedule/contents/' + NOTES_PATH;
        var d = await (await fetch(getUrl, {headers:{'Authorization':'token '+token}})).json();
        var body = {message:'Delete note', content:b64, branch:'master'}; if(d.sha) body.sha = d.sha;
        await fetch(getUrl, {method:'PUT', headers:{'Authorization':'token '+token,'Content-Type':'application/json'}, body:JSON.stringify(body)});
        renderNotes();
    } catch(e){}
}

// ============ REQUIREMENTS CRUD ============
var REQS_PATH = 'publish/requirements.json';
var reqs = [];
var editingReqId = null;

async function loadReqs(){
    try{ var r = await fetch('requirements.json?' + Date.now()); if(r.ok) reqs = await r.json(); else reqs = []; }
    catch(e){ reqs = []; }
    renderReqs();
}
function renderReqs(){
    var h = '';
    if(!reqs.length){ h = '<div style="color:#999;font-size:12px;text-align:center;padding:12px">暂无需求，点击"+ 新增需求"添加</div>'; }
    for(var i=0; i<reqs.length; i++){
        var n = reqs[i];
        h += '<div class="note-card"><div class="note-meta">' + n.time + ' | #' + (reqs.length-i) + '</div><div class="note-text">' + n.text + '</div><div class="note-actions"><button onclick="editReq(' + i + ')">编辑</button><button onclick="delReq(' + i + ')">删除</button></div></div>';
    }
    document.getElementById('reqsList').innerHTML = h;
}
function addReq(){ editingReqId = null; document.getElementById('reqText').value = ''; document.getElementById('reqEditor').classList.add('show'); }
function editReq(idx){ editingReqId = idx; document.getElementById('reqText').value = reqs[idx].text; document.getElementById('reqEditor').classList.add('show'); }
function cancelReq(){ document.getElementById('reqEditor').classList.remove('show'); }
async function saveReq(){
    var text = document.getElementById('reqText').value.trim(); if(!text) return;
    var token = localStorage.getItem('gh_token'); if(!token){ msg('请先设置GitHub Token', 'err'); return; }
    var now = new Date().toLocaleString('zh-CN');
    if(editingReqId !== null){ reqs[editingReqId].text = text; reqs[editingReqId].time = now; }
    else{ reqs.push({text:text, time:now}); }
    reqs.sort(function(a,b){ return b.time.localeCompare(a.time); });
    var b64 = function(s){var b=new TextEncoder().encode(s);return function(ab){var c=0x8000,r=[];for(var i=0;i<ab.length;i+=c)r.push(String.fromCharCode.apply(null,ab.subarray(i,i+c)));return btoa(r.join(""))}(b)}(JSON.stringify(reqs,null,2));
    try{
        var getUrl = 'https://api.github.com/repos/liuyx339-oss/gzu-schedule/contents/' + REQS_PATH;
        var d = await (await fetch(getUrl, {headers:{'Authorization':'token '+token}})).json();
        var body = {message:'Update requirements', content:b64, branch:'master'}; if(d.sha) body.sha = d.sha;
        var r = await (await fetch(getUrl, {method:'PUT', headers:{'Authorization':'token '+token,'Content-Type':'application/json'}, body:JSON.stringify(body)})).json();
        if(r.content){ renderReqs(); cancelReq(); msg('✅ 已保存', 'ok'); } else { msg('保存失败: '+(r.message||'?'), 'err'); }
    } catch(e){ msg('网络错误: '+e.message, 'err'); }
}
async function delReq(idx){
    if(!confirm('确定删除？')) return; reqs.splice(idx,1);
    var token = localStorage.getItem('gh_token'); if(!token) return;
    var b64 = function(s){var b=new TextEncoder().encode(s);return function(ab){var c=0x8000,r=[];for(var i=0;i<ab.length;i+=c)r.push(String.fromCharCode.apply(null,ab.subarray(i,i+c)));return btoa(r.join(""))}(b)}(JSON.stringify(reqs,null,2));
    try{
        var getUrl = 'https://api.github.com/repos/liuyx339-oss/gzu-schedule/contents/' + REQS_PATH;
        var d = await (await fetch(getUrl, {headers:{'Authorization':'token '+token}})).json();
        var body = {message:'Delete requirement', content:b64, branch:'master'}; if(d.sha) body.sha = d.sha;
        await fetch(getUrl, {method:'PUT', headers:{'Authorization':'token '+token,'Content-Type':'application/json'}, body:JSON.stringify(body)});
        renderReqs();
    } catch(e){}
}
</script>
</body>
</html>'''

    replacements = [
        ('__MONTH__', str(data.get('month', ''))),
        ('__JSON_DATA__', json_data),
        ('__TARGET_FULL__', str(TARGET_HOURS_FULL)),
        ('__TARGET_80__', str(TARGET_HOURS_80)),
        ('__STAT_FT__', str(data['statistics']['total_fulltime'])),
        ('__STAT_AVG__', str(data['statistics']['avg_fulltime_hours'])),
        ('__STAT_80__', str(data['statistics']['total_80_hours'])),
        ('__STAT_20__', str(data['statistics']['total_20_hours'])),
        ('__STAT_LN__', str(data['statistics']['total_ln_hours'])),
        ('__STAT_BU_H__', str(data['statistics']['total_backup_hours'])),
    ]
    for placeholder, value in replacements:
        html = html.replace(placeholder, value)

    return html



# ==========================================
# 10. 主入口 (V3 Pipeline)
# ==========================================

def main():
    parser = argparse.ArgumentParser(description='放射/超声排班优化系统 V3')
    parser.add_argument('--month', type=str, default=None, help='排班月份 (YYYY-MM)')
    parser.add_argument('--solver-time', type=int, default=300, help='CP-SAT每阶段求解时间上限(秒)')
    parser.add_argument('--no-feishu', action='store_true', help='跳过飞书API')
    parser.add_argument('--output-dir', type=str, default=None, help='输出目录')
    args = parser.parse_args()

    base_dir = args.output_dir or os.path.dirname(os.path.abspath(__file__))

    print("\n" + "="*60)
    print("🚀 放射/超声 排班优化系统 V3")
    print("   三阶段分层: 80%池 → 20%池 → 备班池 (全覆盖)")
    print("="*60)

    # --- 确定月份 ---
    if args.month:
        parts = args.month.split('-')
        month_year = (int(parts[0]), int(parts[1]))
    else:
        # fallback: 读通用 CSV 推断月份
        fallback_csv = os.path.join(base_dir, "pipeline_output", "Demand_Forecast_Hourly.csv")
        df_tmp = pd.read_csv(fallback_csv, encoding="utf-8-sig")
        df_tmp['ds'] = pd.to_datetime(df_tmp['ds'], errors='coerce')
        df_tmp = df_tmp.dropna(subset=['ds'])
        month_year = (int(df_tmp['ds'].dt.year.iloc[0]), int(df_tmp['ds'].dt.month.iloc[0]))

    month_str = f"{month_year[0]}-{month_year[1]:02d}"
    monthly_csv_name = f"Demand_Forecast_{month_str}_Hourly.csv"
    data_path = os.path.join(base_dir, "pipeline_output", monthly_csv_name)
    if not os.path.exists(data_path):
        # fallback: 旧版通用文件名
        data_path = os.path.join(base_dir, "pipeline_output", "Demand_Forecast_Hourly.csv")
        if not os.path.exists(data_path):
            print(f"❌ 数据文件不存在: {monthly_csv_name}")
            print("   请先运行 prophet_lightGBM.py 生成预测数据")
            sys.exit(1)

    print(f"📅 排班月份: {month_str}")

    # --- 动态计算月目标工时 ---
    global TARGET_HOURS_FULL, TARGET_HOURS_80
    workdays = _count_workdays(month_year[0], month_year[1])
    TARGET_HOURS_FULL = float(workdays * 8)
    TARGET_HOURS_80 = round(TARGET_HOURS_FULL * 0.8, 1)
    print(f"🎯 月目标工时: {TARGET_HOURS_FULL:.0f}h ({workdays}个工作日×8h) | 80%阈值: {TARGET_HOURS_80}h")

    # --- 加载人员 ---
    staff_raw = None
    if not args.no_feishu:
        print("🔍 尝试从飞书加载人员...")
        staff_raw = load_staff_from_feishu()
    if staff_raw is None:
        staff_raw = STAFF_FALLBACK
        print("📋 使用fallback人员列表")
    if "Dustin Huang" not in staff_raw.get("rad_docs_full", []):
        staff_raw.setdefault("rad_docs_full", []).append("Dustin Huang")
    staff = build_staff(staff_raw)
    print(f"\n👥 人员配置:")
    print(f"   放射医生: 全职{staff['放射医生']['fulltime']}  兼职{staff['放射医生']['parttime']}  备班{staff['放射医生']['backup']}")
    print(f"   放射技师: 全职{staff['放射技师']['fulltime']}  兼职{staff['放射技师']['parttime']}  备班{staff['放射技师']['backup']}")
    print(f"   B超医生:  全职{staff['B超医生']['fulltime']}  兼职{staff['B超医生']['parttime']}  备班{staff['B超医生']['backup']}")

    # --- Phase 1: 数据预处理 ---
    hourly_hc, date_strs, all_dates = load_and_preprocess_demand(data_path, month_year)

    # --- Phase 2: L/N 预分配 ---
    ln_schedule, ln_hours, ln_skip_dates = pre_allocate_ln(
        hourly_hc, date_strs, staff, all_dates)

    # --- 加载请假数据 ---
    print(f"\n📅 加载请假数据...")
    leaves_raw = load_leaves_from_feishu()
    leave_constraints = _apply_leave_to_staff(staff, leaves_raw or {}, date_strs) if leaves_raw else {}
    if leave_constraints:
        print(f"   ✅ 请假约束已应用: {len(leave_constraints)} 人")
    # 构建PTO显示数据: {person: set of date_strs}
    pto_dates = defaultdict(set)
    if leaves_raw:
        for person, date_blocks in leaves_raw.items():
            pto_dates[person] = set(date_blocks.keys())

    # --- Phase 3a: Stage 1 — 80%工时池 (先放射医生+放射技师) ---
    stage1_schedule, stage1_hours = solve_stage1_80pct(
        hourly_hc, date_strs, staff, ln_schedule, ln_skip_dates, ln_hours, all_dates,
        roles=['放射医生', '放射技师'], leave_constraints=leave_constraints)

    # --- 收集Dustin放射班日期(用于超声抵扣) ---
    dustin_rad_dates = set()
    for ds, value in stage1_schedule.get(DUSTIN_RAD, {}).items():
        if value:
            shift_str = value[0] if isinstance(value, tuple) else value
            if shift_str:
                dustin_rad_dates.add(ds)
    print(f"\n📌 Dustin上放射班 {len(dustin_rad_dates)} 天 → 超声这些天HC -= 0.5")

    # 调整超声需求(Dustin抵扣)
    us_adjusted_hc = apply_dustin_us_deduction(hourly_hc, dustin_rad_dates, date_strs)

    # --- Phase 3b: Stage 1 — 80%工时池 (超声医生, 使用抵扣后需求) ---
    us_s1_schedule, us_s1_hours = solve_stage1_80pct(
        us_adjusted_hc, date_strs, staff, ln_schedule, ln_skip_dates, ln_hours, all_dates,
        roles=['B超医生'], leave_constraints=leave_constraints)

    # 合并Stage 1结果
    for person, date_shifts in us_s1_schedule.items():
        if person not in stage1_schedule:
            stage1_schedule[person] = {}
        for ds, value in date_shifts.items():
            stage1_schedule[person][ds] = value
    for person, hrs in us_s1_hours.items():
        stage1_hours[person] = hrs

    # --- Phase 4: Stage 2 — 20%工时池 (全部角色使用调整后HC) ---
    stage2_schedule, stage2_hours = solve_stage2_20pct(
        us_adjusted_hc, date_strs, staff, stage1_schedule, stage1_hours,
        ln_schedule, ln_skip_dates, all_dates, leave_constraints=leave_constraints)

    # --- Phase 5: Stage 3 — 备班池全覆盖 (硬约束!) ---
    stage3_schedule, stage3_hours = solve_stage3_backup(
        us_adjusted_hc, date_strs, staff,
        stage1_schedule, stage1_hours,
        stage2_schedule, stage2_hours,
        ln_schedule, ln_skip_dates, all_dates,
        leave_constraints=leave_constraints)

    # --- 将Dustin_US加入B超医生(用于合并/输出) ---
    if DUSTIN_US not in staff['B超医生']['fulltime']:
        staff['B超医生']['fulltime'].append(DUSTIN_US)

    # --- Phase 6: 合并排班 + OnCall分配 ---
    final_schedule, final_hours, category_hours, oncall_schedule, ot_hours = merge_and_oncall(
        stage1_schedule, stage1_hours,
        stage2_schedule, stage2_hours,
        stage3_schedule, stage3_hours,
        date_strs, staff, all_dates)

    # --- Phase 7: Dustin跨角色验证 ---
    dustin_us_available, _ = apply_dustin_cross_role(
        final_schedule, final_hours, category_hours, hourly_hc, date_strs, staff)

    # --- Phase 7B: 超声医生楼层备注 (4/9/B1) ---
    us_notes = assign_ultrasound_notes(final_schedule, date_strs, staff)

    # --- Phase 8: Excel ---
    month_str = f"{month_year[0]}-{month_year[1]:02d}"
    schedule_dir = os.path.join(base_dir, "pipeline_output", "schedule")
    os.makedirs(schedule_dir, exist_ok=True)
    xlsx_path = os.path.join(schedule_dir, f"Schedule_{month_str}_V3.xlsx")
    generate_excel(final_schedule, final_hours, category_hours, hourly_hc,
                   date_strs, staff, oncall_schedule, ot_hours, us_notes,
                   pto_dates, xlsx_path)

    # --- Phase 9: Dashboard ---
    # 输出到 pipeline_output/schedule/ (归档)
    html_path = os.path.join(schedule_dir, f"Schedule_Dashboard_{month_str}_V3.html")
    generate_dashboard_html(final_schedule, final_hours, category_hours, hourly_hc,
                            date_strs, staff, oncall_schedule, us_notes,
                            pto_dates, ot_hours, html_path)

    # 同时输出到 publish/schedule.html (GitHub Pages)
    publish_schedule_path = os.path.join(base_dir, "publish", "schedule.html")
    import shutil
    shutil.copy2(html_path, publish_schedule_path)
    # 归档副本: publish/schedule_2026-06.html (不覆盖历史月份)
    archive_path = os.path.join(base_dir, "publish", f"schedule_{month_str}.html")
    shutil.copy2(html_path, archive_path)
    print(f"   🌐 线上版本: {publish_schedule_path}")
    print(f"   📁 归档副本: {archive_path}")

    # --- 总结 ---
    print("\n" + "="*60)
    print("✅ 排班完成 (V3)！")
    print(f"   📊 Excel: {xlsx_path}")
    print(f"   🌐 仪表盘: {html_path}")
    print("="*60)

    print("\n📋 工时统计 (分类):")
    print("-"*100)
    for role in ['放射医生', '放射技师', 'B超医生']:
        print(f"\n  [{role}]")
        print(f"    {'人员':25} {'80%':8} {'20%':8} {'备班':8} {'L/N':8} {'OT':8} {'总工时':8} {'目标':6} {'OnCall':8}")
        print(f"    {'-'*85}")
        for person in staff[role]['fulltime'] + staff[role]['backup']:
            if person == DUSTIN_US:
                continue
            cat = category_hours.get(person, {"80%": 0, "20%": 0, "备班": 0, "L/N": 0})
            total = final_hours.get(person, 0)
            target = TARGET_HOURS_FULL if not person.startswith("备班") else 0
            ot = ot_hours.get(person, 0)
            oncall_n = len(oncall_schedule.get(person, {}))
            display = DISPLAY_NAME.get(person, person)
            ot_str = f"{ot:5.1f}h" if ot > 0 else "     -"
            print(f"    {display:25} {cat.get('80%',0):6.1f}h {cat.get('20%',0):6.1f}h {cat.get('备班',0):6.1f}h {cat.get('L/N',0):6.1f}h {ot_str} {total:6.1f}h {target:5.0f}h OnCall:{oncall_n}")

    # Dustin汇总
    dustin_total = final_hours.get(DUSTIN_RAD, 0) + final_hours.get(DUSTIN_US, 0)
    print(f"\n  🟡 Dustin 总计: 放射 {final_hours.get(DUSTIN_RAD, 0):.1f}h + 超声 {final_hours.get(DUSTIN_US, 0):.1f}h = {dustin_total:.1f}h / {TARGET_HOURS_FULL}h")


if __name__ == "__main__":
    main()

