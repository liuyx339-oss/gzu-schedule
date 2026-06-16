#!/usr/bin/env python3
"""
部署包生成器
=========================================
从排班 Excel 和 schedule.py 的输出中提取数据，
生成可交付给同事的 deploy_package/ 文件夹。

用法:
  python deploy.py                          # 自动查找最新排班文件
  python deploy.py --xlsx Schedule_2026-06_V3.xlsx  # 指定排班文件
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
from datetime import datetime

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEPLOY_DIR = os.path.join(SCRIPT_DIR, "deploy_package")

# ── 角色配置 ──────────────────────────────────────────────
ROLE_CONFIG = {
    "放射医生": {
        "sheet_name_pattern": "放射医生排班",
        "available_shifts": [
            "D", "D1", "D2", "D3", "D4", "D5", "D6",
            "C", "C1", "L",
            "H1", "H2", "H3", "T",
            "N", "N2", "N3", "L/N",
        ],
        "has_night": True,
    },
    "放射技师": {
        "sheet_name_pattern": "放射技师排班",
        "available_shifts": [
            "D", "D1", "D2", "D3", "D4", "D5", "D6",
            "C", "C1", "L",
            "H1", "H2", "H3", "T",
            "N", "N2", "N3", "L/N",
        ],
        "has_night": True,
    },
    "B超医生": {
        "sheet_name_pattern": "B超医生排班",
        "available_shifts": [
            "D", "D1", "D2", "D3", "D4", "D5", "D6",
            "C", "C1",
            "H1", "H2", "H3", "T",
        ],
        "has_night": False,
    },
}

# ── 班次颜色 & 时间 ──────────────────────────────────────
SHIFT_COLORS = {
    "D": "#4CAF50", "D1": "#66BB6A", "D2": "#81C784", "D3": "#A5D6A7",
    "D4": "#43A047", "D5": "#388E3C", "D6": "#2E7D32",
    "C": "#4CAF50", "C1": "#66BB6A", "L": "#8BC34A",
    "H1": "#29B6F6", "H2": "#4FC3F7", "H3": "#81D4FA", "T": "#B3E5FC",
    "N": "#1a73e8", "N2": "#1565C0", "N3": "#0D47A1",
    "L/N": "#FF9800", "OnCall": "#9E9E9E", "off": "#F5F5F5",
}

SHIFT_TIMES = {
    "D": "08:30-17:30", "D1": "08:30-17:00", "D2": "09:00-17:30",
    "D3": "09:30-18:00", "D4": "09:00-18:00", "D5": "08:30-18:00",
    "D6": "07:30-15:30", "C": "07:40-16:10", "C1": "08:00-16:30",
    "L": "08:00-20:00", "H1": "07:40-11:40", "H2": "08:30-12:30",
    "H3": "13:30-17:30", "T": "08:00-12:00",
    "N": "17:30-08:00", "N2": "17:30-07:30", "N3": "18:00-08:00",
    "L/N": "08:00-08:00",
}


def _find_date_columns(headers: list[str]) -> int:
    """找到第一个日期列的位置（0-based index）"""
    date_pattern = re.compile(r"\d{1,2}月\d{1,2}日")
    for i, h in enumerate(headers):
        if h and date_pattern.search(str(h)):
            return i
    return -1


def _parse_date_header(raw: str) -> str:
    """标准化日期格式: '06月01日' → '06月01日'"""
    return raw.strip()


def _parse_shift_cell(raw_value) -> dict:
    """
    解析排班单元格。
    输入: "D4\n09:00-18:00\n[20%]" 或 None
    输出: {"shift": "D4", "time": "09:00-18:00", "category": "20%"}
    """
    if raw_value is None:
        return {"shift": "", "time": "", "category": ""}
    text = str(raw_value).strip()
    if not text:
        return {"shift": "", "time": "", "category": ""}

    lines = text.split("\n")
    shift = lines[0].strip() if len(lines) > 0 else ""
    time_range = lines[1].strip() if len(lines) > 1 else ""
    # 解析分类标签 [80%], [20%], [备班], [L/N]
    cat_raw = lines[2].strip() if len(lines) > 2 else ""
    category = cat_raw.strip("[]") if cat_raw.startswith("[") else ""

    return {"shift": shift, "time": time_range, "category": category}


def _float_or(val, default=0.0) -> float:
    """安全转换 float"""
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _is_schedule_sheet(headers: list[str]) -> bool:
    """通过表头内容判断是否为排班 sheet（而非需求/统计 sheet）"""
    if len(headers) < 15:
        return False  # 统计汇总 sheet 通常不到 10 列
    # 需求 sheet：第一列是"日期"，其余是 00:00~23:00 时间列
    time_count = sum(1 for h in headers if re.match(r"^\d{2}:\d{2}$", str(h).strip()))
    if time_count > 3:
        return False
    # 排班 sheet：包含日期列（如 "06月01日"）
    date_pattern = re.compile(r"\d{1,2}月\d{1,2}日")
    date_cols = sum(1 for h in headers if date_pattern.search(str(h)))
    return date_cols >= 5  # 至少有 5 个日期列


def parse_schedule_xlsx(xlsx_path: str) -> dict:
    """
    解析排班 Excel，返回与现有仪表盘兼容的 JSON 结构。
    """
    wb = openpyxl.load_workbook(xlsx_path)

    # ── 1. 按内容自动识别 sheet 类型 ──────────────────
    schedule_sheets = []  # [(sheet_name, headers, rows)]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        headers = [str(c) if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        if _is_schedule_sheet(headers):
            schedule_sheets.append(sheet_name)

    # 按角色顺序分配：放射医生、放射技师、B超医生
    role_order = ["放射医生", "放射技师", "B超医生"]
    sheet_map = {}
    for i, sname in enumerate(schedule_sheets):
        if i < len(role_order):
            sheet_map[role_order[i]] = sname
            print(f"  [INFO] {role_order[i]} ← sheet '{sname}'")

    result = {
        "month": "",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "roles": {},
        "statistics": {
            "total_fulltime": 0,
            "total_backup": 0,
            "total_backup_hours": 0.0,
            "avg_fulltime_hours": 0.0,
            "total_80_hours": 0.0,
            "total_20_hours": 0.0,
            "total_ln_hours": 0.0,
        },
    }

    fulltime_hours = []

    for role_key, cfg in ROLE_CONFIG.items():
        if role_key not in sheet_map:
            print(f"  [WARN] 未找到 {role_key} 的排班 sheet，跳过")
            continue

        ws = wb[sheet_map[role_key]]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        if len(rows) < 2:
            print(f"  [WARN] {role_key} sheet 数据不足，跳过")
            continue

        # ── 解析表头 ─────────────────────────────────────
        headers = [str(c) if c is not None else "" for c in rows[0]]
        date_start = _find_date_columns(headers)
        if date_start < 0:
            print(f"  [WARN] {role_key} sheet 未找到日期列，跳过")
            continue

        # 统计列在日期列之前
        stat_headers = headers[:date_start]
        date_headers_raw = headers[date_start:]
        dates = [_parse_date_header(d) for d in date_headers_raw]

        # 设置月份
        if not result["month"] and dates:
            result["month"] = f"{dates[0]} ~ {dates[-1]}"

        # ── 解析人员行 ──────────────────────────────────
        staff_list = []
        for row in rows[1:]:
            if row[0] is None:
                continue
            name = str(row[0]).strip()
            if not name or name == "人员":
                continue

            # 读取统计值（根据列位置灵活适配）
            stat_values = list(row[:date_start])
            date_values = list(row[date_start:])

            # 构建 schedule dict
            schedule = {}
            categories = {}
            for di, dv in enumerate(date_values):
                if di >= len(dates):
                    break
                parsed = _parse_shift_cell(dv)
                schedule[dates[di]] = parsed["shift"]
                categories[dates[di]] = parsed["category"]

            # 计算工时
            # 简单估算：从统计列读取
            hours_80 = _float_or(stat_values[1]) if len(stat_values) > 1 else 0
            hours_20 = _float_or(stat_values[2]) if len(stat_values) > 2 else 0
            hours_backup = _float_or(stat_values[3]) if len(stat_values) > 3 else 0
            hours_ln = _float_or(stat_values[4]) if len(stat_values) > 4 else 0
            total_hours = _float_or(stat_values[5]) if len(stat_values) > 5 else (
                hours_80 + hours_20 + hours_backup + hours_ln
            )
            target = _float_or(stat_values[6]) if len(stat_values) > 6 else 176.0

            # 判断是否为备班人员
            is_backup = "备班" in name

            staff_list.append({
                "name": name,
                "internal_name": name.replace(" ", "_"),
                "hours": total_hours,
                "target": target,
                "hours_80": hours_80,
                "hours_20": hours_20,
                "hours_backup": hours_backup,
                "hours_ln": hours_ln,
                "is_backup": is_backup,
                "oncall_count": 0,
                "schedule": schedule,
                "category": categories,
            })

            if not is_backup:
                fulltime_hours.append(total_hours)
                result["statistics"]["total_fulltime"] += 1
            else:
                result["statistics"]["total_backup"] += 1
                result["statistics"]["total_backup_hours"] += total_hours

            result["statistics"]["total_80_hours"] += hours_80
            result["statistics"]["total_20_hours"] += hours_20
            result["statistics"]["total_ln_hours"] += hours_ln

        # 计算平均值
        if fulltime_hours:
            result["statistics"]["avg_fulltime_hours"] = round(
                sum(fulltime_hours) / len(fulltime_hours), 1
            )

        result["roles"][role_key] = {
            "staff": staff_list,
            "dates": dates,
            "shift_colors": SHIFT_COLORS,
            "shift_times": SHIFT_TIMES,
            "total_days": len(dates),
        }

    wb.close()
    return result


def build_deploy_package(xlsx_path: str | None = None):
    """构建完整部署包"""
    # ── 1. 找到排班 Excel ──────────────────────────────
    if xlsx_path is None:
        import glob as _glob
        # 优先在 pipeline_output/schedule/ 中查找
        schedule_dir = os.path.join(SCRIPT_DIR, "pipeline_output", "schedule")
        xlsx_files = sorted(_glob.glob(os.path.join(schedule_dir, "Schedule_*_V3.xlsx")))
        if not xlsx_files:
            # fallback: 项目根目录
            xlsx_files = sorted(_glob.glob(os.path.join(SCRIPT_DIR, "Schedule_*_V3.xlsx")))
        if not xlsx_files:
            print("[ERROR] 未找到排班 Excel 文件")
            print(f"  搜索路径: {schedule_dir}")
            print("  请先运行 schedule.py 生成排班表")
            sys.exit(1)
        xlsx_path = xlsx_files[-1]
        print(f"[INFO] 自动选择最新排班文件: {os.path.basename(xlsx_path)}")

    if not os.path.exists(xlsx_path):
        print(f"[ERROR] 文件不存在: {xlsx_path}")
        sys.exit(1)

    # ── 2. 解析 ────────────────────────────────────────
    print(f"[INFO] 解析排班数据...")
    data = parse_schedule_xlsx(xlsx_path)
    print(f"  月份: {data['month']}")
    for role_key, role_data in data["roles"].items():
        staff_count = len([s for s in role_data["staff"] if not s["is_backup"]])
        print(f"  {role_key}: {staff_count} 人, {len(role_data['dates'])} 天")

    # ── 3. 准备输出目录 ────────────────────────────────
    os.makedirs(DEPLOY_DIR, exist_ok=True)

    # ── 4. 写入 schedule_data.json ─────────────────────
    json_path = os.path.join(DEPLOY_DIR, "schedule_data.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"[OK] schedule_data.json ({os.path.getsize(json_path) / 1024:.0f} KB)")

    # ── 5. 复制静态文件 ────────────────────────────────
    # server.py 和 启动.bat 从模板目录复制
    template_dir = os.path.join(SCRIPT_DIR, "deploy_package")
    # 这些文件将在后续创建

    print(f"\n[INFO] 部署包基础文件已生成: {DEPLOY_DIR}")
    print(f"  还需创建以下文件完成部署包:")
    print(f"    - server.py")
    print(f"    - 排班仪表盘.html")
    print(f"    - 启动.bat")
    print(f"    - .env.example")
    return data


def main():
    parser = argparse.ArgumentParser(description="生成排班部署包")
    parser.add_argument("--xlsx", type=str, default=None, help="排班 Excel 文件路径")
    parser.add_argument("--output", type=str, default=None, help="输出目录（默认 deploy_package/）")
    args = parser.parse_args()

    if args.output:
        global DEPLOY_DIR
        DEPLOY_DIR = args.output

    build_deploy_package(args.xlsx)
    print("\n[DONE] 部署包基础生成完成")


if __name__ == "__main__":
    main()
